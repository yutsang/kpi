#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
align_to_header.py — 對齊 source_1 全部檔到 表頭.xlsx 標準格式（34 欄）。

背景／結構（合併感知，唔係一行一項目）：
  一個項目佔幾行（行數浮動）；左半(B–F)逐行 key-value 明細；右半(是否該司局範疇→尾)
  每欄垂直合併蓋住成個項目 span = 一個項目一個值。項目邊界 = 「是否該司局範疇的項目」
  欄嘅垂直合併範圍。欄位置逐檔浮動 → 一律靠 row6 子表頭 + row5 分組表頭認欄。

映射（group-aware，因第一輪/第二輪子表頭重複，用 (分組,子表頭) 複合鍵）：見 TARGET_SCHEMA。
  尾段：調整原因 = 各非零調整類型「一格列舉」`1、{類型}：{|金額|}萬澳門元`＼n；
        建議調整金額←潛在調整合計(連正負)、建議調整後金額←調整後投資金額、
        建議接納之調整後金額←跨司工作組確認投資金額；
        該關注事項涉及調整金額←source_2 多 concern 加總（numeric）or |潛在調整合計|（fallback）。
  source_2 per-範疇檔按 承批公司+範疇+項目序號 覆蓋審查欄（--with-overlay，source_2 為準）。

輸出：<root>\\_aligned\\ 鏡像 source_1 樹（**永不寫入 source 樹**）。ss/ 同 附件 = 原檔照抄/best-effort。

Windows 跑：
    set PYTHONIOENCODING=utf-8
    # 1) 先驗一個檔（淨吐文字：逐項目抽咗乜 + 會寫入表頭邊欄；唔寫 xlsx）
    python scripts\\adhoc\\align_to_header.py --root ad-hoc\\workspace ^
        --only "source_1\\旅遊局\\SJM-投資計劃執行情況表二（旅遊局）.xlsx" --preview
    # 2) 同一個檔真正寫低（開嚟眼睇）
    python scripts\\adhoc\\align_to_header.py --root ad-hoc\\workspace ^
        --only "source_1\\旅遊局\\SJM-投資計劃執行情況表二（旅遊局）.xlsx" --with-overlay
    # 3) 冇問題先批全部
    python scripts\\adhoc\\align_to_header.py --root ad-hoc\\workspace --all --with-overlay
"""
from __future__ import annotations

import argparse
import io
import re
import shutil
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Font, Side
try:
    from openpyxl.cell.rich_text import CellRichText as _CellRichText, TextBlock as _TextBlock
    from openpyxl.cell.text import InlineFont as _InlineFont
except ImportError:
    _CellRichText = _TextBlock = _InlineFont = None  # openpyxl < 3.1 冇 CellRichText

PASSWORD = "dicj_kpmg"
EXCEL_EXT = {".xlsx", ".xlsm", ".xls"}
HDR_SCAN = 12                       # 掃頭幾行揾子表頭行
GATE_NOADJ = True                   # 冇調整嘅項目 → 建議調整欄留空（唔填 0）；--fill-zero-adj 關掉
ENCRYPT_OUT = True                  # 輸出用 dicj_kpmg 重新加密（保留 source 密碼）；--no-encrypt 關掉
INSERT_XSGZ_FB_COL = True           # #1：輸出喺『承批公司的反饋意見』後插空白『跨司工作組的反饋意見』欄；--no-extra-col 關掉
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_WRAP_TOP = Alignment(wrap_text=True, vertical="top", horizontal="left")


def _grab_style(cell) -> dict:
    """抄實際 style 物件（唔抄 _style index）—— 跨 workbook 先唔會 index 越界。"""
    return {"font": copy(cell.font), "fill": copy(cell.fill), "border": copy(cell.border),
            "alignment": copy(cell.alignment), "nf": cell.number_format,
            "protection": copy(cell.protection)}


def _apply_style(cell, sty: dict) -> None:
    cell.font = sty["font"]
    cell.fill = sty["fill"]
    cell.border = sty["border"]
    cell.alignment = sty["alignment"]
    cell.number_format = sty["nf"]
    cell.protection = sty["protection"]


def _apply_text_style(dst, src_cell) -> None:
    """跟 source 嘅文字外觀（字體/顏色/底色/數字格式）；框由我哋統一格線負責、
    對齊保留 wrap+top 但借 source 嘅水平對齊。
    Rich text（CellRichText）唔 copy cell-level font：source template 嘅 default font
    本身係 bold，copy 會令所有 rich cell 變全 bold。Run-level <b/> 係唯一可靠 bold 指標。"""
    is_rich = _CellRichText is not None and isinstance(dst.value, _CellRichText)
    if not is_rich:
        dst.font = copy(src_cell.font)
    else:
        # _rich_val 已把每個 run 嘅 bold 明確化（含用 source bold 填「繼承」run），
        # 所以 cell-level bold 清零即可——run-level 完全控制外觀。
        _f = copy(dst.font)
        _f.bold = False
        dst.font = _f
    dst.fill = copy(src_cell.fill)
    dst.number_format = src_cell.number_format
    a = src_cell.alignment
    dst.alignment = Alignment(wrap_text=True, vertical="top",
                              horizontal=(a.horizontal if a and a.horizontal else "left"))


def _col_letters_to_idx(letters: str) -> int:
    idx = 0
    for ch in letters:
        idx = idx * 26 + (ord(ch) - 64)
    return idx


def _make_rich_lookup(src) -> tuple:
    """返 (index_runs, si_cell_map)：
      index_runs  = {si_index: [run_spec, ...]}  （只含有多 run/有格式嘅 shared string）
      si_cell_map = {sheet_name: {(row, col): si_index}}  （t="s" cell → 佢用邊個 si）
    run_spec = (text, b, i, sz, rFont)，b/i ∈ {True, False, None}，None = run 冇明確 <b>/<i>
    （繼承 cell-level，_rich_val 會用 source cell 自己嘅 bold 去填）。

    改用 shared-string INDEX 做 key（唔用純文字）——因為同一段純文字可以喺唔同 si
    有唔同粗體 pattern，用文字做 key 會撞、攞錯 pattern（會展支出果格就係咁錯判）。"""
    if _CellRichText is None or _TextBlock is None or _InlineFont is None:
        return {}, {}
    import zipfile as _zf, xml.etree.ElementTree as _ET, re as _re
    NS   = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
    NSr  = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
    NSpk = 'http://schemas.openxmlformats.org/package/2006/relationships'
    try:
        if isinstance(src, io.BytesIO):
            src.seek(0)
        with _zf.ZipFile(src, 'r') as zf:
            names = set(zf.namelist())
            if 'xl/sharedStrings.xml' not in names:
                return {}, {}
            sst_root = _ET.fromstring(zf.read('xl/sharedStrings.xml'))
            wb_root  = _ET.fromstring(zf.read('xl/workbook.xml')) if 'xl/workbook.xml' in names else None
            rels_root = (_ET.fromstring(zf.read('xl/_rels/workbook.xml.rels'))
                         if 'xl/_rels/workbook.xml.rels' in names else None)
            # sheet name -> sheet xml bytes
            sheet_xmls: dict = {}
            if wb_root is not None and rels_root is not None:
                rid_target = {rel.get('Id'): rel.get('Target', '')
                              for rel in rels_root.findall(f'{{{NSpk}}}Relationship')}
                for sh in wb_root.findall(f'.//{{{NS}}}sheet'):
                    nm  = sh.get('name', '')
                    rid = sh.get(f'{{{NSr}}}id', '')
                    tgt = rid_target.get(rid, '')
                    path = tgt.lstrip('/') if tgt.startswith('/') else 'xl/' + tgt
                    if path in names:
                        sheet_xmls[nm] = zf.read(path)
    except Exception:
        return {}, {}

    # 1) index_runs：逐個 si（保留位置 index）
    index_runs: dict = {}
    for idx, si in enumerate(sst_root.findall(f'{{{NS}}}si')):
        rs = si.findall(f'{{{NS}}}r')
        if not rs:
            continue
        runs, any_fmt = [], False
        for r in rs:
            t_el = r.find(f'{{{NS}}}t')
            txt = (t_el.text or '') if t_el is not None else ''
            if not txt:
                continue
            rpr = r.find(f'{{{NS}}}rPr')
            if rpr is None:
                # 純 run（完全冇 rPr）→ b/i = None = 繼承 cell-level font。
                b = i = None
                sz = rFont = None
            else:
                # 有 rPr → 係完整定義：冇 <b> 即係「明確非粗」(False)，唔會繼承
                # cell-level bold。（呢個先係 Excel rich run 嘅真實行為；之前當
                # None＝繼承令「有 rPr 冇 <b>」段錯誤變粗。）
                any_fmt = True
                be = rpr.find(f'{{{NS}}}b')
                b = (be.get('val', '1') != '0') if be is not None else False
                ie = rpr.find(f'{{{NS}}}i')
                i = (ie.get('val', '1') != '0') if ie is not None else False
                sze = rpr.find(f'{{{NS}}}sz')
                sz = None
                if sze is not None:
                    try: sz = float(sze.get('val', 11))
                    except Exception: pass
                fne = rpr.find(f'{{{NS}}}rFont')
                rFont = fne.get('val', 'Calibri') if fne is not None else None
            runs.append((txt, b, i, sz, rFont))
        if any_fmt and runs:
            index_runs[idx] = runs

    # 2) si_cell_map：逐 sheet 掃 t="s" cell → si index
    _cell_re = _re.compile(rb'<c\b[^>]*?/>|<c\b[^>]*?>.*?</c>', _re.DOTALL)
    si_cell_map: dict = {}
    for sn, xb in sheet_xmls.items():
        m: dict = {}
        for cm in _cell_re.finditer(xb):
            cell = cm.group()
            if b't="s"' not in cell:
                continue
            rm = _re.search(rb'\br="([A-Z]+)(\d+)"', cell)
            vm = _re.search(rb'<v>(\d+)</v>', cell)
            if rm and vm:
                col = _col_letters_to_idx(rm.group(1).decode('ascii'))
                row = int(rm.group(2))
                m[(row, col)] = int(vm.group(1))
        si_cell_map[sn] = m
    return index_runs, si_cell_map


def _rich_val(cell, rich):
    """rich = (index_runs, si_cell_map)。用 cell 自己嘅 (sheet,row,col) 揾返佢個 si index，
    再 build CellRichText——唔靠純文字 key，杜絕同段文字唔同粗體嘅撞 key 錯判。
    每段 bold：有 <b/> 用佢自己；冇 <b/>（含純文字段）→ 用 source cell 自己嘅 bold。"""
    if not rich:
        return cell.value
    index_runs, si_cell_map = rich
    if not index_runs:
        return cell.value
    v = cell.value
    if not isinstance(v, str):
        return v
    try:
        sheet = cell.parent.title
        idx = si_cell_map.get(sheet, {}).get((cell.row, cell.column))
    except Exception:
        idx = None
    runs = index_runs.get(idx) if idx is not None else None
    if runs is None:
        return v
    src_bold = bool(cell.font and cell.font.bold)
    parts = []
    for txt, b, i, sz, rFont in runs:
        kw = {"b": src_bold if b is None else b}
        if i is not None:
            kw["i"] = i
        if sz is not None:
            kw["sz"] = sz
        if rFont is not None:
            kw["rFont"] = rFont
        parts.append(_TextBlock(font=_InlineFont(**kw), text=txt))
    return _CellRichText(*parts)


def _inject_sheet_drawings_raw(src_zip_src, src_sheet_name: str, out_path: Path) -> bool:
    """Post-save ZIP-level：把 source sheet 嘅 drawing XML（含 TextBox/shapes/images）
    完整注入 output xlsx，繞過 openpyxl 唔 load shapes 嘅限制。
    Returns True 代表成功注入咗 drawing。"""
    import zipfile
    import xml.etree.ElementTree as ET

    NS_WB  = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
    NS_PKG = 'http://schemas.openxmlformats.org/package/2006/relationships'
    REL_DRW = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing'

    # ── 1. 從 source ZIP 抽 drawing XML ──────────────────────────────────────
    if isinstance(src_zip_src, io.BytesIO):
        src_zip_src.seek(0)

    extra: dict[str, bytes] = {}
    drw_filename: str = ""
    old_drw_rels: str = ""

    try:
        with zipfile.ZipFile(src_zip_src, 'r') as szf:
            nl = set(szf.namelist())
            wb_root = ET.fromstring(szf.read('xl/workbook.xml'))
            src_idx = next(
                (i for i, sh in enumerate(wb_root.findall(f'.//{{{NS_WB}}}sheet'), 1)
                 if sh.get('name') == src_sheet_name),
                None)
            if src_idx is None:
                return False

            rels_p = f'xl/worksheets/_rels/sheet{src_idx}.xml.rels'
            if rels_p not in nl:
                return False
            rroot = ET.fromstring(szf.read(rels_p))
            drw_rel = next(
                (r for r in rroot.findall(f'{{{NS_PKG}}}Relationship')
                 if r.get('Type') == REL_DRW),
                None)
            if drw_rel is None:
                return False

            drw_filename = drw_rel.get('Target', '').split('/')[-1]
            drw_path = f'xl/drawings/{drw_filename}'
            if drw_path not in nl:
                return False

            extra[drw_path] = szf.read(drw_path)

            old_drw_rels = f'xl/drawings/_rels/{drw_filename}.rels'
            if old_drw_rels in nl:
                drw_rels_bytes = szf.read(old_drw_rels)
                extra[old_drw_rels] = drw_rels_bytes
                for rel in ET.fromstring(drw_rels_bytes).findall(f'{{{NS_PKG}}}Relationship'):
                    tgt = rel.get('Target', '')
                    if tgt.startswith('../media/'):
                        mp = 'xl/media/' + tgt.split('/')[-1]
                        if mp in nl:
                            extra[mp] = szf.read(mp)
    except Exception:
        return False

    if not drw_filename:
        return False

    # ── 2. 讀 output ZIP ─────────────────────────────────────────────────────
    try:
        out_files: dict[str, bytes] = {}
        with zipfile.ZipFile(out_path, 'r') as ozf:
            for name in ozf.namelist():
                out_files[name] = ozf.read(name)
    except Exception:
        return False

    # ── 3. 揾 output sheet 的 index ──────────────────────────────────────────
    try:
        owb = ET.fromstring(out_files['xl/workbook.xml'])
        name_try = (src_sheet_name, src_sheet_name[:31])
        out_idx = next(
            (i for i, sh in enumerate(owb.findall(f'.//{{{NS_WB}}}sheet'), 1)
             if sh.get('name') in name_try),
            None)
        if out_idx is None:
            return False
    except Exception:
        return False

    # ── 4. 新 drawing 檔名（避免衝突）──────────────────────────────────────────
    existing = [n for n in out_files if re.match(r'^xl/drawings/drawing\d+\.xml$', n)]
    used = {int(re.search(r'(\d+)', n).group(1)) for n in existing} if existing else {0}
    new_num = max(used) + 1
    new_fn = f'drawing{new_num}.xml'
    new_drw_path = f'xl/drawings/{new_fn}'
    new_rels_path = f'xl/drawings/_rels/{new_fn}.rels'

    out_files[new_drw_path] = extra[f'xl/drawings/{drw_filename}']
    if old_drw_rels in extra:
        out_files[new_rels_path] = extra[old_drw_rels]
    for k, v in extra.items():
        if k.startswith('xl/media/') and k not in out_files:
            out_files[k] = v

    # ── 4b. 更新 [Content_Types].xml（必須，否則 Excel 報 "found a problem"）────
    CT_PATH = '[Content_Types].xml'
    DRW_CT  = 'application/vnd.openxmlformats-officedocument.drawing+xml'
    _MEDIA_CT = {
        'png': 'image/png', 'jpg': 'image/jpeg', 'jpeg': 'image/jpeg',
        'gif': 'image/gif', 'bmp': 'image/bmp', 'tiff': 'image/tiff',
        'emf': 'image/x-emf', 'wmf': 'image/x-wmf', 'svg': 'image/svg+xml',
    }
    if CT_PATH in out_files:
        ct = out_files[CT_PATH].decode('utf-8')
        # 加 drawing Override（如未登記）
        if new_fn not in ct:
            ct = ct.replace(
                '</Types>',
                f'<Override PartName="/xl/drawings/{new_fn}" ContentType="{DRW_CT}"/></Types>')
        # 加 media Default（如未登記）
        for mp in (k for k in extra if k.startswith('xl/media/')):
            ext = mp.rsplit('.', 1)[-1].lower()
            mime = _MEDIA_CT.get(ext)
            if mime and f'Extension="{ext}"' not in ct:
                ct = ct.replace(
                    '</Types>',
                    f'<Default Extension="{ext}" ContentType="{mime}"/></Types>')
        out_files[CT_PATH] = ct.encode('utf-8')

    # ── 5. 加 drawing relationship 入 output sheet rels ──────────────────────
    sheet_rels_p = f'xl/worksheets/_rels/sheet{out_idx}.xml.rels'
    rel_entry = (f'<Relationship Id="rId_drw{new_num}" Type="{REL_DRW}" '
                 f'Target="../drawings/{new_fn}"/>')
    if sheet_rels_p in out_files:
        s = out_files[sheet_rels_p].decode('utf-8')
        s = s.replace('</Relationships>', rel_entry + '</Relationships>')
        out_files[sheet_rels_p] = s.encode('utf-8')
    else:
        ns = 'http://schemas.openxmlformats.org/package/2006/relationships'
        out_files[sheet_rels_p] = (
            f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            f'<Relationships xmlns="{ns}">{rel_entry}</Relationships>'
        ).encode('utf-8')

    # ── 5b. 喺 sheet XML 加返 <drawing r:id> 元素 ─────────────────────────────
    # 冇呢個 tag，Excel 唔會 render 個 drawing（印章貼咗都唔會出現）——舊版漏咗。
    sheet_xml_p = f'xl/worksheets/sheet{out_idx}.xml'
    if sheet_xml_p in out_files:
        sx = out_files[sheet_xml_p].decode('utf-8')
        if '<drawing ' not in sx and '<drawing/>' not in sx:
            # 保證 root 有宣告 xmlns:r，否則 <drawing r:id> 會係未宣告 prefix
            R_NS = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
            if 'xmlns:r=' not in sx.split('>', 1)[0]:
                sx = re.sub(r'(<worksheet\b)', rf'\1 xmlns:r="{R_NS}"', sx, count=1)
            drawing_el = f'<drawing r:id="rId_drw{new_num}"/>'
            # <drawing> 喺 schema 尾段；要喺呢啲後續元素之前插（如存在），否則插喺尾
            trailing = ('<legacyDrawing', '<legacyDrawingHF', '<drawingHF', '<picture',
                        '<oleObjects', '<controls', '<webPublishItems', '<tableParts',
                        '<extLst')
            pos = -1
            for a in trailing:
                i = sx.find(a)
                if i != -1 and (pos == -1 or i < pos):
                    pos = i
            if pos != -1:
                sx = sx[:pos] + drawing_el + sx[pos:]
            else:
                sx = sx.replace('</worksheet>', drawing_el + '</worksheet>')
            out_files[sheet_xml_p] = sx.encode('utf-8')

    # ── 6. 寫返 output ────────────────────────────────────────────────────────
    try:
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, 'w', compression=zipfile.ZIP_DEFLATED) as ozf:
            for name, data in out_files.items():
                ozf.writestr(name, data)
        out_path.write_bytes(buf.getvalue())
        return True
    except Exception:
        return False


def base_grid(cell, data_style: dict) -> None:
    """每個 data cell 打底：template data style + 完整四邊框 + wrap-top。"""
    _apply_style(cell, data_style)
    cell.border = _BORDER
    cell.alignment = _WRAP_TOP


def copy_full_cell(dst, src_cell) -> None:
    """值 + 全 style 照抄（跨 workbook 安全：抄 style 物件唔抄 index）。"""
    if src_cell.value is not None:
        dst.value = src_cell.value
    _apply_style(dst, _grab_style(src_cell))


def encrypt_file(plain: Path, enc: Path, log) -> bool:
    """用 dicj_kpmg 重新加密（保留 source 密碼保護）。msoffcrypto-tool CLI -e。"""
    import subprocess
    import sys
    base = [str(plain), str(enc), "-e", "-p", PASSWORD]
    for cmd in ([sys.executable, "-m", "msoffcrypto"] + base,
                ["msoffcrypto-tool"] + base):
        try:
            r = subprocess.run(cmd, capture_output=True, text=True)
        except FileNotFoundError:
            continue
        if r.returncode == 0 and enc.exists() and enc.stat().st_size > 0:
            return True
    log("  ⚠ 加密失敗（msoffcrypto-tool -e 唔 work）→ 輸出未加密；喺 Windows： pip install -U msoffcrypto-tool")
    return False


# ── 正規化 ────────────────────────────────────────────────────────────────
_PUNC = str.maketrans({
    "（": "(", "）": ")", "：": ":", "，": ",", "、": ",", "／": "/",
    "　": "", " ": "", " ": "", "\"": "", "“": "", "”": "", "'": "",
})


def _s(x) -> str:
    return "" if x is None else str(x).strip().replace("\r", "")


def nkey(x) -> str:
    """認欄用嘅正規化鍵：去空白/引號、全形標點→半形。"""
    return _s(x).replace("\n", "").translate(_PUNC)


# 子表頭同義變體（source 名 → 表頭 canonical 名 nkey）
VARIANTS = {
    nkey("畢馬威的分析"): nkey("KPMG分析"),   # 只喺第一/二輪組；潛在調整組另有規則
}
# 任何司局專名「XX的回覆／XX整體的回覆」（跨司工作組/博監局/經濟局/社文司/…）→ 表頭「跨司工作組的回覆」
_RE_REPLY = re.compile(r".+的回覆$")
# 「KPMG希望進一步向XX瞭解的事項」（XX = 跨司工作組/博監局/…，各範疇專名）→ 表頭「…向跨司工作組瞭解的事項」
_RE_ASK = re.compile(r"^KPMG希望進一步向.+瞭解的事項$")


def canon_sub(raw: str) -> str:
    k = nkey(raw)
    k = re.sub(r"\([^()]*\)$", "", k)          # 去尾部括號註（「(萬澳門元)」「(如不一致…)」）
    if k in VARIANTS:
        return VARIANTS[k]
    if _RE_REPLY.match(k):                      # 收晒各司局專名「…的回覆」
        return nkey("跨司工作組的回覆")
    if _RE_ASK.match(k):                        # 收晒「KPMG希望進一步向{司局}瞭解的事項」
        return nkey("KPMG希望進一步向跨司工作組瞭解的事項")
    return k


# ── 表頭目標 schema（34 欄；idx = 表頭資料欄位置；rule 見 resolve）──────────
# group: row5 分組（"" = 無/左半/單欄）；sub: row6 子表頭；rule: 點填
G1 = "與跨司工作組的第一輪問題諮詢"
G2 = "與跨司工作組的第二輪問題諮詢"
GT = "與承批公司溝通潛在調整事項"          # 用前綴 match（表頭全名長）
LEFT = "__LEFT__"

TARGET_SCHEMA = [
    # 左半（positional，右對齊照抄）——只標記，唔逐欄配名
    {"grp": "", "sub": "是否該司局範疇的項目", "rule": ("copy_any", "是否該司局範疇的項目")},
    # 第一輪
    {"grp": G1, "sub": "是否有希望諮詢的問題", "rule": ("copy", G1, "是否有希望諮詢的問題")},
    {"grp": G1, "sub": "問題狀態", "rule": ("copy", G1, "問題狀態")},
    {"grp": G1, "sub": "KPMG提出日期", "rule": ("copy", G1, "KPMG提出日期")},
    {"grp": G1, "sub": "KPMG分析", "rule": ("copy", G1, "KPMG分析")},
    {"grp": G1, "sub": "承批公司管理層解釋", "rule": ("copy", G1, "承批公司管理層解釋")},
    {"grp": G1, "sub": "KPMG希望進一步向跨司工作組瞭解的事項", "rule": ("copy", G1, "KPMG希望進一步向跨司工作組瞭解的事項")},
    {"grp": G1, "sub": "跨司工作組的回覆", "rule": ("copy", G1, "跨司工作組的回覆")},
    # 第二輪
    {"grp": G2, "sub": "問題狀態", "rule": ("copy", G2, "問題狀態")},
    {"grp": G2, "sub": "KPMG提出日期", "rule": ("copy", G2, "KPMG提出日期")},
    {"grp": G2, "sub": "KPMG分析", "rule": ("copy", G2, "KPMG分析")},
    {"grp": G2, "sub": "承批公司管理層解釋", "rule": ("copy", G2, "承批公司管理層解釋")},
    {"grp": G2, "sub": "KPMG希望進一步向跨司工作組瞭解的事項", "rule": ("copy", G2, "KPMG希望進一步向跨司工作組瞭解的事項")},
    {"grp": G2, "sub": "跨司工作組的回覆", "rule": ("copy", G2, "跨司工作組的回覆")},
    # 與承批公司溝通潛在調整事項（source_1 只有前 2 條 + 跨司反饋/畢馬威分析）
    {"grp": GT, "sub": "畢馬威關注事項", "rule": ("copy", GT, "畢馬威關注事項")},
    {"grp": GT, "sub": "承批公司的反饋意見", "rule": ("copy", GT, "承批公司的反饋意見")},
    {"grp": GT, "sub": "是否需進一步與跨司工作組溝通", "rule": ("yn_followup",)},          # 預設否；source_2=是
    {"grp": GT, "sub": "需溝通關注事項", "rule": ("blank",)},                             # source_2
    {"grp": GT, "sub": "該關注事項涉及調整金額", "rule": ("abs_total",)},                  # |潛在調整合計|
    {"grp": GT, "sub": "跨司工作組主責部門針對該關注事項已給的反饋意見", "rule": ("copy", GT, "跨司工作組的反饋意見")},
    {"grp": GT, "sub": "KPMG需與跨司工作組進一步確認的問題", "rule": ("copy", GT, "畢馬威的分析")},   # 假設
    {"grp": GT, "sub": "跨司工作組最新反饋意見", "rule": ("blank",)},                      # source_2
    # 畢馬威審查意見
    {"grp": "畢馬威審查意見", "sub": "建議調整金額", "rule": ("seed", "潛在調整合計")},
    {"grp": "畢馬威審查意見", "sub": "調整原因", "rule": ("enum",)},
    {"grp": "畢馬威審查意見", "sub": "建議調整後金額", "rule": ("seed", "調整後投資金額")},
    # 跨司工作組審閱意見
    {"grp": "跨司工作組審閱意見", "sub": "項目分析意見", "rule": ("blank",)},
    {"grp": "跨司工作組審閱意見", "sub": "建議接納之調整後金額", "rule": ("seed", "跨司工作組確認投資金額")},
    {"grp": "", "sub": "對比畢馬威審查意見與跨司工作組審閱意見是否一致", "rule": ("blank",)},
]

# 潛在調整組內「唔係調整類型」嘅子表頭（其餘全部當調整類型入列舉）
ADJ_NON_TYPE = {nkey(x) for x in
                ["申報投資金額", "潛在調整合計", "調整後投資金額", "跨司工作組確認投資金額"]}
# source_2 覆蓋會蓋嘅表頭欄（sub nkey）
OVERLAY_SUBS = [nkey(x) for x in
                ["是否需進一步與跨司工作組溝通", "需溝通關注事項", "該關注事項涉及調整金額",
                 "跨司工作組主責部門針對該關注事項已給的反饋意見",
                 "KPMG需與跨司工作組進一步確認的問題", "跨司工作組最新反饋意見",
                 "畢馬威關注事項", "承批公司的反饋意見"]]

# 唔按項目合併嘅欄（右半唔 merge 落 span，改逐行填同一個值）
# 只剩 W(23)=是否需進一步與跨司工作組溝通；AA(KPMG需與跨司)已改回按項目合併
NO_MERGE_SUBS = {nkey("是否需進一步與跨司工作組溝通")}


# ── I/O ──────────────────────────────────────────────────────────────────
def load_wb(path: Path):
    """非 read-only（要 merged_cells）；加密就 dicj_kpmg 解。
    load 後立即 build rich-text lookup（zip 仍開，archive 未關），cache 到 wb._kpi_rich_table。
    wb._kpi_src_zip_src 儲 source ZIP（Path 或解密 BytesIO），供 post-save drawing injection 用。"""
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        wb._kpi_rich_table = _make_rich_lookup(path)
        wb._kpi_src_zip_src = path
        return wb
    except Exception:
        import msoffcrypto
        buf = io.BytesIO()
        with open(path, "rb") as f:
            off = msoffcrypto.OfficeFile(f)
            off.load_key(password=PASSWORD)
            off.decrypt(buf)
        buf.seek(0)
        wb = openpyxl.load_workbook(buf, data_only=True)
        wb._kpi_rich_table = _make_rich_lookup(buf)
        buf.seek(0)
        wb._kpi_src_zip_src = buf
        return wb


def num(v):
    """試 parse 成數字（去逗號）；唔係數字回 None。"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = _s(v).replace(",", "").replace("，", "")
    if re.fullmatch(r"-?\d+(\.\d+)?", s):
        return float(s)
    return None


def fmt_amt(x: float) -> str:
    return str(int(x)) if abs(x - round(x)) < 1e-9 else f"{x:.2f}".rstrip("0").rstrip(".")


# ── 表頭解析 ──────────────────────────────────────────────────────────────
def group_map(ws, grow: int, maxcol: int) -> dict[int, str]:
    """row5 分組表頭：橫向合併 fill-forward → {col: group_nkey}。"""
    gm: dict[int, str] = {}
    merged = {}
    for m in ws.merged_cells.ranges:
        if m.min_row <= grow <= m.max_row and m.max_col > m.min_col:
            v = _s(ws.cell(m.min_row, m.min_col).value)
            for c in range(m.min_col, m.max_col + 1):
                merged[c] = v
    for c in range(1, maxcol + 1):
        gm[c] = nkey(merged.get(c, _s(ws.cell(grow, c).value)))
    return gm


def detect_sub_row(ws, maxrow: int, maxcol: int) -> int:
    """揀同已知子表頭重疊最多嗰行做 row6（子表頭）。"""
    known = {nkey(r["sub"]) for r in TARGET_SCHEMA} | {nkey("非博彩投資項目性質")}
    best, best_hit = 1, -1
    for r in range(1, min(HDR_SCAN, maxrow) + 1):
        vals = {canon_sub(ws.cell(r, c).value) for c in range(1, maxcol + 1)}
        hit = len(known & vals)
        if hit > best_hit:
            best_hit, best = hit, r
    return best


def find_anchor_col(ws, subrow: int, maxcol: int) -> int | None:
    tgt = nkey("是否該司局範疇的項目")
    for c in range(1, maxcol + 1):
        if canon_sub(ws.cell(subrow, c).value) == tgt:
            return c
    return None


# ── 抽取一個 sheet 嘅項目 ─────────────────────────────────────────────────
class Project:
    __slots__ = ("r0", "r1", "left", "by_gs", "adj", "seed", "seq", "override")

    def __init__(self, r0, r1):
        self.r0, self.r1 = r0, r1
        self.left: list[list] = []          # 逐行左半原值 [[cell,...],...]
        self.by_gs: dict[tuple, object] = {}  # (grp_nkey, sub_canon) -> value
        self.adj: list[tuple[str, float]] = []  # [(類型顯示名, 金額), ...] 非零
        self.seed: dict[str, object] = {}   # 潛在調整合計/調整後投資金額/跨司工作組確認投資金額
        self.seq: str = ""                  # 項目序號名稱（join key，如 "76-吸引外國客源…"）
        self.override: dict[str, object] = {}  # source_2 覆蓋：canon_sub -> value（為準，蓋過計算值）

    def has_adj(self) -> bool:
        """有冇實際調整：有非零列舉類型、或可推得非零調整總額。
        期後頁常見冇逐項拆解、亦冇填『潛在調整合計』，只填 調整後投資金額＋申報投資金額
        → 要用 _adj_total（同 build_enum 一致）先開得閘，否則 enum 永遠被 gate 走（dead code）。"""
        if self.adj:
            return True
        t = _adj_total(self)
        return t is not None and abs(t) > 1e-9


def merged_val(ws, c: int, r0: int, r1: int):
    for r in range(r0, r1 + 1):
        v = ws.cell(r, c).value
        if _s(v) != "":
            return v
    return None


def project_spans(ws, anchor: int, data0: int, maxrow: int) -> list[tuple[int, int]]:
    """用 anchor 欄垂直合併定項目邊界；未合併但有值 = 單行項目。"""
    spans = []
    covered = set()
    for m in ws.merged_cells.ranges:
        if m.min_col <= anchor <= m.max_col and m.min_row >= data0 and m.max_row > m.min_row:
            spans.append((m.min_row, m.max_row))
            covered.update(range(m.min_row, m.max_row + 1))
    for r in range(data0, maxrow + 1):
        if r in covered:
            continue
        if _s(ws.cell(r, anchor).value) != "":
            spans.append((r, r))
    return sorted(spans)


_RE_ITEM_LABEL = "投資項目序號及名稱"
_RE_ITEM_LABEL_NK = nkey(_RE_ITEM_LABEL)
# 項目碼：76- / CE001- / OPCE006- / 項目3-（英文碼 0-5 字母 + 數字，或「項目N」）
_RE_SEQ_CODE = re.compile(r"^(?:[A-Za-z]{0,5}\d+|項目\d+)\s*[-–]")


def _derive_seq(first: list[str]) -> str:
    """項目 join key 名：優先攞「投資項目序號及名稱：」label 後面嗰格（VML/Wynn 名喺隔籬格），
    否則攞似項目碼（76-/CE001-/項目3-）嘅格，最後退回次欄。"""
    for i, x in enumerate(first):
        if nkey(x).startswith(_RE_ITEM_LABEL_NK):
            tail = re.sub(r"^.*?投資項目序號及名稱[:：]?\s*", "", x).strip()   # 同格已附名（SJM）
            if tail:
                return tail
            for y in first[i + 1:]:                                          # 名喺 label 右邊格
                if _s(y):
                    return _s(y)
    for x in first:                                                          # 直接似項目碼嘅格
        if _RE_SEQ_CODE.match(x):
            return x
    return first[1] if len(first) > 1 else (first[0] if first else "")


def find_label_col(ws, subrow: int, anchor: int, maxrow: int) -> int | None:
    """揾「投資項目序號及名稱：」標籤欄（左半、每個項目首行都有）。"""
    for r in range(subrow + 1, min(subrow + 80, maxrow) + 1):
        for c in range(1, max(anchor, 2)):
            if nkey(ws.cell(r, c).value).startswith(_RE_ITEM_LABEL):
                return c
    return None


def _row_blank(ws, r: int, maxcol: int) -> bool:
    return all(_s(ws.cell(r, c).value) == "" for c in range(1, maxcol + 1))


def _is_band_row(ws, r: int, maxcol: int) -> bool:
    """中間重複表頭帶：一個新『非博彩投資項目性質範疇』開始時，會重覆一條
    「非博彩投資項目性質 / 金額」+「資料要求」表頭帶（B:C 合併）。呢啲行係範疇分隔，
    唔屬任何項目 → 唔可以併入上一個項目 span、要獨立過帶。"""
    for c in range(1, min(maxcol, 6) + 1):
        k = nkey(ws.cell(r, c).value)
        if k.startswith("非博彩投資項目性質") or k == "資料要求":
            return True
    return False


def spans_by_label(ws, label_col: int, data0: int, maxrow: int, maxcol: int) -> list[tuple[int, int]]:
    """項目邊界 = 標籤行；一個項目由佢標籤行去到下個標籤行前一行（尾部空行 / 範疇表頭帶剪走）。
    唔靠 anchor 合併高度，故唔會漏『佔多過合併高度』嘅項目明細行；但下一個範疇嘅重覆表頭帶
    （非博彩投資項目性質 / 資料要求）唔可以被吞入上一個項目 → 由尾剪走、獨立過帶。"""
    starts = [r for r in range(data0, maxrow + 1)
              if nkey(ws.cell(r, label_col).value).startswith(_RE_ITEM_LABEL)]
    spans = []
    for i, s in enumerate(starts):
        e = starts[i + 1] - 1 if i + 1 < len(starts) else maxrow
        while e > s and (_row_blank(ws, e, maxcol) or _is_band_row(ws, e, maxcol)):
            e -= 1
        spans.append((s, e))
    return spans


def extract(ws, log) -> tuple[list[Project], int, int, dict[int, str], int]:
    maxrow, maxcol = ws.max_row or 0, ws.max_column or 0
    subrow = detect_sub_row(ws, maxrow, maxcol)
    grow = max(1, subrow - 1)
    gm = group_map(ws, grow, maxcol)
    anchor = find_anchor_col(ws, subrow, maxcol)
    if anchor is None:
        return [], subrow, 0, gm, maxcol, {}, maxrow
    # 每欄嘅 (group, sub)
    col_gs: dict[int, tuple] = {}
    for c in range(1, maxcol + 1):
        col_gs[c] = (gm.get(c, ""), canon_sub(ws.cell(subrow, c).value))
    left_cols = [c for c in range(1, anchor)]         # 左半 = anchor 之前
    right_cols = [c for c in range(anchor, maxcol + 1)]
    data0 = subrow + 1
    label_col = find_label_col(ws, subrow, anchor, maxrow)
    if label_col:
        spans = spans_by_label(ws, label_col, data0, maxrow, maxcol)
        am = project_spans(ws, anchor, data0, maxrow)
        taller = sum(1 for (s, e) in spans
                     for (a0, a1) in am if a0 == s and (e - s) > (a1 - a0))
        if taller:
            log(f"      （項目分段用標籤行：{len(spans)} 個；其中 {taller} 個高過 anchor 合併，"
                f"已收回被合併切走嘅明細行）")
    else:
        spans = project_spans(ws, anchor, data0, maxrow)
        log("      （揾唔到「投資項目序號及名稱」標籤欄 → 退回用 anchor 合併分段）")
    projs: list[Project] = []
    for r0, r1 in spans:
        p = Project(r0, r1)
        for r in range(r0, r1 + 1):
            p.left.append([ws.cell(r, c).value for c in left_cols])
        # 序號名稱 join key = 左半第一行第一個似 "數字-名" 或次欄
        first = [ _s(x) for x in (p.left[0] if p.left else []) ]
        p.seq = _derive_seq(first)
        for c in right_cols:
            grp, sub = col_gs[c]
            val = merged_val(ws, c, r0, r1)
            # 潛在調整組：拆調整類型 / seed
            if sub in {nkey("潛在調整合計"), nkey("調整後投資金額"), nkey("跨司工作組確認投資金額"), nkey("申報投資金額")}:
                p.seed[sub] = val
            elif _is_adj_type(grp, sub, ws.cell(subrow, c).value):
                n = num(val)
                if n is not None and abs(n) > 1e-9:
                    p.adj.append((_s(ws.cell(subrow, c).value), n))
            else:
                if val is not None:
                    p.by_gs[(grp, sub)] = val
        projs.append(p)
    if label_col:
        label_starts = {s for s, _ in spans}
        orphan = [a0 for (a0, a1) in am if a0 not in label_starts
                  and not any(s <= a0 <= e for (s, e) in spans)]
        log(f"      項目({len(projs)}): {[p.seq for p in projs]}")
        if orphan:
            log(f"      ⚠ anchor 有合併起點但唔喺任何標籤項目內: rows {orphan} → 核實係咪漏咗真項目")
    return projs, subrow, anchor, gm, maxcol, col_gs, maxrow


def _is_adj_type(grp: str, sub: str, raw: str) -> bool:
    """只認「投資金額的潛在調整事項及調整後金額」拆解組嘅調整類型欄。
    要排除「與承批公司溝通潛在調整事項…」溝通組（組名一樣含『潛在調整』，
    但佢啲欄係文字工作流，唔係金額拆解）。"""
    if "潛在調整" not in grp or "溝通" in grp:
        return False
    return sub not in ADJ_NON_TYPE


def _adj_total(p: Project):
    """項目調整總額：優先『潛在調整合計』；否則『調整後投資金額 − 申報投資金額』。"""
    n = num(p.seed.get(nkey("潛在調整合計")))
    if n is not None:
        return n
    after = num(p.seed.get(nkey("調整後投資金額")))
    before = num(p.seed.get(nkey("申報投資金額")))
    if after is not None and before is not None:
        return after - before
    return None


def build_enum(p: Project) -> str:
    lines = []
    for i, (typ, amt) in enumerate(p.adj, 1):
        lines.append(f"{i}、{typ}：{fmt_amt(abs(amt))}萬澳門元")
    if not lines:                       # 只有調整總數、冇逐項拆解（期後常見）→ 用總額補一條
        t = _adj_total(p)
        if t is not None and abs(t) > 1e-9:
            lines.append(f"1、投資金額調整：{fmt_amt(abs(t))}萬澳門元")
    return "\n".join(lines)


# ── 解析目標欄值 ──────────────────────────────────────────────────────────
def cell_value(tpl, c: int, p: Project):
    """一個表頭右半欄嘅最終值：source_2 覆蓋為準，否則按 rule 計算。"""
    _grp, sub = tpl.col_gs[c]
    if sub in p.override and _s(p.override[sub]) != "":
        return p.override[sub]
    return resolve(tpl.rules[c], p)


def resolve(rule, p: Project):
    kind = rule[0]
    if kind == "blank":
        return None
    if kind == "yn_followup":
        return "否"          # #5 預設「否」；source_2 有覆蓋（喺 cell_value 前置）先變「是」
    if kind == "copy_any":
        sub = canon_sub(rule[1])
        for (g, s), v in p.by_gs.items():
            if s == sub:
                return v
        return None
    if kind == "copy":
        want_sub = canon_sub(rule[2])          # 分組用前綴寬鬆 match、子表頭正規化 match
        for (g, s), v in p.by_gs.items():
            if s == want_sub and _grp_match(g, rule[1]):
                return v
        return None
    if kind == "enum":
        if GATE_NOADJ and not p.has_adj():
            return None
        return build_enum(p) or None
    if kind == "abs_total":
        if GATE_NOADJ and not p.has_adj():
            return None
        t = _adj_total(p)                       # 同 enum 一致：合計 → 否則 調整後−申報
        return abs(t) if t is not None else None
    if kind == "seed":
        # 建議調整金額(←潛在調整合計) / 建議調整後金額(←調整後投資金額) / 建議接納之調整後金額
        # 一律跟 source_1（default source_1；冇調整都照顯示 source_1 數，唔 blank、唔填 0 覆蓋）。
        # source_2 覆蓋唔掂呢幾欄（唔喺 OVERLAY_SUBS）→ 一定係 source_1 為底。
        return p.seed.get(nkey(rule[1]))
    return None


_RE_ROUND1 = re.compile(r"(第一|首|1)輪")
_RE_ROUND2 = re.compile(r"(第二|2)輪")


def _round_of(s: str) -> str:
    """分組名裡面嘅諮詢輪次：第二輪→'2'、第一輪/首輪→'1'、冇→''。用嚟分開兩輪唔好撈亂。"""
    if _RE_ROUND2.search(s):
        return "2"
    if _RE_ROUND1.search(s):
        return "1"
    return ""


def _grp_match(got: str, want: str) -> bool:
    """分組名寬鬆 match：want 係短前綴（如 GT），got 係檔內全名 nkey。
    但『第一輪/第二輪問題諮詢』共用前綴『與跨司工作組』→ 舊 w[:6] 前綴兩輪都會中（#2 撈亂
    第一/第二輪 columns）；故先比輪次：兩邊都有輪次而唔同 → 一定唔 match。"""
    g, w = nkey(got), nkey(want)
    rg, rw = _round_of(g), _round_of(w)
    if rg and rw and rg != rw:
        return False
    return w in g or g in w or g.startswith(w[:6])


# ── 表頭 template（header rows 1..subrow）───────────────────────────────────
class Template:
    def __init__(self, hf: Path, log):
        wb = load_wb(hf)
        ws = wb.active
        self.maxcol = ws.max_column or 0
        self.subrow = detect_sub_row(ws, ws.max_row or 0, self.maxcol)
        self.grow = max(1, self.subrow - 1)
        self.gm = group_map(ws, self.grow, self.maxcol)
        self.anchor = find_anchor_col(ws, self.subrow, self.maxcol)
        # 逐欄 (grp, sub) + 每欄嘅 rule（跟 TARGET_SCHEMA 對）
        self.col_gs = {c: (self.gm.get(c, ""), canon_sub(ws.cell(self.subrow, c).value))
                       for c in range(1, self.maxcol + 1)}
        self.rules = self._attach_rules(ws, log)
        # header 區塊（值+style）、col widths、header 內 merges
        self.hdr_cells = {}
        for r in range(1, self.subrow + 1):
            for c in range(1, self.maxcol + 1):
                cell = ws.cell(r, c)
                self.hdr_cells[(r, c)] = (cell.value, _grab_style(cell))
        self.widths = {get_column_letter(c): (ws.column_dimensions[get_column_letter(c)].width)
                       for c in range(1, self.maxcol + 1)}
        self.row_heights = {r: ws.row_dimensions[r].height for r in range(1, self.subrow + 1)}
        self.hdr_merges = [str(m) for m in ws.merged_cells.ranges if m.max_row <= self.subrow]
        self.data_style = _grab_style(ws.cell(self.subrow, self.maxcol))
        wb.close()

    def _attach_rules(self, ws, log):
        """按 (grp,sub) 幫每個表頭欄搭 TARGET_SCHEMA rule；左半 = LEFT。"""
        rules = {}
        left_cols = [c for c in range(1, (self.anchor or 1))]
        for c in left_cols:
            rules[c] = ("__left__", c)
        # 建 schema lookup by (grp_match, sub)
        for c in range(self.anchor or 1, self.maxcol + 1):
            grp, sub = self.col_gs[c]
            if not sub:                      # row6 空（縱向合併表頭）→ 用 row5 文字做 sub
                sub = canon_sub(ws.cell(self.grow, c).value)
            match = None
            for row in TARGET_SCHEMA:
                if canon_sub(row["sub"]) == sub and (row["grp"] == "" or _grp_match(grp, row["grp"])):
                    match = row["rule"]
                    break
            if match is None:
                match = ("blank",)
                log(f"      ⚠ 表頭欄 {get_column_letter(c)} ({grp}|{sub}) 冇對應 rule → 留空")
            rules[c] = match
        return rules

    def insert_blank_col(self, after_sub: str, grp: str, sub: str, log=None) -> None:
        """#1：喺 `after_sub` 嗰欄之後插一條 **output-only** 空白欄（留白、之後按項目合併）。
        表頭.xlsx 冇呢欄 → 純粹喺 template 資料結構加一欄，令輸出多一欄；由項目名驅動（唔靠欄位）。
        右半欄（唔喺 NO_MERGE_SUBS）會照常按項目 span 合併；rule=blank → 值永遠空。"""
        from openpyxl.utils import range_boundaries
        tgt = canon_sub(after_sub)
        P = None
        for c in range(self.anchor or 1, self.maxcol + 1):
            if self.col_gs[c][1] == tgt:
                P = c + 1
                break
        if P is None:
            if log:
                log(f"      ⚠ 插欄揾唔到『{after_sub}』→ 唔插新欄")
            return
        old_max = self.maxcol
        # col_gs / rules / gm 位移（由大到細，避免覆蓋）
        for c in range(old_max, P - 1, -1):
            self.col_gs[c + 1] = self.col_gs.pop(c)
            self.rules[c + 1] = self.rules.pop(c)
            if c in self.gm:
                self.gm[c + 1] = self.gm.pop(c)
        self.col_gs[P] = (nkey(grp), canon_sub(sub))
        self.rules[P] = ("blank",)
        self.gm[P] = nkey(grp)
        # hdr_cells 位移 + 新欄抄左鄰(P-1)風格；row6 填新子表頭，其餘留空（row5 靠延伸合併蓋住）
        for r in range(1, self.subrow + 1):
            for c in range(old_max, P - 1, -1):
                if (r, c) in self.hdr_cells:
                    self.hdr_cells[(r, c + 1)] = self.hdr_cells.pop((r, c))
            nb = self.hdr_cells.get((r, P - 1))
            style = nb[1] if nb else self.data_style
            self.hdr_cells[(r, P)] = (sub if r == self.subrow else None, style)
        self.maxcol = old_max + 1
        # widths 位移（新欄借左鄰闊度）
        old_w = dict(self.widths)
        self.widths = {}
        for c in range(1, self.maxcol + 1):
            src = c if c < P else (P - 1 if c == P else c - 1)
            w = old_w.get(get_column_letter(src))
            if w is not None:
                self.widths[get_column_letter(c)] = w
        # hdr_merges 位移 / 延伸（新欄若喺某組內 → 該組合併延長 1 欄）
        new_merges = []
        for mrg in self.hdr_merges:
            mn_c, mn_r, mx_c, mx_r = range_boundaries(mrg)
            if mx_c < P:
                a, b = mn_c, mx_c
            elif mn_c >= P:
                a, b = mn_c + 1, mx_c + 1
            else:
                a, b = mn_c, mx_c + 1
            new_merges.append(f"{get_column_letter(a)}{mn_r}:{get_column_letter(b)}{mx_r}")
        self.hdr_merges = new_merges
        if log:
            log(f"      ＋插 output-only 空白欄『{sub}』喺第 {P} 欄（{get_column_letter(P)}）")


# ── source 欄 → 目標欄 對位圖 ───────────────────────────────────────────────
def build_col_plan(tpl: Template, src_anchor: int, src_col_gs: dict, src_maxcol: int) -> dict:
    """目標欄 c → ('src', 源欄) 直抄(值+style) / ('derive', rule) 計算欄。
    左半：以 anchor 為尾右對齊 offset；右半：按 rule 嘅 (分組,子表頭) 揾源欄。"""
    plan: dict[int, tuple] = {}
    for c in range(1, tpl.anchor or 1):                 # 左半 offset 對位
        sc = src_anchor - (tpl.anchor - c)
        plan[c] = ("src", sc) if sc >= 1 else ("blank",)
    for c in range(tpl.anchor or 1, tpl.maxcol + 1):    # 右半 by name
        rule = tpl.rules[c]
        if rule[0] in ("copy", "copy_any"):
            want_sub = canon_sub(rule[1] if rule[0] == "copy_any" else rule[2])
            want_grp = None if rule[0] == "copy_any" else rule[1]
            sc = next((scol for scol, (g, s) in src_col_gs.items()
                       if s == want_sub and (want_grp is None or _grp_match(g, want_grp))), None)
            plan[c] = ("src", sc) if sc else ("derive", rule)
        else:
            plan[c] = ("derive", rule)
    return plan


def unmapped_source_cols(tpl: Template, src_anchor: int, src_col_gs: dict) -> list:
    """回 [(源欄, 分組, 子標題)]：source 右半欄冇對應目標 rule、又唔係 seed／調整類型
    → data 會靜靜哋唔帶過（風險 B）。"""
    referenced = []
    for c in range(tpl.anchor or 1, tpl.maxcol + 1):
        rule = tpl.rules[c]
        if rule[0] == "copy_any":
            referenced.append((None, canon_sub(rule[1])))
        elif rule[0] == "copy":
            referenced.append((rule[1], canon_sub(rule[2])))
    out = []
    for sc, (g, s) in sorted(src_col_gs.items()):
        if sc < src_anchor or not s:
            continue
        if s in ADJ_NON_TYPE or _is_adj_type(g, s, s):
            continue                               # seed / 調整類型（會入列舉）
        if not any((rg is None or _grp_match(g, rg)) and s == rs for rg, rs in referenced):
            out.append((sc, g, s))
    return out


def warn_unmapped_source(tpl: Template, src_anchor: int, src_col_gs: dict, log) -> int:
    cols = unmapped_source_cols(tpl, src_anchor, src_col_gs)
    for sc, g, s in cols:
        log(f"      ⚠ source 欄 {get_column_letter(sc)} ({g}|{s}) 冇對應目標 → data 唔會帶過")
    return len(cols)


# ── 寫一個對齊 sheet（照 source 行序、跟 source 文字 style、column 對位）──────
def write_sheet(ws_out, tpl: Template, ws_src, projs: list[Project],
                src_anchor: int, src_subrow: int, src_col_gs: dict,
                src_maxcol: int, src_maxrow: int, log):
    # rich-text lookup dict（load_wb 已 build + cache；dict lookup by plain-text key）
    _rt = getattr(ws_src.parent, '_kpi_rich_table', {})

    # 1) header 區（用 template 34 欄標準表頭 + 佢自己嘅 style）
    for (r, c), (v, sty) in tpl.hdr_cells.items():
        cell = ws_out.cell(r, c, v)
        _apply_style(cell, sty)
    for col, w in tpl.widths.items():
        if w:
            ws_out.column_dimensions[col].width = w
    for r, h in tpl.row_heights.items():                 # #6 表頭行高跟 表頭.xlsx
        if h is not None:
            ws_out.row_dimensions[r].height = h
    for m in tpl.hdr_merges:
        ws_out.merge_cells(m)

    # #1 前置行（rows 1 → src_subrow-2）：公司名/日期/標題等 source 專有內容
    # 覆蓋 template 同位置（template 呢幾行係通用格式，source 有公司名）
    if src_subrow > 2:
        for sr in range(1, src_subrow - 1):
            for sc in range(1, src_maxcol + 1):
                s_cell = ws_src.cell(sr, sc)
                if s_cell.value is not None:
                    t_cell = ws_out.cell(sr, sc)
                    t_cell.value = _rich_val(s_cell, _rt)
                    _apply_text_style(t_cell, s_cell)
        # 前置行橫向合併（唔衝突就照抄，衝突就跳過）
        for mg in ws_src.merged_cells.ranges:
            if 1 <= mg.min_row <= mg.max_row <= src_subrow - 2:
                try:
                    ws_out.merge_cells(
                        start_row=mg.min_row, end_row=mg.max_row,
                        start_column=mg.min_col, end_column=mg.max_col)
                except Exception:
                    pass

    # #2 形狀/文字框（Confidencial 印章等）→ post-save ZIP-level injection（見 _inject_sheet_drawings_raw）
    # deepcopy _drawing 唔 work：openpyxl 唔 load TextBox shapes → 由 process_file 注入

    plan = build_col_plan(tpl, src_anchor, src_col_gs, src_maxcol)
    inv = {sc: c for c, k in plan.items() if k[0] == "src" for sc in (k[1],)}  # 源欄→目標欄

    # #2 欄寬跟原來：source 欄 → 目標欄，抄返 source 嘅 column width。
    # derive/插入欄（冇 source 對應）保留上面 template 嘅寬度。
    from openpyxl.utils import get_column_letter as _gcl
    for c in range(1, tpl.maxcol + 1):
        k = plan.get(c)
        if k and k[0] == "src":
            sdim = ws_src.column_dimensions.get(_gcl(k[1]))
            if sdim is not None and sdim.width:
                ws_out.column_dimensions[_gcl(c)].width = sdim.width

    proj_by_start = {p.r0: p for p in projs}
    src_data0 = src_subrow + 1
    out_row = tpl.subrow + 1
    src_row = src_data0
    right = list(range(tpl.anchor or 1, tpl.maxcol + 1))
    proj_spans_out: list[tuple[int, int, int]] = []      # (out_row, span, src_r0)
    passthru: list[tuple[int, int]] = []                 # (out_row, src_row) 非項目行

    # ── ② 範疇 roll-up：範疇表頭帶「是否需進一步與跨司工作組溝通」= 是（該範疇任一項目要 source_2 跟進）──
    col_comm = next((c for c in right
                     if tpl.col_gs[c][1] == nkey("是否需進一步與跨司工作組溝通")), None)
    band_key_of: dict[int, int] = {}          # band 行 → 佢所屬範疇 key（連續 band 行嘅首行）
    group_starts: list[int] = []
    _prev = None
    for r in range(src_data0, src_maxrow + 1):
        if not _is_band_row(ws_src, r, src_maxcol):
            continue
        if _prev is not None and r == _prev + 1:
            band_key_of[r] = group_starts[-1]
        else:
            group_starts.append(r); band_key_of[r] = r
        _prev = r

    def _scope_key(r0: int):
        cand = [g for g in group_starts if g <= r0]
        return cand[-1] if cand else None

    def _needs_followup(p: Project) -> bool:
        if col_comm is not None and _s(cell_value(tpl, col_comm, p)).startswith("是"):
            return True
        return any(_s(v) != "" for v in p.override.values())     # 有任何 source_2 覆蓋值 = 有跟進

    roll: dict[int, bool] = {}
    for p in projs:
        k = _scope_key(p.r0)
        if k is not None:
            roll[k] = roll.get(k, False) or _needs_followup(p)

    while src_row <= src_maxrow:
        p = proj_by_start.get(src_row)
        if p is not None:
            span = p.r1 - p.r0 + 1
            for rr in range(out_row, out_row + span):    # 打底格線
                for c in range(1, tpl.maxcol + 1):
                    base_grid(ws_out.cell(rr, c), tpl.data_style)
            for i in range(span):                        # #2 行高跟原來
                sh = ws_src.row_dimensions[p.r0 + i].height
                if sh is not None:
                    ws_out.row_dimensions[out_row + i].height = sh
            # 左半：逐行直抄（值 + source 文字 style）
            for i in range(span):
                for c in range(1, tpl.anchor or 1):
                    k = plan[c]
                    if k[0] == "src":
                        s = ws_src.cell(p.r0 + i, k[1])
                        if s.value is not None:
                            ws_out.cell(out_row + i, c).value = _rich_val(s, _rt)
                        _apply_text_style(ws_out.cell(out_row + i, c), s)
            # 右半：anchor 行寫值（source_2 覆蓋優先）+ 跟源欄文字 style
            for c in right:
                cell = ws_out.cell(out_row, c)
                v = cell_value(tpl, c, p)
                cell.value = v
                k = plan[c]
                _grp, sub = tpl.col_gs[c]
                overridden = sub in p.override and _s(p.override[sub]) != ""
                if k[0] == "src" and not overridden:
                    src_cell = ws_src.cell(p.r0, k[1])
                    # 直接從 source 抄 value（_rich_val 保留 run-level bold via ZIP XML）
                    if src_cell.value is not None:
                        cell.value = _rich_val(src_cell, _rt)
                    _apply_text_style(cell, src_cell)
                # W(是否需進一步與跨司工作組溝通)：唔合併 → 逐行填同一值，
                #   否則 unmerge 後其餘行留白。AA(KPMG需與跨司)已改回按項目合併。
                if sub in NO_MERGE_SUBS and span > 1 and _s(v) != "":
                    for i in range(1, span):
                        rep = ws_out.cell(out_row + i, c)
                        rep.value = v
                        if k[0] == "src" and not overridden:
                            _apply_text_style(rep, ws_src.cell(p.r0, k[1]))
            proj_spans_out.append((out_row, span, p.r0))
            out_row += span
            src_row = p.r1 + 1
        else:
            if not _row_blank(ws_src, src_row, src_maxcol):   # 非項目行 → 原樣帶過
                for c in range(1, tpl.maxcol + 1):
                    base_grid(ws_out.cell(out_row, c), tpl.data_style)
                sh = ws_src.row_dimensions[src_row].height     # #2 行高跟原來
                if sh is not None:
                    ws_out.row_dimensions[out_row].height = sh
                for sc in range(1, src_maxcol + 1):
                    tc = inv.get(sc)
                    if tc:
                        s = ws_src.cell(src_row, sc)
                        if s.value is not None:
                            ws_out.cell(out_row, tc).value = _rich_val(s, _rt)
                        _apply_text_style(ws_out.cell(out_row, tc), s)
                # 範疇表頭帶 → 填「是否需進一步與跨司工作組溝通」roll-up（方便按範疇 filter）
                if col_comm is not None and src_row in band_key_of:
                    ws_out.cell(out_row, col_comm).value = \
                        "是" if roll.get(band_key_of[src_row], False) else "否"
                passthru.append((out_row, src_row))
                out_row += 1
            src_row += 1

    # 2) merges：右半項目 span 內合併（全部 cell 已有框 → 四邊齊）
    #    NO_MERGE_SUBS 欄（W 是否需進一步）唔合併 → 跳過（逐行已填值）
    for orow, span, _r0 in proj_spans_out:
        if span > 1:
            for c in right:
                if tpl.col_gs[c][1] in NO_MERGE_SUBS:
                    continue
                ws_out.merge_cells(start_row=orow, end_row=orow + span - 1,
                                   start_column=c, end_column=c)
    # 非項目行嘅橫向合併（小標題橫跨幾欄）→ 重映射後保留
    src_hmerge = [(m.min_row, m.min_col, m.max_col) for m in ws_src.merged_cells.ranges
                  if m.min_row == m.max_row and m.max_col > m.min_col]
    row_of = {sr: orow for orow, sr in passthru}
    for sr, c0, c1 in src_hmerge:
        if sr in row_of:
            t0, t1 = inv.get(c0), inv.get(c1)
            if t0 and t1 and t1 > t0:
                ws_out.merge_cells(start_row=row_of[sr], end_row=row_of[sr],
                                   start_column=min(t0, t1), end_column=max(t0, t1))
    # #6 資料行高：跟返 source 原檔（項目 span 逐行對 + 非項目行對）
    src_to_out = {}
    for orow, span, r0 in proj_spans_out:
        for i in range(span):
            src_to_out[r0 + i] = orow + i
    src_to_out.update(row_of)
    for sr, orow in src_to_out.items():
        h = ws_src.row_dimensions[sr].height
        if h is not None:
            ws_out.row_dimensions[orow].height = h
    return out_row


def light_copy_sheet(ws_out, ws_src):
    """附件/非標準 sheet → 原樣 best-effort 複製（值 + 全 style + 合併 + 欄寬 + 行高）。"""
    for r in range(1, (ws_src.max_row or 0) + 1):
        for c in range(1, (ws_src.max_column or 0) + 1):
            copy_full_cell(ws_out.cell(r, c), ws_src.cell(r, c))
    for m in ws_src.merged_cells.ranges:
        ws_out.merge_cells(str(m))
    for c in range(1, (ws_src.max_column or 0) + 1):
        L = get_column_letter(c)
        w = ws_src.column_dimensions[L].width
        if w:
            ws_out.column_dimensions[L].width = w
    for r, dim in ws_src.row_dimensions.items():
        if dim.height is not None:
            ws_out.row_dimensions[r].height = dim.height
    # #8 附件內嵌圖片 best-effort 過帶（WMF 已喺 openpyxl load 時 drop、救唔返）
    try:
        from openpyxl.drawing.image import Image as _XLImage
        for im in getattr(ws_src, "_images", []) or []:
            try:
                ref = getattr(im, "ref", None)
                if ref is None:
                    continue
                ni = _XLImage(ref)
                anc = getattr(im, "anchor", None)
                if anc is not None:
                    ni.anchor = anc
                ws_out.add_image(ni)
            except Exception:
                pass
    except Exception:
        pass
    # #2 形狀/文字框 drawing（私密Confidencial 等 TextBox）
    try:
        from copy import deepcopy as _dc
        drw = getattr(ws_src, "_drawing", None)
        if drw is not None:
            ws_out._drawing = _dc(drw)
    except Exception:
        pass


# ── source_2 overlay ──────────────────────────────────────────────────────
def find_overlay_file(root: Path, scope: str, company: str) -> Path | None:
    base = root / "source_2"
    if not base.exists():
        return None
    cands = []
    for p in base.rglob("*"):
        if p.is_dir() or p.name.startswith("~$") or p.suffix.lower() not in EXCEL_EXT:
            continue
        rel = p.relative_to(base).as_posix()
        if scope and scope in rel and company and company in rel:
            cands.append(p)
    return sorted(cands, key=lambda x: len(x.as_posix()))[0] if cands else None


def _concern_sum(ws, p, col_gs) -> float | None:
    """source_2 多個關注事項嘅調整金額加總（numeric，方便 Excel 加總）。
    逐行掃『該關注事項涉及調整金額』非空數字，去重後加 |abs|，返回總和；冇 → None。"""
    col_amt = next((c for c, (g, s) in col_gs.items()
                    if s == nkey("該關注事項涉及調整金額")), None)
    if col_amt is None:
        return None
    total, prev, found = 0.0, None, False
    for r in range(p.r0, p.r1 + 1):
        a = num(ws.cell(r, col_amt).value)
        if a is None or a == prev:       # 逐行重複（冇 merge）→ 去重
            continue
        prev = a
        total += abs(a)
        found = True
    return total if found else None


def build_overlay(path: Path, log) -> dict[str, dict]:
    """source_2 per-範疇檔 → {項目序號: {sub_canon: value}}（掃全部 sheet，只攞 OVERLAY_SUBS）。"""
    try:
        wb = load_wb(path)
    except Exception as e:
        log(f"      ⚠ overlay 開唔到 {path.name}: {e}")
        return {}
    out: dict[str, dict] = {}
    log(f"      overlay 檔 sheets: {wb.sheetnames}")
    for sn in wb.sheetnames:
        ws = wb[sn]
        projs, subrow, anchor, gm, maxcol, col_gs, maxrow = extract(ws, log)
        if not anchor or not projs:
            log(f"      overlay sheet {sn!r}: anchor={anchor} 項目={len(projs)} → 跳過")
            continue
        found = 0
        for p in projs:
            d = {}
            for (g, s), v in p.by_gs.items():
                if s not in OVERLAY_SUBS or _s(v) == "":
                    continue
                # 該關注事項涉及調整金額：由下面 _concern_enum 統一砌（多關注事項串一格），呢度跳過
                if s == nkey("該關注事項涉及調整金額"):
                    continue
                d[s] = v
            # source_2 多關注事項 → 各 concern 金額加總（numeric）填入該欄（方便加總）
            cs = _concern_sum(ws, p, col_gs)
            if cs is not None:
                d[nkey("該關注事項涉及調整金額")] = cs
            if d:
                out[_seqkey(p.seq)] = d
                found += 1
        log(f"      overlay sheet {sn!r}: {len(projs)} 項目, {found} 個有覆蓋值 "
            f"(序號: {[_seqkey(p.seq) for p in projs]})")
    wb.close()
    return out


def _seqkey(seq: str) -> str:
    """join key：去 label 前綴後，攞項目碼（76/CE001/項目3）做鍵；否則純數字或全名 fallback。
    source_1／source_2 兩邊都行同一 _derive_seq→_seqkey，同一項目 → 同一鍵，先對得返。"""
    s = re.sub(r"^.*?投資項目序號及名稱[:：]?\s*", "", _s(seq)).strip()
    m = re.match(r"^([A-Za-z]{0,5}\d+|項目\d+)\s*[-–]", s)
    if m:
        return nkey(m.group(1))
    m2 = re.match(r"^\s*(\d+)", s)
    return m2.group(1) if m2 else nkey(s)


def apply_overlay(projs, overlay: dict, tpl: Template, log):
    hits = 0
    for p in projs:
        d = overlay.get(_seqkey(p.seq))
        if not d:
            continue
        hits += 1
        for s, v in d.items():
            p.override[s] = v               # 為準，寫入時蓋過計算值
    log(f"      overlay 命中 {hits}/{len(projs)} 個項目")


# ── preview ───────────────────────────────────────────────────────────────
def preview_sheet(sn, projs, tpl: Template, log):
    log(f"\n  ── sheet {sn!r}：{len(projs)} 個項目 ──")
    for p in projs[:6]:
        log(f"\n    ▸ 項目「{p.seq}」 rows {p.r0}-{p.r1}（{p.r1-p.r0+1}行）")
        log(f"       左半首行: {[_s(x) for x in (p.left[0] if p.left else [])][:6]}")
        if p.adj:
            log(f"       調整類型(非零): " + "; ".join(f"{t}={fmt_amt(a)}" for t, a in p.adj))
        t = _adj_total(p)
        log(f"       seed: 申報={_s(p.seed.get(nkey('申報投資金額')))} "
            f"潛在調整合計={_s(p.seed.get(nkey('潛在調整合計')))} "
            f"調整後={_s(p.seed.get(nkey('調整後投資金額')))} "
            f"跨司確認={_s(p.seed.get(nkey('跨司工作組確認投資金額')))} "
            f"→_adj_total={fmt_amt(t) if t is not None else 'None'} has_adj={p.has_adj()}")
        # 表頭尾段會寫乜
        for c in range(tpl.anchor or 1, tpl.maxcol + 1):
            g, s = tpl.col_gs[c]
            if s in {nkey("調整原因"), nkey("該關注事項涉及調整金額"), nkey("建議調整金額"),
                     nkey("建議調整後金額"), nkey("建議接納之調整後金額"),
                     nkey("需溝通關注事項"), nkey("是否需進一步與跨司工作組溝通"),
                     nkey("畢馬威關注事項"), nkey("承批公司的反饋意見"),
                     nkey("跨司工作組主責部門針對該關注事項已給的反饋意見"),
                     nkey("KPMG需與跨司工作組進一步確認的問題"),
                     nkey("跨司工作組最新反饋意見")}:
                v = cell_value(tpl, c, p)
                if _s(v):
                    tag = " [source_2]" if s in p.override and _s(p.override[s]) != "" else ""
                    disp = _s(v).replace("\n", " ⏎ ")
                    log(f"       → 表頭[{get_column_letter(c)} {s}]{tag} = {disp[:120]}")
    if len(projs) > 6:
        log(f"    …（另 {len(projs)-6} 個項目略）")


# ── 每檔處理 ───────────────────────────────────────────────────────────────
# 保留作參考／audit 用；scope 偵測已改用資料夾名，唔再靠呢張清單
SCOPE_HINTS = ["旅遊局", "博監局", "經濟局", "文化局", "體育局", "郵電局", "其他範疇", "治安警"]


def infer_scope_company(rel: str) -> tuple[str, str]:
    # scope = source_1 下面第一層資料夾（範疇名）——最穩陣，唔會漏範疇。
    # 例如 source_1/會議展覽範疇/Galaxy-…xlsx → scope=會議展覽範疇。
    parts = Path(rel).parts
    scope = ""
    if len(parts) >= 2 and parts[0].lower() == "source_1":
        scope = parts[1]
    elif len(parts) >= 2:
        scope = parts[-2]
    m = re.search(r"([A-Za-z]{2,})[-\-]", Path(rel).name)
    company = m.group(1) if m else ""
    return scope, company


def is_attachment_sheet(name: str) -> bool:
    return any(k in name for k in ("附件", "承批公司附件", "attach", "Attach"))


def is_junk_sheet(name: str) -> bool:
    """UpSlide 外掛留低嘅垃圾 sheet（UPSLIDE_Undo 等）→ 唔抄。"""
    return name.strip().upper().startswith("UPSLIDE")


def is_dup_file(rel: str) -> bool:
    """使用者為咗開嚟 check 而整嘅同名副本（… - Copy / 副本 / 複本）→ 批次跳過。"""
    stem = Path(rel).stem
    return any(k in stem for k in (" - Copy", "-Copy", "- Copy", "副本", "複本"))


def process_file(root: Path, rel: str, tpl: Template, out_dir: Path,
                 preview: bool, with_overlay: bool, log):
    src = root / rel
    log(f"\n{'='*74}\n# {rel}")
    if not src.exists():
        log("  ✗ 唔存在"); return
    wb = load_wb(src)
    scope, company = infer_scope_company(rel)
    overlay = {}
    if with_overlay:
        ofile = find_overlay_file(root, scope, company)
        if ofile:
            log(f"  overlay 檔: {ofile.relative_to(root).as_posix()}")
            overlay = build_overlay(ofile, log)
            if not overlay:
                log("  ⚠ overlay 抽唔到任何覆蓋值（下面 base 為準）")
        else:
            log(f"  （冇 source_2 per-範疇檔 match scope={scope} company={company}）")

    out_wb = openpyxl.Workbook()
    out_wb.remove(out_wb.active)
    aligned_sheets: list[str] = []   # source sheet names processed by write_sheet（用於 drawing injection）
    for sn in wb.sheetnames:
        ws = wb[sn]
        if is_junk_sheet(sn):
            log(f"  ── sheet {sn!r}：外掛垃圾 sheet → 唔抄（跳過）")
            continue
        if is_attachment_sheet(sn):
            log(f"  ── sheet {sn!r}：附件 → 原樣照抄（best-effort）")
            if not preview:
                light_copy_sheet(out_wb.create_sheet(sn), ws)
            continue
        projs, subrow, anchor, gm, maxcol, col_gs, maxrow = extract(ws, log)
        if not anchor or not projs:
            log(f"  ── sheet {sn!r}：揾唔到「是否該司局範疇」anchor 或冇項目 → 原樣照抄")
            if not preview:
                light_copy_sheet(out_wb.create_sheet(sn), ws)
            continue
        warn_unmapped_source(tpl, anchor, col_gs, log)      # 風險 B：漏欄示警
        if overlay:
            apply_overlay(projs, overlay, tpl, log)
        if preview:
            preview_sheet(sn, projs, tpl, log)
        else:
            write_sheet(out_wb.create_sheet(sn[:31]), tpl, ws, projs,
                        anchor, subrow, col_gs, maxcol, maxrow, log)
            aligned_sheets.append(sn)
    src_zip_src = getattr(wb, '_kpi_src_zip_src', None)
    wb.close()
    if preview:
        return
    out_path = out_dir / rel
    out_path.parent.mkdir(parents=True, exist_ok=True)
    if len(out_wb.sheetnames) == 0:
        out_wb.create_sheet("空")
    if ENCRYPT_OUT:                              # 保留 source 密碼保護
        tmp = out_path.with_suffix(".plain.xlsx")
        out_wb.save(tmp)
        # post-save ZIP injection：shapes/TextBox（Confidencial 等）
        if src_zip_src and aligned_sheets:
            for sn in aligned_sheets:
                _inject_sheet_drawings_raw(src_zip_src, sn, tmp)
        if encrypt_file(tmp, out_path, log):
            tmp.unlink(missing_ok=True)
            log(f"  ✓ 寫入 {out_path.relative_to(root).as_posix()}（已加密 dicj_kpmg）")
        else:
            shutil.move(str(tmp), str(out_path))
            log(f"  ✓ 寫入 {out_path.relative_to(root).as_posix()}（未加密）")
    else:
        out_wb.save(out_path)
        log(f"  ✓ 寫入 {out_path.relative_to(root).as_posix()}")


def copy_asis(root: Path, rel: str, out_dir: Path, log):
    src, dst = root / rel, out_dir / rel
    dst.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src, dst)
    log(f"  ⧉ 照抄 {rel}")


def iter_source1(root: Path):
    base = root / "source_1"
    for p in sorted(base.rglob("*")):
        if p.is_dir() or p.name.startswith("~$"):
            continue
        if p.suffix.lower() in EXCEL_EXT:
            yield p.relative_to(root).as_posix()


def _is_encrypted(path: Path) -> bool:
    """加密 OOXML 係 OLE/CFB 容器（magic D0CF11E0）。"""
    try:
        with open(path, "rb") as f:
            return f.read(4) == b"\xd0\xcf\x11\xe0"
    except Exception:
        return False


def _quiet(*_a, **_k):
    pass


# ── read-only 巡查全部 source_1：逐檔逐 sheet 報 A–F 狀態（唔寫檔）─────────────
def audit(root: Path, tpl: Template, log):
    files = list(iter_source1(root))
    fA, fB, fC, fE, fAtt, ok = [], [], [], [], [], []
    tpl_left = [tpl.col_gs[c][1] for c in range(1, tpl.anchor or 1)]
    log(f"# AUDIT（read-only，唔寫檔）：source_1 共 {len(files)} 檔  "
        f"／ template 左半 {(tpl.anchor or 1) - 1} 欄、anchor {get_column_letter(tpl.anchor) if tpl.anchor else '?'}")
    log(f"# template 左半名（右對齊基準，尾欄=貼 anchor 前）: {tpl_left}\n")
    for rel in files:
        if any(pp.lower() == "ss" for pp in Path(rel).parts):
            log(f"▸ {rel}   [ss/ → 照抄，跳過分析]")
            continue
        if is_dup_file(rel):
            log(f"▸ {rel}   [重複檔（使用者 check 用）→ 批次跳過]")
            continue
        src = root / rel
        enc = "加密" if _is_encrypted(src) else "無密碼"
        try:
            wb = load_wb(src)
        except Exception as e:
            log(f"▸ {rel}   ✗ 開唔到: {type(e).__name__}: {e}")
            fA.append(f"{rel} (開唔到)")
            continue
        log(f"▸ {rel}   [{enc}]  sheets={wb.sheetnames}")
        scope, company = infer_scope_company(rel)
        for sn in wb.sheetnames:
            ws = wb[sn]
            if is_junk_sheet(sn):
                log(f"    · {sn!r}  外掛垃圾 → 唔抄（跳過）")
                continue
            if is_attachment_sheet(sn):
                log(f"    · {sn!r}  附件 → best-effort 照抄")
                fAtt.append(f"{rel}::{sn}")
                continue
            projs, subrow, anchor, gm, maxcol, col_gs, maxrow = extract(ws, _quiet)
            older = any(y in sn for y in ("2024", "2023", "2022"))
            if not anchor or not projs:
                log(f"    · {sn!r}  ✗ 冇 anchor／冇項目 → 對齊唔到,只會照抄"
                    f"{'（舊年份）' if older else ''}")
                fA.append(f"{rel}::{sn}")
                continue
            leftn = anchor - 1
            tleft = (tpl.anchor or 1) - 1
            offset = (tpl.anchor or 1) - anchor
            unm = unmapped_source_cols(tpl, anchor, col_gs)
            adjn = sum(1 for _sc, (g, s) in col_gs.items() if _is_adj_type(g, s, s))
            flags = []
            if leftn != tleft:
                flags.append(f"左半{leftn}欄≠template{tleft}(E)"); fE.append(f"{rel}::{sn}")
            if unm:
                flags.append(f"漏欄{len(unm)}(B)"); fB.append(f"{rel}::{sn}")
            if older:
                flags.append("舊年份會被夾成34欄(C)"); fC.append(f"{rel}::{sn}")
            status = "  ".join(flags) if flags else "✓ 全對齊"
            log(f"    · {sn!r}  anchor={get_column_letter(anchor)}  左半{leftn}欄(offset{offset:+d})  "
                f"項目{len(projs)}  調整類型欄{adjn}  →  {status}")
            if leftn != tleft:                        # E：印左半名，眼睇右對齊落邊
                src_left = [col_gs[c][1] for c in range(1, anchor)]
                log(f"        左半名(source): {src_left}")
                log(f"        右對齊落 template 欄 {get_column_letter(offset + 1)}..{get_column_letter(tpl.anchor - 1)}"
                    f"（template 頭 {offset} 欄留空）")
            for sc, g, s in unm:
                log(f"        ⚠ 漏: 欄{get_column_letter(sc)} ({g} | {s})")
            if not flags:
                ok.append(f"{rel}::{sn}")
        of = find_overlay_file(root, scope, company)
        log(f"    overlay(source_2): {('✓ ' + of.relative_to(root).as_posix()) if of else '✗ 冇對應檔（該檔無覆蓋）'}")
        wb.close()
    log("\n" + "=" * 64 + "\n# A–F 風險匯總")
    log(f"A 對齊唔到(冇anchor/項目)，只會照抄 : {len(fA)}")
    for x in fA:
        log(f"    - {x}")
    log(f"B 有漏欄(會靜靜哋掉數)              : {len(fB)}")
    for x in fB:
        log(f"    - {x}")
    log(f"C 含舊年份 sheet(會一齊夾34欄)      : {len(fC)}")
    for x in fC:
        log(f"    - {x}")
    log(f"E 左半欄數 ≠ template(識別欄會右移) : {len(fE)}")
    for x in fE:
        log(f"    - {x}")
    log(f"F 附件(best-effort 照抄)           : {len(fAtt)}")
    log(f"✓ 全對齊冇 flag 嘅 sheet           : {len(ok)}")
    log("D 加密：audit 唔加密；實跑每檔會用 dicj_kpmg 重加密，只有 log 出「⚠ 加密失敗」先算 fail")


# ── read-only 數字對數（reconciliation）：唔寫檔，查每個項目金額拼唔拼得返 ──────
TOL_ADJ = 0.5           # 對數容差（萬澳門元）；細過呢個當「拍得住」（避開 rounding）


def _project_number_checks(p: Project, tol: float) -> tuple[list[str], list[str]]:
    """回 (識別_issues, 映射_issues)。
    識別 = source 本身／我哋抽取有冇拼錯（#3『調整金額弄亂』會喺呢度爆）：
        · 潛在調整合計 ?= Σ(逐項調整類型)
        · 調整後投資金額 ?= 申報投資金額 + 調整
    映射 = 我哋寫落表頭尾段嗰幾個數 ?= source seed（seed／enum／abs_total 邏輯啱唔啱）。"""
    idi, mpi = [], []
    total = num(p.seed.get(nkey("潛在調整合計")))
    before = num(p.seed.get(nkey("申報投資金額")))
    after = num(p.seed.get(nkey("調整後投資金額")))
    sigma = sum(a for _t, a in p.adj) if p.adj else None
    # 識別①：合計 = Σ逐項調整類型
    if total is not None and sigma is not None and abs(total - sigma) > tol:
        idi.append(f"潛在調整合計={fmt_amt(total)} ≠ Σ逐項調整={fmt_amt(sigma)}"
                   f"（差{fmt_amt(total - sigma)}）")
    # 識別②：調整後 = 申報 + 調整
    adj_id = total if total is not None else sigma
    if after is not None and before is not None and adj_id is not None:
        exp = before + adj_id
        if abs(after - exp) > tol:
            idi.append(f"調整後={fmt_amt(after)} ≠ 申報{fmt_amt(before)}+調整{fmt_amt(adj_id)}"
                       f"={fmt_amt(exp)}（差{fmt_amt(after - exp)}）")
    # 映射①：建議調整金額 ← 潛在調整合計
    g_adj = num(resolve(("seed", "潛在調整合計"), p))
    if g_adj is not None and total is not None and abs(g_adj - total) > tol:
        mpi.append(f"建議調整金額 寫={fmt_amt(g_adj)} ≠ 潛在調整合計={fmt_amt(total)}")
    # 映射②：建議調整後金額 ← 調整後投資金額
    g_after = num(resolve(("seed", "調整後投資金額"), p))
    if g_after is not None and after is not None and abs(g_after - after) > tol:
        mpi.append(f"建議調整後金額 寫={fmt_amt(g_after)} ≠ 調整後投資金額={fmt_amt(after)}")
    # 映射③：該關注事項涉及調整金額 ← |調整總額|
    g_abs = resolve(("abs_total",), p)
    tot_abs = _adj_total(p)
    if g_abs is not None and tot_abs is not None and abs(g_abs - abs(tot_abs)) > tol:
        mpi.append(f"該關注事項涉及調整金額 寫={fmt_amt(g_abs)} ≠ |調整總額|={fmt_amt(abs(tot_abs))}")
    # 映射④：調整原因 列舉金額加總（逐項→Σ|類型|；只有總數→|總額|）
    enum = build_enum(p)
    if enum:
        esum = sum(float(x) for x in re.findall(r"[:：]\s*(-?\d+(?:\.\d+)?)\s*萬", enum))
        exp_e = (sum(abs(a) for _t, a in p.adj) if p.adj
                 else (abs(tot_abs) if tot_abs is not None else None))
        if exp_e is not None and abs(esum - exp_e) > tol:
            mpi.append(f"調整原因 列舉加總={fmt_amt(esum)} ≠ 應有={fmt_amt(exp_e)}")
    return idi, mpi


def verify(root: Path, tpl: Template, tol: float, log):
    """逐個 source_1 sheet 逐項目對數：識別算式（#3 會爆）＋我哋映射寫嘅數；
    另出每 sheet／全域 Σ申報／Σ調整後／Σ合計 頂線，俾你同自己 total 拍。唔寫 xlsx。"""
    files = list(iter_source1(root))
    n_proj = n_idfail = n_mpfail = 0
    G_before = G_after = G_total = 0.0
    G_s2_amt = 0.0            # Σ source_2 該關注事項涉及調整金額
    G_s2_hit = G_s2_miss = 0  # source_2 命中/未命中 項目數
    log(f"# VERIFY 數字對數（read-only，唔寫檔）：source_1 共 {len(files)} 檔  容差 {fmt_amt(tol)} 萬\n")
    for rel in files:
        if any(pp.lower() == "ss" for pp in Path(rel).parts) or is_dup_file(rel):
            continue
        try:
            wb = load_wb(root / rel)
        except Exception as e:
            log(f"▸ {rel}   ✗ 開唔到: {type(e).__name__}: {e}")
            continue
        printed = False
        file_projs: list = []          # 全檔項目（供 source_2 對數）
        for sn in wb.sheetnames:
            ws = wb[sn]
            if is_junk_sheet(sn) or is_attachment_sheet(sn):
                continue
            projs, subrow, anchor, gm, maxcol, col_gs, maxrow = extract(ws, _quiet)
            if not anchor or not projs:
                continue
            file_projs.extend(projs)
            s_before = s_after = s_total = 0.0
            fails = []
            for p in projs:
                n_proj += 1
                idi, mpi = _project_number_checks(p, tol)
                if idi:
                    n_idfail += 1
                if mpi:
                    n_mpfail += 1
                if idi or mpi:
                    fails.append((p, idi, mpi))
                s_before += num(p.seed.get(nkey("申報投資金額"))) or 0
                s_after += num(p.seed.get(nkey("調整後投資金額"))) or 0
                s_total += num(p.seed.get(nkey("潛在調整合計"))) or 0
            G_before += s_before
            G_after += s_after
            G_total += s_total
            if not printed:
                log(f"▸ {rel}")
                printed = True
            log(f"    · {sn!r}  項目{len(projs)}  Σ申報={fmt_amt(s_before)} "
                f"Σ調整後={fmt_amt(s_after)} Σ合計={fmt_amt(s_total)}"
                + (f"  ⚠ 對唔到 {len(fails)} 個項目" if fails else "  ✓ 全對得返"))
            for p, idi, mpi in fails[:25]:
                for msg in idi:
                    log(f"        ✗[識別] 「{p.seq}」 rows {p.r0}-{p.r1}: {msg}")
                for msg in mpi:
                    log(f"        ✗[映射] 「{p.seq}」 rows {p.r0}-{p.r1}: {msg}")

        # ── #3 source_2 對數：套返 per-範疇 overlay，對「該關注事項涉及調整金額」──
        scope, company = infer_scope_company(rel)
        ofile = find_overlay_file(root, scope, company)
        if ofile and file_projs:
            overlay = build_overlay(ofile, _quiet)
            f_amt = 0.0
            f_hit = f_miss = 0
            miss_seqs = []
            for p in file_projs:
                d = overlay.get(_seqkey(p.seq))
                if d:
                    f_hit += 1
                    a = num(d.get(nkey("該關注事項涉及調整金額")))
                    if a:
                        f_amt += abs(a)
                else:
                    # 有潛在調整但 source_2 揾唔到跟進 → 值得核
                    t = _adj_total(p)
                    if t is not None and abs(t) > tol:
                        f_miss += 1
                        if len(miss_seqs) < 6:
                            miss_seqs.append(p.seq)
            G_s2_amt += f_amt
            G_s2_hit += f_hit
            G_s2_miss += f_miss
            if not printed:
                log(f"▸ {rel}")
                printed = True
            log(f"    ⟐ source_2 {ofile.name}: 命中 {f_hit} 項目  "
                f"Σ涉及調整金額={fmt_amt(f_amt)}"
                + (f"  ⚠ {f_miss} 個有調整但 source_2 冇跟進" if f_miss else ""))
            for sq in miss_seqs:
                log(f"        ⚠[source_2缺] 「{sq}」有潛在調整但 source_2 揾唔到")
        wb.close()
    log("\n" + "=" * 64 + "\n# 對數匯總")
    log(f"項目總數                    : {n_proj}")
    log(f"識別/抽取算式對唔返(調整)   : {n_idfail}   ← #3『調整金額弄亂』會喺呢度爆")
    log(f"我哋映射寫錯數              : {n_mpfail}   ← 應為 0（seed/enum/abs_total 邏輯）")
    log(f"Σ申報投資金額              : {fmt_amt(G_before)} 萬")
    log(f"Σ調整後投資金額            : {fmt_amt(G_after)} 萬")
    log(f"Σ潛在調整合計              : {fmt_amt(G_total)} 萬")
    log(f"頂線對數 Σ調整後-Σ申報      : {fmt_amt(G_after - G_before)} 萬  "
        f"(應 ≈ Σ合計 {fmt_amt(G_total)}，差 {fmt_amt(G_after - G_before - G_total)})")
    log("\n# source_2（跨司工作組跟進）對數")
    log(f"source_2 命中項目            : {G_s2_hit}")
    log(f"有調整但 source_2 冇跟進     : {G_s2_miss}   ← 值得逐個核")
    log(f"Σ source_2 涉及調整金額     : {fmt_amt(G_s2_amt)} 萬  "
        f"(對 |Σ潛在調整合計|={fmt_amt(abs(G_total))}，差 {fmt_amt(G_s2_amt - abs(G_total))})")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--root", default="ad-hoc/workspace")
    ap.add_argument("--header-file", default="表頭.xlsx")
    ap.add_argument("--only", help="淨處理呢一個相對 root 嘅檔（驗證用）")
    ap.add_argument("--all", action="store_true", help="批 source_1 全部檔")
    ap.add_argument("--audit", action="store_true",
                    help="read-only 巡查全部 source_1，逐檔逐 sheet 報 A–F 風險（唔寫檔）")
    ap.add_argument("--verify", action="store_true",
                    help="read-only 對數：逐項目查調整算式＋映射寫嘅數，出 Σ 頂線（唔寫檔）")
    ap.add_argument("--tol", type=float, default=TOL_ADJ,
                    help="對數容差（萬澳門元，預設 0.5）")
    ap.add_argument("--preview", action="store_true", help="只吐抽取+映射文字，唔寫 xlsx")
    ap.add_argument("--with-overlay", action="store_true",
                    help="（已預設開，保留兼容）套 source_2 per-範疇覆蓋")
    ap.add_argument("--no-overlay", action="store_true",
                    help="關閉 source_2 覆蓋（預設會套 source_2；呢個先會唔套）")
    ap.add_argument("--fill-zero-adj", action="store_true",
                    help="冇調整都照抄 0（預設留空）")
    ap.add_argument("--no-encrypt", action="store_true",
                    help="輸出唔加密（預設用 dicj_kpmg 重新加密）")
    ap.add_argument("--no-extra-col", action="store_true",
                    help="唔插『跨司工作組的反饋意見』output-only 空白欄（#1）")
    ap.add_argument("--out", default="_aligned", help="輸出子資料夾（root 底下）")
    a = ap.parse_args()
    global GATE_NOADJ, ENCRYPT_OUT
    GATE_NOADJ = not a.fill_zero_adj
    ENCRYPT_OUT = not a.no_encrypt
    # source_2 覆蓋預設開；--no-overlay 先關（--with-overlay 保留但已係預設）
    use_overlay = not a.no_overlay
    root = Path(a.root)
    out_dir = root / a.out
    lines = []

    def log(s=""):
        lines.append(s); print(s)

    hf = root / a.header_file
    if not hf.exists():
        raise SystemExit(f"✗ 冇表頭: {hf}")
    tpl = Template(hf, log)
    log(f"# 表頭 template：{tpl.maxcol} 欄, 子表頭 row{tpl.subrow}, anchor 欄 "
        f"{get_column_letter(tpl.anchor) if tpl.anchor else '?'}")

    if a.audit:
        audit(root, tpl, log)
        rp = root / "_align_audit.txt"
        rp.write_text("\n".join(lines), encoding="utf-8")
        print(f"\n✓ audit 寫入 {rp.resolve()}（貼返嚟俾我）")
        return

    if a.verify:
        verify(root, tpl, a.tol, log)
        rp = root / "_align_verify.txt"
        rp.write_text("\n".join(lines), encoding="utf-8")
        print(f"\n✓ verify 寫入 {rp.resolve()}（貼返嚟俾我）")
        return

    if INSERT_XSGZ_FB_COL and not a.no_extra_col:      # #1 output-only 空白欄
        tpl.insert_blank_col("承批公司的反饋意見", GT, "跨司工作組的反饋意見", log)

    if a.only:
        targets = [a.only]
    elif a.all:
        targets = list(iter_source1(root))
    else:
        raise SystemExit("俾 --only <rel> 或 --all")

    for rel in targets:
        parts = Path(rel).parts
        if any(pp.lower() == "ss" for pp in parts):
            if not a.preview:
                copy_asis(root, rel, out_dir, log)
            continue
        if a.all and is_dup_file(rel):
            log(f"  ⏭ 跳過重複檔（使用者 check 用）: {rel}")
            continue
        try:
            process_file(root, rel, tpl, out_dir, a.preview, use_overlay, log)
        except Exception as e:
            import traceback
            log(f"  ✗ 出錯: {type(e).__name__}: {e}")
            log(traceback.format_exc())

    rp = root / ("_align_preview.txt" if a.preview else "_align_log.txt")
    rp.write_text("\n".join(lines), encoding="utf-8")
    print(f"\n✓ log 寫入 {rp.resolve()}（貼返嚟俾我）")


if __name__ == "__main__":
    main()
