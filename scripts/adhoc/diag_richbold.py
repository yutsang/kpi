#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
diag_richbold.py — 查真實 source 檔嘅 bold 到底喺邊、_make_rich_lookup 收唔收到、
key 對唔對得上 cell.value。用嚟解「輸出完全冇 bold」。

用法：
    python scripts\\adhoc\\diag_richbold.py ^
        "ad-hoc\\workspace\\source_1\\旅遊局\\SJM-投資計劃執行情況表二（旅遊局）.xlsx"

會印：
  [A] sharedStrings 裡面有幾多個 multi-run（rich）entry、頭幾個 key + run 嘅 b 狀態
  [B] 逐 sheet 掃 cell：
        - value 係 str 且喺 lookup（rich 命中）幾多個
        - value 係 str 但唔喺 lookup（可能 key mismatch）幾多個，抽樣印出對唔上嘅 value
        - cell.font.bold=True（cell-level 整格 bold）幾多個，抽樣
  [C] 結論提示
"""
import io
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
import align_to_header as A  # noqa: E402


def _load(path: Path):
    """returns (wb, rich_lookup_text_view)。
    _make_rich_lookup 而家回傳 (index_runs, si_cell_map)；呢度砌返一個
    {plain_text: runs} view 俾下面顯示用（純顯示，撞唔撞 key 唔緊要）。"""
    import openpyxl

    def _textview(rich):
        ir = rich[0] if isinstance(rich, tuple) else (rich or {})
        return {''.join(t for t, *_ in runs): runs for runs in ir.values()}

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        return wb, _textview(A._make_rich_lookup(path))
    except Exception:
        import msoffcrypto
        buf = io.BytesIO()
        with open(path, "rb") as f:
            off = msoffcrypto.OfficeFile(f)
            off.load_key(password=A.PASSWORD)
            off.decrypt(buf)
        buf.seek(0)
        wb = openpyxl.load_workbook(buf, data_only=True)
        buf.seek(0)
        return wb, _textview(A._make_rich_lookup(buf))


def _open_any(path: Path, rich=True):
    import openpyxl
    try:
        return openpyxl.load_workbook(path, data_only=True, rich_text=rich)
    except Exception:
        import msoffcrypto
        buf = io.BytesIO()
        with open(path, "rb") as f:
            off = msoffcrypto.OfficeFile(f)
            off.load_key(password=A.PASSWORD)
            off.decrypt(buf)
        buf.seek(0)
        return openpyxl.load_workbook(buf, data_only=True, rich_text=rich)


def _dump_uv(path: Path, label: str, max_cells: int = 6):
    """開檔（讀 rich text），dump U/V 欄 cell 嘅逐 run bold（睇實邊邊有冇 bold）。"""
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.utils import get_column_letter
    wb = _open_any(path, rich=True)
    print(f"  --- {label}: {path.name} ---")
    shown = 0
    for sn in wb.sheetnames:
        ws = wb[sn]
        for row in ws.iter_rows():
            for c in row:
                col = get_column_letter(c.column)
                if col not in ("U", "V"):
                    continue
                v = c.value
                if v is None or (isinstance(v, str) and not v.strip()):
                    continue
                cell_bold = bool(c.font and c.font.bold)
                if isinstance(v, CellRichText):
                    runs = [(p.text, getattr(p.font, 'b', None)) if isinstance(p, TextBlock)
                            else (str(p), 'plain') for p in v]
                    print(f"    {sn}!{col}{c.row}: RICH cell-bold={cell_bold}")
                    for t, b in runs[:6]:
                        print(f"        b={b}  {t[:45]!r}")
                else:
                    print(f"    {sn}!{col}{c.row}: plain cell-bold={cell_bold}  {str(v)[:45]!r}")
                shown += 1
                if shown >= max_cells:
                    return
    if shown == 0:
        print("    （U/V 欄冇搵到有內容嘅 cell）")


def _find_uv(path: Path, substr: str, label: str):
    """搵晒含 substr 嘅 U/V cell，印座標 + cell.font.bold + 每段 run 嘅明確 b。"""
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.utils import get_column_letter
    wb = _open_any(path, rich=True)
    print(f"  --- {label}: {path.name} ---")
    n = 0
    for sn in wb.sheetnames:
        ws = wb[sn]
        for row in ws.iter_rows():
            for c in row:
                if get_column_letter(c.column) not in ("U", "V"):
                    continue
                v = c.value
                txt = (str(v) if not isinstance(v, CellRichText)
                       else ''.join(p.text if isinstance(p, TextBlock) else str(p) for p in v))
                if not txt or substr not in txt:
                    continue
                cb = bool(c.font and c.font.bold)
                n += 1
                print(f"    {sn}!{get_column_letter(c.column)}{c.row}  cell.font.bold={cb}  "
                      f"type={'RICH' if isinstance(v, CellRichText) else 'plain'}")
                if isinstance(v, CellRichText):
                    for p in v:
                        if isinstance(p, TextBlock):
                            print(f"        run b={getattr(p.font,'b',None)!r}  {p.text[:40]!r}")
                        else:
                            print(f"        run (plain)  {str(p)[:40]!r}")
    if n == 0:
        print(f"    （冇 U/V cell 含 {substr!r}）")


def _uv_patterns(path: Path):
    """{plain_text: Counter( bold_pattern )}，bold_pattern = 每個非空白 run 嘅
    effective bold 組成嘅 tuple。同一文字可以有多個 cell（多個 pattern）→ 用 Counter
    做 multiset，先唔會被行重排／同文字唔同粗體嘅情況呃到（假 mismatch）。"""
    from collections import Counter, defaultdict
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.utils import get_column_letter
    wb = _open_any(path, rich=True)
    d = defaultdict(Counter)
    for sn in wb.sheetnames:
        ws = wb[sn]
        for row in ws.iter_rows():
            for c in row:
                if get_column_letter(c.column) not in ("U", "V"):
                    continue
                v = c.value
                if not isinstance(v, CellRichText):
                    continue
                cb = bool(c.font and c.font.bold)
                pat = []
                for p in v:
                    if isinstance(p, TextBlock):
                        b = getattr(p.font, 'b', None)
                        eff = bool(b) if b is not None else cb
                        txt = p.text
                    else:
                        eff, txt = cb, str(p)
                    if txt.strip():
                        pat.append(eff)
                d[str(v)][tuple(pat)] += 1
    return d


def _uv_diff(src_path: Path, out_path: Path):
    """比較 source vs output 每個 U/V 文字嘅粗體 pattern MULTISET（唔理行位置）。
    只有 output 嘅 pattern 集合同 source 唔同先算真錯判。"""
    print("== U/V bold pattern 對比（multiset，唔理行位置）：只列真唔一致 ==")
    sp = _uv_patterns(src_path)
    op = _uv_patterns(out_path)
    mism = 0
    for text, s_pats in sp.items():
        o_pats = op.get(text)
        if o_pats is None:
            print(f"  ⚠ output 冇呢個文字嘅 U/V cell: {text[:36]!r}")
            mism += 1
            continue
        if s_pats != o_pats:
            mism += 1
            print(f"  ✗ {text[:30]!r}")
            print(f"      source patterns: {[list(map(int, p)) for p in s_pats]}")
            print(f"      output patterns: {[list(map(int, p)) for p in o_pats]}")
    if mism == 0:
        print("  ✓ 所有 U/V 文字嘅粗體 pattern 集合，source 同 output 完全一致（冇真錯判）。")
    else:
        print(f"  共 {mism} 個文字真唔一致。")
    print()


def _bold_count(path: Path):
    """開一個 xlsx（可能加密），數 cell-level bold + rich-run bold 幾多，抽樣。"""
    import openpyxl
    try:
        wb = openpyxl.load_workbook(path, data_only=True, rich_text=True)
    except Exception:
        import msoffcrypto
        buf = io.BytesIO()
        with open(path, "rb") as f:
            off = msoffcrypto.OfficeFile(f)
            off.load_key(password=A.PASSWORD)
            off.decrypt(buf)
        buf.seek(0)
        wb = openpyxl.load_workbook(buf, data_only=True, rich_text=True)
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    cell_bold = run_bold = 0
    samples = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        for row in ws.iter_rows():
            for c in row:
                try:
                    if c.font and c.font.bold:
                        cell_bold += 1
                        if len(samples) < 6:
                            samples.append((sn, c.coordinate, 'cell', str(c.value)[:25]))
                except Exception:
                    pass
                if isinstance(c.value, CellRichText):
                    for p in c.value:
                        if isinstance(p, TextBlock) and p.font is not None and getattr(p.font, 'b', None):
                            run_bold += 1
                            if len(samples) < 12:
                                samples.append((sn, c.coordinate, 'run', p.text[:20]))
                            break
    return cell_bold, run_bold, samples


def main():
    if len(sys.argv) < 2:
        print("俾一個 source xlsx 路徑（第二個參數可俾 output xlsx 對比 bold）")
        return
    path = Path(sys.argv[1])
    if not path.exists():
        print("✗ 唔存在:", path)
        return

    # 若有第二個參數 → 直接對比 source vs output 嘅 bold 數 + U/V 逐 run
    if len(sys.argv) >= 3:
        out = Path(sys.argv[2])
        print("== BOLD 對比 source vs output ==")
        for tag, p in [("SOURCE", path), ("OUTPUT", out)]:
            if not p.exists():
                print(f"  {tag}: ✗ 唔存在 {p}")
                continue
            cb, rb, sm = _bold_count(p)
            print(f"  {tag} {p.name}: cell-level-bold={cb}  rich-run-bold={rb}")
        if out.exists():
            print()
            _uv_diff(path, out)      # 只報唔一致嘅格 —— 直接揪出錯判
        # 第 3 個參數 = 搜尋字串：dump 兩邊含此字串嘅 U/V cell 全部原始細節
        if len(sys.argv) >= 4:
            sub = sys.argv[3]
            print(f"\n== 定點搜尋 U/V cell 含 {sub!r} ==")
            _find_uv(path, sub, "原來 SOURCE")
            if out.exists():
                _find_uv(out, sub, "新出 OUTPUT")
            print()
        print("\n== U/V 欄逐 run 對比（原來 vs 新出）==")
        _dump_uv(path, "原來 SOURCE")
        if out.exists():
            _dump_uv(out, "新出 OUTPUT")
        print()

    wb, rl = _load(path)

    print(f"# {path.name}")
    print(f"[A] rich lookup entries（multi-run/有格式嘅 shared string）: {len(rl)}")
    with_bold = [(k, runs) for k, runs in rl.items()
                 if any(b is True for _t, b, _i, _sz, _fn in runs)]
    print(f"    當中有 run-level <b/>=True 嘅 entry: {len(with_bold)}")
    for k, runs in with_bold[:6]:
        print(f"    key={k[:40]!r}")
        for t, b, i, sz, fn in runs[:8]:
            print(f"        run b={b} i={i} sz={sz} font={fn}  {t[:40]!r}")
    if not with_bold:
        print("    ⚠ 冇任何 entry 有 run-level <b/>=True → U/V 嘅『粗體』唔係靠 <b/> run，"
              "可能係 cell-level bold、或者用咗其他方式（要再睇 raw XML）。")
    if not rl:
        print("    ⚠⚠ lookup 係空！代表 sharedStrings.xml 冇 <r> run，或者讀唔到 →"
              " 所有 run-level bold 一定會冇。")

    # [A2] 專睇 U/V 欄嘅 cell（畢馬威關注事項 / 承批公司的反饋意見）
    print("\n[A2] U/V 欄 cell 逐 run 拆解（頭幾個有內容嘅）:")
    from openpyxl.utils import get_column_letter
    shown_uv = 0
    for sn in wb.sheetnames:
        ws = wb[sn]
        for row in ws.iter_rows():
            for c in row:
                col = get_column_letter(c.column)
                if col in ("U", "V") and isinstance(c.value, str) and c.value.strip():
                    runs = rl.get(c.value)
                    src_bold = bool(c.font and c.font.bold)
                    if runs:
                        print(f"    {sn}!{col}{c.row}: rich, cell-bold={src_bold}")
                        for t, b, i, sz, fn in runs[:6]:
                            print(f"        run b={b}  {t[:45]!r}")
                    else:
                        print(f"    {sn}!{col}{c.row}: NON-rich（唔喺lookup）, cell-bold={src_bold}"
                              f"  {c.value[:45]!r}")
                    shown_uv += 1
                    if shown_uv >= 8:
                        break
            if shown_uv >= 8:
                break
        if shown_uv >= 8:
            break

    print("\n[B] 逐 sheet 掃 cell（只計 value 係文字嘅）:")
    tot_hit = tot_miss = tot_cellbold = 0
    miss_samples, cellbold_samples, hit_samples = [], [], []
    for sn in wb.sheetnames:
        ws = wb[sn]
        hit = miss = cbold = 0
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, str) and v.strip():
                    if v in rl:
                        hit += 1
                        if len(hit_samples) < 4:
                            hit_samples.append((sn, c.coordinate, v[:30]))
                    else:
                        miss += 1
                        if len(miss_samples) < 8:
                            miss_samples.append((sn, c.coordinate, v[:40]))
                try:
                    if c.font and c.font.bold:
                        cbold += 1
                        if len(cellbold_samples) < 6:
                            cellbold_samples.append((sn, c.coordinate, str(v)[:30]))
                except Exception:
                    pass
        tot_hit += hit; tot_miss += miss; tot_cellbold += cbold
        print(f"    · {sn!r}: rich命中={hit}  未命中(str)={miss}  cell-level-bold={cbold}")

    print(f"\n  總: rich命中={tot_hit}  未命中={tot_miss}  cell-level-bold={tot_cellbold}")
    if hit_samples:
        print("  rich 命中樣本:")
        for sn, co, t in hit_samples:
            print(f"     {sn}!{co}: {t!r}")
    if miss_samples:
        print("  未命中(str, 唔喺 lookup)樣本 —— 若呢啲本應 rich，即係 key mismatch:")
        for sn, co, t in miss_samples:
            print(f"     {sn}!{co}: {t!r}")
    if cellbold_samples:
        print("  cell-level bold 樣本（呢啲應該行 non-rich path、copy src font）:")
        for sn, co, t in cellbold_samples:
            print(f"     {sn}!{co}: {t!r}")

    print("\n[C] 判讀:")
    if not rl:
        print("  → lookup 空：sharedStrings 冇 run 或讀唔到。run-level bold 冇得救於呢條 path。")
    elif tot_hit == 0:
        print("  → lookup 有嘢但 0 命中：key 對唔上 cell.value（openpyxl 讀出嚟嘅文字同"
              " raw run 串接唔一致）。要對齊 key。")
    else:
        print(f"  → lookup 有命中（{tot_hit}）。若輸出仍冇 bold，問題喺 _rich_val 之後"
              "（apply/save/encrypt）或者 base_grid 覆蓋。")
    if tot_cellbold:
        print(f"  → 有 {tot_cellbold} 格 cell-level bold；呢啲行 non-rich path，"
              "輸出理應有 bold。若都冇，即係 _apply_text_style 個 copy 冇生效。")


if __name__ == "__main__":
    main()
