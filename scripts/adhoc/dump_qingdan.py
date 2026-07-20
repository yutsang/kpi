#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
dump_qingdan.py — 望實 data\\投資項目清單 6 個清單檔（+ 任何指定 xlsx）嘅結構，
俾 align_to_header 加「項目編號」欄 + cross-check 用。

用法：
    # 掃 data\\投資項目清單\\ 下面全部 xlsx
    python scripts\\adhoc\\dump_qingdan.py "ad-hoc\\workspace\\data\\投資項目清單"
    # 或直接指定一個/多個檔（例如新 SJM 檔）
    python scripts\\adhoc\\dump_qingdan.py "ad-hoc\\workspace\\data\\3.SJM-投資計劃執行情況表二（旅遊局）.xlsx"

每個檔每個 sheet 印：
  - sheet 名、維度
  - 頭 ~3 行（當表頭）逐欄文字
  - 之後 ~6 行資料（逐欄），等睇到「項目編號」欄喺邊、咩格式（76 / CE001 / 項目3 / B11.1…）
加密檔自動用 dicj_kpmg 解。
"""
import io
import sys
from pathlib import Path

PASSWORD = "dicj_kpmg"


def _open(path: Path):
    import openpyxl
    try:
        return openpyxl.load_workbook(path, data_only=True)
    except Exception:
        import msoffcrypto
        buf = io.BytesIO()
        with open(path, "rb") as f:
            off = msoffcrypto.OfficeFile(f)
            off.load_key(password=PASSWORD)
            off.decrypt(buf)
        buf.seek(0)
        return openpyxl.load_workbook(buf, data_only=True)


def _c(v):
    if v is None:
        return ""
    s = str(v).replace("\n", "⏎")
    return s[:24]


def dump_file(path: Path, hdr_rows=3, data_rows=6):
    from openpyxl.utils import get_column_letter
    print(f"\n{'='*74}\n# {path.name}")
    try:
        wb = _open(path)
    except Exception as e:
        print(f"  ✗ 開唔到: {type(e).__name__}: {e}")
        return
    for sn in wb.sheetnames:
        ws = wb[sn]
        mc = min(ws.max_column or 0, 20)
        mr = ws.max_row or 0
        print(f"\n  ── sheet {sn!r}  ({mr} 行 × {ws.max_column} 欄，印頭 {mc} 欄) ──")
        # 表頭幾行
        for r in range(1, min(hdr_rows, mr) + 1):
            cells = [f"{get_column_letter(c)}={_c(ws.cell(r, c).value)}"
                     for c in range(1, mc + 1) if ws.cell(r, c).value is not None]
            print(f"    [hdr r{r}] " + " | ".join(cells))
        # 揾第一行有嘢嘅資料行做起點（跳過空行）
        start = None
        for r in range(1, mr + 1):
            if any(ws.cell(r, c).value not in (None, "") for c in range(1, mc + 1)):
                start = r
                break
        start = (start or 1)
        print(f"    ── 資料樣本（由 r{start} 起 {data_rows} 行）──")
        shown = 0
        for r in range(start, mr + 1):
            vals = [ws.cell(r, c).value for c in range(1, mc + 1)]
            if all(v in (None, "") for v in vals):
                continue
            cells = [f"{get_column_letter(c+1)}={_c(v)}" for c, v in enumerate(vals) if v not in (None, "")]
            print(f"    [r{r}] " + " | ".join(cells))
            shown += 1
            if shown >= data_rows:
                break
    wb.close()


def main():
    args = sys.argv[1:]
    if not args:
        print("俾一個資料夾或 xlsx 路徑")
        return
    targets = []
    for a in args:
        p = Path(a)
        if p.is_dir():
            targets += sorted(x for x in p.rglob("*")
                              if x.suffix.lower() in (".xlsx", ".xlsm") and not x.name.startswith("~$"))
        elif p.exists():
            targets.append(p)
        else:
            print(f"✗ 唔存在: {p}")
    for t in targets:
        dump_file(t)


if __name__ == "__main__":
    main()
