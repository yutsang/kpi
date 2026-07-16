#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
inspect_headers.py — adhoc：對齊 表頭.xlsx 前，先 inspect source_1 / source_2。

喺 Windows（檔所在嗰部機）跑：
    set PYTHONIOENCODING=utf-8
    python scripts\\adhoc\\inspect_headers.py --root adhoc\\workspace

做乜：
  1. 讀 <root>\\表頭.xlsx → 目標欄（standard header）。
  2. 行勻 <root>\\source_1 所有 Excel/CSV（**跳過任何叫 ss 嘅 subfolder**），
     開密碼檔用 dicj_kpmg，偵測 header 行，同 表頭 對欄（缺/多/次序）。
  3. source_2 一齊列（結構 + 欄），佢係「有問題/要挑出嚟」嗰批。
  4. 出：資料夾樹 + 每檔欄 diff + summary。寫 <root>\\_inspect_report.txt，同時印出嚟。

report 貼返嚟俾我（或者放 Mac results/），我再設計對齊。純讀，唔改任何檔。
"""
from __future__ import annotations

import argparse
import io
import sys
from pathlib import Path

import openpyxl

PASSWORD = "dicj_kpmg"
EXCEL_EXT = {".xlsx", ".xlsm", ".xls"}
DATA_EXT = EXCEL_EXT | {".csv"}
SKIP_DIRS = {"ss"}          # user：source_1 內 ss subfolder 唔理
MAX_SCAN_ROWS = 15          # 掃頭幾行揾 header 行


def _norm(x) -> str:
    return "" if x is None else str(x).strip().replace("\n", " ").replace("\r", " ")


def load_wb(path: Path):
    """開 workbook（read-only）。加密就用 dicj_kpmg 解。回 (wb, encrypted_bool)。"""
    try:
        return openpyxl.load_workbook(path, read_only=True, data_only=True), False
    except Exception:
        try:
            import msoffcrypto
        except ImportError:
            raise RuntimeError("疑似加密檔，但冇裝 msoffcrypto → 喺 Windows 跑： pip install msoffcrypto-tool")
        buf = io.BytesIO()
        with open(path, "rb") as f:
            off = msoffcrypto.OfficeFile(f)
            off.load_key(password=PASSWORD)
            off.decrypt(buf)
        buf.seek(0)
        return openpyxl.load_workbook(buf, read_only=True, data_only=True), True


def detect_header(ws, target: list[str]) -> tuple[int, list[str]]:
    """掃前幾行，揀同 target 重疊最多嗰行做 header；都唔中就用第一行非空。"""
    tset = {_norm(t) for t in target if _norm(t)}
    best_i, best_cols, best_hit = 0, [], -1
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=MAX_SCAN_ROWS, values_only=True)):
        cols = [_norm(c) for c in row]
        rows.append(cols)
        hit = len(tset & {c for c in cols if c}) if tset else sum(1 for c in cols if c)
        if hit > best_hit:
            best_hit, best_i, best_cols = hit, i, cols
    # trim 尾部空欄
    while best_cols and best_cols[-1] == "":
        best_cols.pop()
    return best_i + 1, best_cols


def read_csv_header(path: Path, target: list[str]) -> tuple[int, list[str]]:
    import csv
    for enc in ("utf-8-sig", "utf-8", "cp950", "gbk", "latin-1"):
        try:
            with open(path, "r", encoding=enc, newline="") as f:
                rows = [r for _, r in zip(range(MAX_SCAN_ROWS), csv.reader(f))]
            break
        except Exception:
            rows = []
    tset = {_norm(t) for t in target if _norm(t)}
    best_i, best_cols, best_hit = 0, [], -1
    for i, r in enumerate(rows):
        cols = [_norm(c) for c in r]
        hit = len(tset & {c for c in cols if c}) if tset else sum(1 for c in cols if c)
        if hit > best_hit:
            best_hit, best_i, best_cols = hit, i, cols
    while best_cols and best_cols[-1] == "":
        best_cols.pop()
    return best_i + 1, best_cols


def diff_cols(target: list[str], got: list[str]) -> dict:
    t = [c for c in (_norm(x) for x in target) if c]
    g = [c for c in (_norm(x) for x in got) if c]
    ts, gs = set(t), set(g)
    return {
        "missing": [c for c in t if c not in gs],      # 表頭有、檔冇
        "extra": [c for c in g if c not in ts],        # 檔有、表頭冇
        "exact": t == g,                                # 一模一樣（含次序）
        "same_set": ts == gs,                           # 齊全但次序可能唔同
    }


def walk_files(root: Path):
    """yield 檔，跳過 SKIP_DIRS。"""
    for p in sorted(root.rglob("*")):
        if p.is_dir():
            continue
        if any(part.lower() in SKIP_DIRS for part in p.relative_to(root).parts):
            continue
        if p.suffix.lower() in DATA_EXT and not p.name.startswith("~$"):
            yield p


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--root", default="adhoc/workspace", help="workspace 根（含 表頭.xlsx / source_1 / source_2）")
    ap.add_argument("--header-file", default="表頭.xlsx")
    a = ap.parse_args()
    root = Path(a.root)
    out = []

    def log(s=""):
        out.append(s)
        print(s)

    if not root.exists():
        sys.exit(f"✗ root 唔存在: {root.resolve()}")

    # 1) 表頭
    hf = root / a.header_file
    target: list[str] = []
    log(f"# workspace: {root.resolve()}")
    if hf.exists():
        wb, enc = load_wb(hf)
        ws = wb.active
        hrow, target = detect_header(ws, [])
        log(f"\n## 表頭.xlsx  (sheet={ws.title!r}, header 喺第 {hrow} 行, {'加密' if enc else '無密碼'})")
        log(f"目標欄 ({len(target)}):")
        for i, c in enumerate(target, 1):
            log(f"  {i:>2}. {c}")
        wb.close()
    else:
        log(f"\n⚠ 揾唔到 {hf} — 冇目標欄，下面只列各檔欄")

    # 2/3) source_1 / source_2
    for src in ("source_1", "source_2"):
        base = root / src
        log(f"\n{'='*70}\n## {src}")
        if not base.exists():
            log(f"  (冇 {base})")
            continue
        files = list(walk_files(base))
        # 資料夾樹
        log(f"  檔數（跳過 {SKIP_DIRS}）: {len(files)}")
        dirs = sorted({p.parent.relative_to(base).as_posix() for p in files})
        log("  子資料夾:")
        for d in dirs:
            n = sum(1 for p in files if p.parent.relative_to(base).as_posix() == d)
            log(f"    {d if d != '.' else '(根)'}/   [{n} 檔]")
        # 每檔欄 + diff
        n_exact = n_sameset = n_diff = n_err = 0
        for p in files:
            rel = p.relative_to(base).as_posix()
            try:
                if p.suffix.lower() == ".csv":
                    hrow, cols = read_csv_header(p, target)
                    sheets, enc = ["(csv)"], False
                else:
                    wb, enc = load_wb(p)
                    sheets = wb.sheetnames
                    hrow, cols = detect_header(wb.active, target)
                    wb.close()
                d = diff_cols(target, cols) if target else {"missing": [], "extra": [], "exact": None, "same_set": None}
                tag = "EXACT ✓" if d["exact"] else ("欄齊但次序唔同" if d["same_set"] else "有 diff")
                if d["exact"]:
                    n_exact += 1
                elif d["same_set"]:
                    n_sameset += 1
                else:
                    n_diff += 1
                log(f"\n  ▸ {rel}   [{tag}]  ({'加密' if enc else '無密碼'}, header 第{hrow}行, {len(cols)} 欄)")
                if len(sheets) > 1:
                    log(f"      sheets: {sheets}")
                if target and not d["exact"]:
                    if d["missing"]:
                        log(f"      缺(表頭有檔冇): {d['missing']}")
                    if d["extra"]:
                        log(f"      多(檔有表頭冇): {d['extra']}")
                if not target:
                    log(f"      欄: {cols}")
            except Exception as e:
                n_err += 1
                log(f"\n  ▸ {rel}   [讀取失敗] {type(e).__name__}: {e}")
        log(f"\n  -- {src} summary: EXACT={n_exact}  欄齊次序異={n_sameset}  有diff={n_diff}  失敗={n_err}")

    rp = root / "_inspect_report.txt"
    rp.write_text("\n".join(out), encoding="utf-8")
    print(f"\n✓ 報告寫入 {rp.resolve()}（貼返嚟俾我 / 放 Mac results/）")


if __name__ == "__main__":
    main()
