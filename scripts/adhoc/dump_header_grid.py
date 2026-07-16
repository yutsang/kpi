#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
dump_header_grid.py — 合併感知咁吐幾個檔嘅「表頭層級 + 行結構」。

點解要合併感知：表格唔係一行一項目 —— 一個項目佔幾行（行數唔固定）、識別欄(項目名/金額/範疇)
係跨行垂直合併，明細（如各類調整）散落項目下面幾行；仲有水平合併嘅分組表頭。用 read_only 當平面
grid 會睇漏。所以用非 read-only 開，攞返 merged_cells，列出：
  A. 表頭層級（逐欄 row1..N 堆疊值）
  B. 合併範圍：垂直合併(跨行=項目級 key) / 水平合併(跨欄=分組表頭)
  C. 逐行 anchor 值（睇一個項目佔幾行、邊幾欄 merge、邊幾欄逐行變）

Windows 跑：
    set PYTHONIOENCODING=utf-8
    python scripts\\adhoc\\dump_header_grid.py --root ad-hoc\\workspace
    python scripts\\adhoc\\dump_header_grid.py --root ad-hoc\\workspace --pair 旅遊局 SJM

純讀，唔改檔。輸出寫 <root>\\_header_grid.txt。
"""
from __future__ import annotations

import argparse
import io
from pathlib import Path

import openpyxl

PASSWORD = "dicj_kpmg"
DEFAULT_FILES = [
    "表頭.xlsx",
    r"source_1\其他範疇\MGM-投資計劃執行情況表二（其他範疇）.xlsx",
    r"source_1\博監局\SJM-投資計劃執行情況表二（博監局）.xlsx",
    r"source_2\2025年年度投資計劃執行情況審查專項工作關注事項 - 承批公司回覆-跨司工作组反馈意见-待进一步確認事項.xlsx",
]


def load_wb(path: Path):
    """非 read-only（要攞 merged_cells）；加密就 dicj_kpmg 解。"""
    try:
        return openpyxl.load_workbook(path, data_only=True)
    except Exception:
        try:
            import msoffcrypto
        except ImportError:
            raise RuntimeError("加密檔要 pip install msoffcrypto-tool")
        buf = io.BytesIO()
        with open(path, "rb") as f:
            off = msoffcrypto.OfficeFile(f)
            off.load_key(password=PASSWORD)
            off.decrypt(buf)
        buf.seek(0)
        return openpyxl.load_workbook(buf, data_only=True)


def col_letter(i: int) -> str:
    s = ""
    while i > 0:
        i, r = divmod(i - 1, 26)
        s = chr(65 + r) + s
    return s


def _norm(x) -> str:
    return "" if x is None else str(x).strip().replace("\n", " / ").replace("\r", " ")


def _clip(s: str, n: int = 28) -> str:
    return s if len(s) <= n else s[: n - 1] + "…"


def dump_header_stack(ws, hdr_rows: int, out) -> None:
    """A. 逐欄印 row1..hdr_rows 堆疊值（睇真正欄名同分組）。"""
    grid = [[_norm(c) for c in row]
            for row in ws.iter_rows(min_row=1, max_row=hdr_rows, values_only=True)]
    ncols = max((len(r) for r in grid), default=0)
    while ncols > 0 and all(len(r) < ncols or r[ncols - 1] == "" for r in grid):
        ncols -= 1
    out(f"    ── A. 表頭層級 (row1..{hdr_rows}, {ncols} 欄) ──")
    for c in range(ncols):
        stack = [grid[r][c] for r in range(len(grid)) if c < len(grid[r]) and grid[r][c]]
        if stack:
            out(f"      [{c+1:>2} {col_letter(c+1):>2}] " + "  ↑  ".join(stack))


def dump_structure(ws, nrows: int, out) -> None:
    """B+C. 合併範圍 + 逐行 anchor 值。"""
    maxrow = ws.max_row or 0
    maxcol = ws.max_column or 0
    win = min(nrows, maxrow)
    out(f"    dims: rows={maxrow} cols={maxcol}  (結構顯示前 {win} 行)")

    vmerges, hmerges = [], []
    top_left, covered = {}, {}
    for m in ws.merged_cells.ranges:
        if m.min_row > win:
            continue
        val = _norm(ws.cell(m.min_row, m.min_col).value)
        top_left[(m.min_row, m.min_col)] = (m.max_row - m.min_row + 1, m.max_col - m.min_col + 1)
        for rr in range(m.min_row, m.max_row + 1):
            for cc in range(m.min_col, m.max_col + 1):
                if (rr, cc) != (m.min_row, m.min_col):
                    covered[(rr, cc)] = True
        if m.max_row > m.min_row:
            vmerges.append((m.min_col, m.min_row, m.max_col, m.max_row, val))
        elif m.max_col > m.min_col:
            hmerges.append((m.min_row, m.min_col, m.max_col, val))

    out("    ── B1. 垂直合併 (跨行 = 項目級 key；一個 merge = 一個項目佔幾行) ──")
    if vmerges:
        for c, r0, c1, r1, v in sorted(vmerges):
            cs = f"×{c1 - c + 1}欄" if c1 > c else ""
            out(f"      [{col_letter(c)}{r0}:{col_letter(c1)}{r1}] {r1 - r0 + 1}行{cs} = {_clip(v, 40)}")
    else:
        out(f"      (前 {win} 行冇垂直合併)")

    out("    ── B2. 水平合併 (跨欄 = 分組表頭) ──")
    if hmerges:
        for r, c0, c1, v in sorted(hmerges):
            out(f"      [{col_letter(c0)}{r}:{col_letter(c1)}{r}] = {_clip(v, 40)}")
    else:
        out(f"      (前 {win} 行冇水平合併)")

    out("    ── C. 逐行 anchor 值 (只印有值格；(合R×C)=合併起點) ──")
    for r in range(1, win + 1):
        cells = []
        for c in range(1, maxcol + 1):
            v = _norm(ws.cell(r, c).value)
            if not v:
                continue
            mark = ""
            if (r, c) in top_left and top_left[(r, c)] != (1, 1):
                sp = top_left[(r, c)]
                mark = f"(合{sp[0]}×{sp[1]})"
            cells.append(f"{c}:{_clip(v)}{mark}")
        out(f"      R{r:>2}: " + (" | ".join(cells) if cells else "(空)"))


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--root", default="ad-hoc/workspace")
    ap.add_argument("--files", nargs="*", default=None, help="相對 root 嘅檔（預設 4 個代表）")
    ap.add_argument("--hdr-rows", type=int, default=8, help="表頭層級睇頭幾行")
    ap.add_argument("--struct-rows", type=int, default=30, help="行結構睇頭幾行")
    ap.add_argument("--all-sheets", action="store_true", help="每個 sheet 都 dump（預設只第一個）")
    ap.add_argument("--pair", nargs=2, metavar=("範疇", "公司"),
                    help="自動揾 source_1/source_2 內路徑含呢兩個關鍵字嘅檔一齊 dump（睇 join key）")
    a = ap.parse_args()
    root = Path(a.root)
    files = list(a.files) if a.files else list(DEFAULT_FILES)
    if a.pair:
        kw_scope, kw_company = a.pair
        for src in ("source_1", "source_2"):
            base = root / src
            if not base.exists():
                continue
            for p in sorted(base.rglob("*")):
                if p.is_dir() or p.name.startswith("~$"):
                    continue
                if p.suffix.lower() not in (".xlsx", ".xlsm", ".xls"):
                    continue
                relstr = p.relative_to(root).as_posix()
                if kw_scope in relstr and kw_company in relstr:
                    files.append(relstr)

    lines: list[str] = []

    def out(s=""):
        lines.append(s)
        print(s)

    for rel in files:
        p = root / rel
        out("\n" + "=" * 78)
        out(f"# {rel}")
        if not p.exists():
            out(f"  ✗ 唔存在: {p}")
            continue
        try:
            wb = load_wb(p)
        except Exception as e:
            out(f"  ✗ 開唔到: {type(e).__name__}: {e}")
            continue
        out(f"  sheets: {wb.sheetnames}")
        targets = wb.sheetnames if a.all_sheets else [wb.sheetnames[0]]
        for sn in targets:
            out(f"\n  ── sheet: {sn!r} ──")
            ws = wb[sn]
            dump_header_stack(ws, a.hdr_rows, out)
            dump_structure(ws, a.struct_rows, out)
        wb.close()

    rp = root / "_header_grid.txt"
    rp.write_text("\n".join(lines), encoding="utf-8")
    print(f"\n✓ 寫入 {rp.resolve()}（貼返嚟俾我）")


if __name__ == "__main__":
    main()
