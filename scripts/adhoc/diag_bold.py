# -*- coding: utf-8 -*-
"""
Diagnostic: inspect bold formatting in source xlsx vs aligned output.

Usage (Windows):
  python scripts/adhoc/diag_bold.py <source_xlsx> [<aligned_xlsx>]

Prints for each non-empty cell in sheet 1:
  row | col | value[:35] | cell.font.bold | in_rich_lookup | rt_runs_bold
"""
import sys, io, zipfile, xml.etree.ElementTree as ET
from pathlib import Path
from copy import copy

PASSWORD = "dicj_kpmg"
NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'


def decrypt(path: Path) -> io.BytesIO:
    import msoffcrypto
    buf = io.BytesIO()
    with open(path, "rb") as f:
        off = msoffcrypto.OfficeFile(f)
        off.load_key(password=PASSWORD)
        off.decrypt(buf)
    buf.seek(0)
    return buf


def load(path: Path):
    import openpyxl
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        src = path
    except Exception:
        src = decrypt(path)
        wb = openpyxl.load_workbook(src, data_only=True)
    return wb, src


def build_rich_lookup(src):
    """Returns {plain_text: [(run_text, b_in_rpr)]} for si entries with runs."""
    if isinstance(src, io.BytesIO):
        src.seek(0)
    result = {}
    try:
        with zipfile.ZipFile(src, 'r') as zf:
            if 'xl/sharedStrings.xml' not in zf.namelist():
                return {}
            root = ET.parse(zf.open('xl/sharedStrings.xml')).getroot()
    except Exception as e:
        return {"_error": str(e)}

    for si in root.findall(f'{{{NS}}}si'):
        rs = si.findall(f'{{{NS}}}r')
        if not rs:
            continue
        runs = []
        plain = ""
        for r in rs:
            t_el = r.find(f'{{{NS}}}t')
            txt = (t_el.text or '') if t_el is not None else ''
            plain += txt
            rpr = r.find(f'{{{NS}}}rPr')
            b_el = rpr.find(f'{{{NS}}}b') if rpr is not None else None
            if b_el is not None:
                b_val = b_el.get('val', '1') != '0'
            else:
                b_val = None  # not specified in run
            runs.append((txt[:20], b_val))
        result[plain] = runs
    return result


def inspect(path: Path, label: str):
    print(f"\n{'='*70}")
    print(f"  {label}: {path.name}")
    print(f"{'='*70}")

    wb, src = load(path)
    ws = wb.worksheets[0]
    rich = build_rich_lookup(src)

    print(f"  Sheet: {ws.title} | Rich lookup entries: {len(rich)}")
    print(f"  {'R':>3} {'C':>3}  {'font.bold':>10}  {'in_lookup':>9}  {'run b_vals':20}  value[:40]")
    print(f"  {'-'*3} {'-'*3}  {'-'*10}  {'-'*9}  {'-'*20}  {'-'*40}")

    shown = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            v = str(cell.value)
            fb = cell.font.bold if cell.font else '?'
            in_lk = v in rich
            run_b = ""
            if in_lk:
                runs = rich[v]
                run_b = str([b for _, b in runs])[:20]
            print(f"  {cell.row:>3} {cell.column:>3}  {str(fb):>10}  {str(in_lk):>9}  {run_b:<20}  {v[:40]!r}")
            shown += 1
            if shown >= 60:
                print("  ... (truncated at 60 cells)")
                return

    # summary stats
    all_bold = [c.font.bold for row in ws.iter_rows() for c in row
                if c.value is not None and c.font]
    from collections import Counter
    print(f"\n  font.bold summary: {Counter(all_bold)}")


def inspect_rich(src_path: Path, out_path: Path | None):
    """專門睇 rich cells：source 嘅 rich lookup entries 對應哪些 cells，
    佢哋嘅 src.font.bold 係幾多，output 對應 cells 嘅 font.bold 係幾多。"""
    wb_s, src = load(src_path)
    ws_s = wb_s.worksheets[0]
    rich = build_rich_lookup(src)

    # 讀 output
    wb_o = None
    ws_o = None
    if out_path and out_path.exists():
        try:
            import openpyxl
            wb_o = openpyxl.load_workbook(out_path, data_only=True)
            ws_o = wb_o.worksheets[0]
        except Exception as e:
            print(f"  [WARNING] Cannot load output file: {e}")
            ws_o = None

    print(f"\n{'='*70}")
    print(f"  RICH CELL DETAILS — SOURCE: {src_path.name}")
    if ws_o:
        print(f"  vs OUTPUT: {out_path.name}")
    print(f"  Rich lookup has {len(rich)} entries")
    print(f"{'='*70}")

    # 先印 lookup entries 本身（頭 10 個）
    print("\n  [A] First 10 sharedString entries with runs:")
    print(f"  {'plain_text[:40]':44}  runs [(txt[:15], b_in_rpr)]")
    for i, (k, v) in enumerate(list(rich.items())[:10]):
        print(f"  {k[:44]!r:46}  {v}")

    # 再印 source cells that are in rich lookup
    print(f"\n  [B] Source cells in rich lookup  (src.font.bold | run b_vals | value[:35])")
    print(f"  {'R':>4} {'C':>3}  {'src.bold':>9}  {'run b_vals':24}  value[:35]")
    found = 0
    for row in ws_s.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            v = str(cell.value)
            if v not in rich:
                continue
            runs = rich[v]
            run_b = str([b for _, b in runs])[:24]
            sb = cell.font.bold if cell.font else '?'
            # check output
            ob = '—'
            if ws_o:
                try:
                    oc = ws_o.cell(cell.row, cell.column)
                    ob = oc.font.bold if oc.font else '?'
                except Exception:
                    ob = 'err'
            print(f"  {cell.row:>4} {cell.column:>3}  {str(sb):>9}  {run_b:<24}  {v[:35]!r}  → out.bold={ob}")
            found += 1
            if found >= 30:
                print("  ... truncated at 30 rich cells")
                break
        if found >= 30:
            break

    if found == 0:
        print("  (no cells matched — cell.value might not match lookup keys)")
        print("  Showing first 5 lookup keys vs first 5 cell values in row 9+:")
        keys = list(rich.keys())[:5]
        print("  LOOKUP KEYS:", [k[:30] for k in keys])
        vals = []
        for row in ws_s.iter_rows(min_row=9):
            for cell in row:
                if cell.value is not None:
                    vals.append(str(cell.value)[:30])
            if len(vals) >= 10:
                break
        print("  CELL VALUES (row 9+):", vals[:10])

    # font.bold summary for source
    all_bold = [c.font.bold for row in ws_s.iter_rows() for c in row
                if c.value is not None and c.font]
    from collections import Counter
    print(f"\n  [C] SOURCE font.bold summary (non-empty cells): {Counter(all_bold)}")
    if ws_o:
        all_bold_o = [c.font.bold for row in ws_o.iter_rows() for c in row
                      if c.value is not None and c.font]
        print(f"  [C] OUTPUT font.bold summary (non-empty cells): {Counter(all_bold_o)}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python diag_bold.py <source.xlsx> [<aligned.xlsx>]")
        sys.exit(1)

    src_path = Path(sys.argv[1])
    out_path = Path(sys.argv[2]) if len(sys.argv) >= 3 else None
    inspect_rich(src_path, out_path)
