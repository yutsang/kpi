#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
dump_qingdan_codes.py — 專睇『投資項目清單』每個檔嘅「項目編號 / 項目序號」欄真正格式，
俾 align 把輸出「項目編號」欄改成跟清單格式（唔再淨抽表二序號如 SJM「76」）。

每個清單檔每個 sheet：
  1. 揾表頭行（含『項目序號』或『項目編號』或『項目名稱』嘅行）。
  2. 印晒該行**全部**欄（欄letter = 表頭label），等睇清楚有幾多條 code 欄、邊條先係正式編號。
  3. 揾出所有 header 含 序號/編號/名稱 嘅欄，dump **全部資料行**（cap 400）嗰幾欄嘅完整值，
     等睇到每個 entity 真實格式（76 / 076 / IV008 / 項目35 / B11.1 / 前綴…）。

用法（Windows）：
    python scripts\\adhoc\\dump_qingdan_codes.py "ad-hoc\\workspace\\data\\投資項目清單"
    或指定單一檔（例如淨睇 SJM）：
    python scripts\\adhoc\\dump_qingdan_codes.py "…\\3.SJM.2025年度…投资项目清单.xlsx"
加密檔自動用 dicj_kpmg 解。
"""
import io
import re
import sys
from pathlib import Path

PASSWORD = "dicj_kpmg"
_CODE_HINT = re.compile(r"序號|編號|編碼|名稱|項目")


def _open(path: Path):
    import openpyxl
    try:
        return openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception:
        import msoffcrypto
        buf = io.BytesIO()
        with open(path, "rb") as f:
            off = msoffcrypto.OfficeFile(f)
            off.load_key(password=PASSWORD)
            off.decrypt(buf)
        buf.seek(0)
        return openpyxl.load_workbook(buf, data_only=True, read_only=True)


def _s(v):
    return "" if v is None else str(v).replace("\n", "⏎").strip()


def dump_file(path: Path, max_rows=400):
    from openpyxl.utils import get_column_letter
    print(f"\n{'='*78}\n# {path.name}")
    try:
        wb = _open(path)
    except Exception as e:
        print(f"  ✗ 開唔到: {type(e).__name__}: {e}")
        return
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            rows.append(row)
            if i > max_rows:
                break
        if not rows:
            continue
        # 揾表頭行：頭 15 行入面，最多「序號/編號/名稱/項目」關鍵字嗰行
        hdr_r, best = None, -1
        for r in range(min(15, len(rows))):
            hits = sum(1 for v in rows[r] if _CODE_HINT.search(_s(v)))
            if hits > best and hits >= 1:
                best, hdr_r = hits, r
        if hdr_r is None:
            print(f"\n  ── sheet {sn!r}：揾唔到表頭（跳過）──")
            continue
        hdr = rows[hdr_r]
        print(f"\n  ── sheet {sn!r}（表頭 row {hdr_r+1}）全部欄 ──")
        for c, v in enumerate(hdr):
            if _s(v):
                print(f"      {get_column_letter(c+1)} = {_s(v)}")
        # code 欄 = header 含關鍵字
        code_cols = [c for c, v in enumerate(hdr) if _CODE_HINT.search(_s(v))]
        print(f"\n  ── sheet {sn!r}：code/名稱欄逐行值（{get_column_letter(1)}起）──")
        shown = 0
        for r in range(hdr_r + 1, len(rows)):
            cells = [(get_column_letter(c+1), _s(rows[r][c])) for c in code_cols
                     if c < len(rows[r]) and _s(rows[r][c])]
            if not cells:
                continue
            print("      " + " | ".join(f"{L}={v[:40]}" for L, v in cells))
            shown += 1
            if shown >= 60:
                print(f"      …（餘下略，已印 60 行）")
                break
    wb.close()


def main():
    args = sys.argv[1:]
    if not args:
        print("俾一個資料夾或清單 xlsx 路徑")
        return
    targets = []
    for a in args:
        p = Path(a)
        if p.is_dir():
            targets += sorted(x for x in p.rglob("*")
                              if x.suffix.lower() in (".xlsx", ".xlsm") and not x.name.startswith("~$"))
        elif p.exists():
            targets.append(p)
        else:
            print(f"✗ 唔存在: {p}")
    for t in targets:
        dump_file(t)


if __name__ == "__main__":
    main()
