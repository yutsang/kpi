#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
inspect_coverage.py — 一次過survey 全部 inputs（feed + 清單 + 表2），估計報告每頁 coverage。
輸出俾 Claude 判斷：邊啲頁由 inputs 拎到、邊啲要外部（KPI底稿/組織圖/相片）。

用法（KMPG 網內）：
  python scripts\\report\\inspect_coverage.py "tableau_combined_25.csv" ^
      --qingdan "data\\投資項目清單\\3. MGM.…投资项目清单.xlsx" --entity mgm [--biao2 "data\\表2"]
⚠ 輸出係機密（有 database 內容）→ 貼落 results/ 或直接 paste，切勿 commit。
"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
try:
    import pandas as pd
    import openpyxl
except ImportError:
    print("✗ pip install pandas openpyxl"); sys.exit(1)

# 報告每頁對應嘅 data 需求（我睇 scan 整理）→ 用嚟報 coverage
PAGE_NEEDS = [
    ("p1 封面", "entity 全名 + 日期", "code/config"),
    ("p3 縮寫定義", "承批公司/子公司/縮寫清單", "boilerplate（半機密）"),
    ("p4 目錄", "章節+頁碼", "code"),
    ("p5 股權架構圖+主體公司表", "股權/控股/執行主體 + 各主體投資金額%", "❓ 探 清單/feed"),
    ("p6 整體投資支出概況", "主體公司 × 投資金額（博彩/非博彩）", "❓ 探"),
    ("p11 整體執行概況(2欄)", "範疇×計劃/報告/完成率 + 敘述", "✅ feed+清單"),
    ("p15 潛在調整匯總", "7類調整×金額", "✅ feed"),
    ("p19-26 過往期後", "24/23期後 範疇×金額", "✅ feed"),
    ("p28-40 主要發現", "逐調整類型×項目 發現/管理層/跨司", "✅ 清單+表2"),
    ("p42-63 其他信息/單項", "單項審查 + 設施/活動 + 程序匯總", "✅ feed（程序=boilerplate）"),
    ("p73-82 六項KPI", "6個財務KPI 計算+數值", "❓ 探 KPI 欄"),
    ("p93 現場走訪", "capex≥2000萬 項目", "✅ feed"),
    ("p101 藝術品展出清單", "逐件藝術品 + 展出狀態", "❓ 探"),
    ("p103 多功能娛樂補充圖", "相片", "⛔ 相片"),
    ("p84 風險領域/工作範圍", "審查方法論", "boilerplate"),
]


def _probe(cols, *keywords):
    return [c for c in cols if any(k in str(c) for k in keywords)]


def main():
    args = sys.argv[1:]
    entity = qingdan = biao2_dir = None
    for flag in ("--entity", "--qingdan", "--biao2"):
        if flag in args:
            i = args.index(flag); v = args[i + 1]; del args[i:i + 2]
            if flag == "--entity":
                entity = v.lower()
            elif flag == "--qingdan":
                qingdan = v
            else:
                biao2_dir = v
    if not args:
        print("俾 feed csv 路徑（--qingdan 清單 --entity mgm）"); return

    print("=" * 80)
    print("報告每頁 data 需求（我睇 scan 整理）：")
    for pg, need, src in PAGE_NEEDS:
        print(f"  {pg:<26} 需要：{need:<32} 來源：{src}")

    # ── feed ──
    print("\n" + "=" * 80 + "\n# FEED（tableau）")
    df = pd.read_csv(args[0], low_memory=False)
    if entity and "entity" in df.columns:
        df = df[df["entity"].astype(str).str.lower() == entity]
    print(f"  行數:{len(df)}  欄數:{len(df.columns)}")
    print(f"  欄:{list(df.columns)}")
    for c in ("entity", "year_bucket", "ng_scope", "ng_label", "vertical_label",
              "final_capex_opex", "調整一級"):
        if c in df.columns:
            vals = df[c].dropna().astype(str).unique()[:12]
            print(f"  {c} 值({df[c].nunique()}): {list(vals)}")
    # feed 有冇 主體/執行公司/KPI 欄？
    print(f"  ❓主體/執行公司欄: {_probe(df.columns, '主體', '執行公司', '控股', '子公司', '公司', 'entity_sub')}")
    print(f"  ❓KPI/收入/留宿欄: {_probe(df.columns, 'KPI', '收入', '留宿', '晚數', '毛收入', '國際')}")

    # ── 清單 ──
    if qingdan:
        print("\n" + "=" * 80 + "\n# 清單（Database）")
        wb = openpyxl.load_workbook(qingdan, data_only=True, read_only=True)
        ws = next((wb[s] for s in wb.sheetnames if s.lower().startswith("database")), wb[wb.sheetnames[0]])
        rows = [r for i, r in enumerate(ws.iter_rows(values_only=True)) if i < 40]
        hr = next((i for i, r in enumerate(rows) if r and any("項目類型" in str(c or "") for c in r)), 2)
        hdr = [str(c or "").replace("\n", "") for c in rows[hr]]
        print(f"  sheets:{wb.sheetnames}")
        print(f"  表頭行:{hr+1}  欄數:{len(hdr)}")
        print(f"  ❓股權/主體/執行欄: {_probe(hdr, '股權', '主體', '執行公司', '控股', '子公司', '承批主體')}")
        print(f"  ❓KPI欄: {_probe(hdr, 'KPI', '收入', '留宿', '晚數', '毛收入', '增長率', '國際住客')}")
        print(f"  ❓藝術品欄: {_probe(hdr, '藝術品', '展出', '館藏', '拍賣')}")
        print(f"  範疇/性質欄: {_probe(hdr, '項目性質', '項目類型', '範疇')}")
        # 藝術品：搵含「藝術品」嘅 row（項目名/內容）
        name_c = next((i for i, h in enumerate(hdr) if h.strip() == "項目名稱"), None)
        if name_c is not None:
            arts = [str(r[name_c]) for r in rows[hr + 1:] if name_c < len(r) and r[name_c]
                    and any(k in str(r[name_c]) for k in ("藝術", "博物館", "珍寶"))]
            print(f"  藝術相關項目名: {arts[:6]}")

    # ── 表2 ──
    if biao2_dir:
        print("\n" + "=" * 80 + "\n# 表2")
        try:
            import biao2
            b2 = biao2.load_biao2(biao2_dir, entity or "mgm", log=lambda *a: None)
            print(f"  {len(b2)} 個 (博彩?,碼) 有 finding；有附件檔?")
            d = Path(biao2_dir)
            atts = [p.name for p in d.rglob("*.xls*") if "附件" in p.name and entity in p.name.lower()]
            print(f"  附件檔（可能有藝術品/子項明細）: {atts}")
        except Exception as e:
            print(f"  ⚠ {e}")

    print("\n→ paste 全部俾 Claude，佢會逐頁定 coverage + 開始砌")


if __name__ == "__main__":
    main()
