#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
dump_qingdan_category_rates.py — 由項目清單直接算『逐範疇 2025 計劃完成率』（＝報告做法：
分母含零投資項目）。用嚟確認 group by 項目性質(D) 出嘅完成率 == 報告（會展37.2%/社區23.8%/海上44.9%…）。

清單欄（MGM Database，表頭第3行）：
  C 項目類型（博彩/非博彩）、D 項目性質（範疇）、E 承批公司項目序號、F 項目名稱、
  P/Q 2023計劃/實際、CM/CP 2024計劃/實際合計、EV/EY 2025計劃/實際合計。

用法：python scripts\\report\\dump_qingdan_category_rates.py "data\\投資項目清單\\3. MGM.…投资项目清单.xlsx"
"""
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("✗ pip install openpyxl"); sys.exit(1)


def _num(v):
    try:
        return float(str(v).replace(",", "").strip())
    except (TypeError, ValueError):
        return 0.0


def _find(hdr, *needles, year=None):
    """揾第一個表頭 contains 全部 needles（可指定 year 前綴）嘅欄 index。"""
    for ci, h in enumerate(hdr):
        hs = str(h or "")
        if year and str(year) not in hs:
            continue
        if all(n in hs for n in needles):
            return ci
    return None


def main():
    if len(sys.argv) < 2:
        print("俾清單 xlsx 路徑"); return
    wb = openpyxl.load_workbook(sys.argv[1], data_only=True, read_only=True)
    ws = next((wb[s] for s in wb.sheetnames if s.lower().startswith("database")), wb[wb.sheetnames[0]])
    rows = list(ws.iter_rows(values_only=True))
    # 表頭行：揾含「項目類型」嗰行
    hr = next((i for i, r in enumerate(rows[:14]) if r and any("項目類型" in str(c or "") for c in r)), 2)
    hdr = [str(c or "").replace("\n", "") for c in rows[hr]]

    c_type = _find(hdr, "項目類型")
    c_sub = _find(hdr, "項目性質")
    c_code = _find(hdr, "承批公司項目序號")
    c_name = _find(hdr, "項目名稱")
    cols = {
        "2023計劃": _find(hdr, "預計投資金額", year=2023),
        "2023實際": _find(hdr, "實際投資金額", year=2023),
        "2024計劃": _find(hdr, "預計投資金額", "合計", year=2024),
        "2024實際": _find(hdr, "實際投資金額", "合計", year=2024),
        "2025計劃": _find(hdr, "預計投資金額", "合計", year=2025),
        "2025實際": _find(hdr, "實際投資金額", "合計", year=2025),
    }
    print(f"表頭行 {hr+1}；欄命中：類型={c_type} 性質={c_sub} 序號={c_code} " +
          " ".join(f"{k}={v}" for k, v in cols.items()))

    # 逐 (項目類型, 項目性質) 匯總
    agg = {}
    for r in rows[hr + 1:]:
        if not r or c_code is None or c_code >= len(r) or not r[c_code]:
            continue
        typ = str(r[c_type] or "").strip() if c_type is not None else ""
        sub = str(r[c_sub] or "").strip() if c_sub is not None else ""
        key = (typ, sub)
        a = agg.setdefault(key, {"n": 0, **{k: 0.0 for k in cols}})
        a["n"] += 1
        for k, ci in cols.items():
            if ci is not None and ci < len(r):
                a[k] += _num(r[ci])

    for yr in ("2023", "2024", "2025"):
        print(f"\n{'='*78}\n# {yr} 逐範疇 計劃 / 實際 / 完成率（含零投資項目）")
        print(f"{'項目類型':<10}{'項目性質(範疇)':<20}{'n':>3}{'計劃':>12}{'實際':>12}{'完成率':>9}")
        for (typ, sub), a in sorted(agg.items()):
            plan, act = a[f"{yr}計劃"], a[f"{yr}實際"]
            if plan == 0 and act == 0:
                continue
            rate = f"{act/plan*100:.1f}%" if plan else "—"
            print(f"{typ:<10}{sub:<20}{a['n']:>3}{plan:>12,.0f}{act:>12,.0f}{rate:>9}")

    print("\n→ 對返報告：會展37.2% / 社區23.8% / 海上44.9% / 文化56.2% / 美食77.8% 啱唔啱？")
    print("  啱 = 我就 wire『清單 group by 項目性質 出完成率』入 build_overview。")


if __name__ == "__main__":
    main()
