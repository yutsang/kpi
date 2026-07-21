#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
dump_qingdan_cols.py — 列出投資項目清單 Database sheet 全部欄（Excel 字母 + 表頭 + 一個樣本值），
幫手定位 narrative（build_narrative.py）要用嘅**補充說明欄**：
  KPMG 分析發現、管理層解釋、調整事項備註/理由、計劃/實際投資內容、項目狀況、期後調整說明…

用法（Windows）：
    python scripts\\report\\dump_qingdan_cols.py "data\\投資項目清單\\3. MGM.…投资项目清单.xlsx"
輸出 stdout（貼返嚟俾我）：每欄一行  字母 | 表頭 | 樣本值。
"""
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("✗ pip install openpyxl"); sys.exit(1)

# 幫手標記：表頭 / 樣本含呢啲字 = 好可能係 narrative 素材
HINT = ["發現", "解釋", "說明", "原因", "理由", "備註", "分析", "內容", "狀況", "情況",
        "偏離", "調整", "訪談", "走訪", "management", "finding"]


def main():
    if len(sys.argv) < 2:
        print("俾清單 xlsx 路徑"); return
    wb = openpyxl.load_workbook(sys.argv[1], data_only=True, read_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = []
        for i, r in enumerate(ws.iter_rows(values_only=True)):
            rows.append(r)
            if i > 80:
                break
        hdr_r = None
        for ri in range(min(14, len(rows))):
            joined = "".join(str(v) for v in (rows[ri] or []) if v)
            if "承批公司項目序號" in joined or ("項目名稱" in joined and "項目類型" in joined):
                hdr_r = ri; break
        if hdr_r is None:
            continue
        hdr = rows[hdr_r]
        print(f"\n{'='*90}\n# sheet {sn!r}  header 喺第 {hdr_r+1} 行  共 {len(hdr)} 欄")
        print(f"  {'欄':>3} | {'表頭':<40} | 樣本值（第一個非空）")
        print(f"  {'-'*3}-+-{'-'*40}-+-{'-'*44}")
        for ci in range(len(hdr)):
            letter = get_column_letter(ci + 1)
            h = hdr[ci]
            sample = ""
            for ri in range(hdr_r + 1, min(hdr_r + 60, len(rows))):
                v = rows[ri][ci] if ci < len(rows[ri]) else None
                if v not in (None, ""):
                    sample = str(v).replace("\n", " ").strip()[:70]; break
            hs = "" if h is None else str(h).replace("\n", " ").strip()[:40]
            if not hs and not sample:
                continue
            mark = "  ⟸ narrative?" if any(k in (hs + sample).lower() for k in [x.lower() for x in HINT]) else ""
            print(f"  {letter:>3} | {hs:<40} | {sample}{mark}")
        break     # 淨係頭一個有 Database 結構嘅 sheet
    print("\n（有『⟸ narrative?』標記嘅欄 = 我砌 build_narrative.py 會用嘅補充素材，貼返嚟俾我 pin 實欄名）")


if __name__ == "__main__":
    main()
