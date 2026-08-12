#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
inspect_sources.py — 盤點【報告三個 input】到底有幾多可用欄，同埋我哋而家用咗幾多。

背景：清單 Database 有 ~184 欄，我哋 build_narrative 只 map 咗 10 個概念；
表2 有標準 33 欄（關注事項／建議調整金額／調整原因／跨司兩輪意見…），但 biao2.py 而家係
「盲抓」（見到 ≥30 字嘅 cell 就當 finding，唔知邊欄係咩）。呢個 script 出一份逐欄清單，
等我可以改成【按欄名 structured 抽】，把未用嘅補充資料寫入報告。

用法（Windows，kpi-main 底下）：
    python scripts\\report\\inspect_sources.py
    python scripts\\report\\inspect_sources.py --entity mgm --rows 2500

輸出：console（分批，每批 ≤--batch 行，方便 paste）+ 檔 sources_audit.txt（同一內容）。
每行 = TSV：來源 / 檔或sheet / 欄index / 欄名 / 有值行數 / 相異值數 / 我哋用咗未 / 樣本
「用咗未」：USED=已接入報告；—=未用（＝可以攞嚟寫嘢）。
"""
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

QINGDAN_DIR = "data/投資項目清單"
BIAO2_DIR = "data/表2"
FEED = "tableau_combined_25.csv"

# 我哋而家真係接入咗報告嘅欄（關鍵字）→ 其餘全部係「未用」
USED_QINGDAN = ["承批公司項目序號", "項目名稱", "項目類型", "項目性質", "項目狀況",
                "實際實施地點", "計劃投資內容", "實際投資內容", "分析發現", "投資偏離",
                "管理層解釋", "變更原因", "期後調整内容", "期後調整內容", "調整事項備註",
                "調整備註", "跨司工作組的回", "跨司工作组的回", "跨司工作組回", "第二輪意見",
                "KPMG回覆", "KPMG回复", "預計投資金額"]
USED_FEED = ["entity", "dicj code", "project", "year_bucket", "報告年", "ng_scope", "ng_label",
             "ng_code", "vertical_label", "final_capex_opex", "調整一級",
             "調整前_萬", "調整_萬", "調整後_萬"]


def _clean(v):
    return re.sub(r"\s+", " ", str(v if v is not None else "")).strip()


def _used(name, used_list):
    n = _clean(name)
    return "USED" if any(k in n for k in used_list) else "—"


def _col_stats(rows, ci, hdr, used_list, max_sample=2):
    vals, seen = [], set()
    n = 0
    for r in rows:
        if ci >= len(r) or r[ci] is None:
            continue
        s = _clean(r[ci])
        if not s:
            continue
        n += 1
        seen.add(s[:60])
        if len(vals) < max_sample and len(s) > 1:
            vals.append(s[:90])
    if n == 0:
        return None
    return [str(ci), _clean(hdr)[:60] or "(無標題)", str(n), str(len(seen)),
            _used(hdr, used_list), " ⁄ ".join(vals)]


def _find_header(rows, keys, maxscan=14):
    """揾表頭行：包含任一 key 嘅第一行。回 (row_idx, header_list)。"""
    for ri in range(min(maxscan, len(rows))):
        row = rows[ri] or []
        if any(any(k in _clean(v) for k in keys) for v in row):
            return ri, [_clean(v) for v in row]
    return None, []


# 表2 標準概念（由 mgm sources audit 2026-08-12 實測得出）→ 用嚟認【detail 表頭行】
B2_CONCEPTS = [
    "投資項目序號及名稱", "實施時間", "擬投資金額", "已投放金額", "是否該司局範疇",
    "是否有希望諮詢的問題", "問題狀態", "KPMG提出日期", "KPMG分析", "承批公司管理層解釋",
    "KPMG希望進一步", "跨司工作組的回覆", "畢馬威關注事項", "承批公司的反饋意見",
    "跨司工作組的反饋意見", "是否需進一步", "需溝通關注事項", "該關注事項涉及調整金額",
    "跨司工作組主責部門", "KPMG需與跨司工作組", "跨司工作組最新反饋意見",
    "建議調整金額", "調整原因", "建議調整後金額", "項目分析意見", "建議接納之調整後金額",
    "項目編號", "資料要求",
]


def _detail_header(rows, band=10):
    """表2 真正嘅 detail 表頭行 = 頭 band 行入面 match 到最多概念嗰行。
    ⚠ 之前用『第一個含關鍵字嘅行』會揾到 group 行（上面一行），令 detail 欄名全部變 (無標題)。
    回 (row_idx, n_match)。"""
    best, bn = None, 0
    for ri in range(min(band, len(rows))):
        n = sum(1 for v in (rows[ri] or []) if any(c in _clean(v) for c in B2_CONCEPTS))
        if n > bn:
            bn, best = n, ri
    return best, bn


def audit_qingdan(entity, maxrows, out):
    import openpyxl
    d = Path(QINGDAN_DIR)
    files = [p for p in sorted(d.rglob("*.xlsx"))
             if entity.lower() in p.name.lower() and not p.name.startswith("~$")] if d.exists() else []
    if not files:
        out.append(f"# 清單：揾唔到（{QINGDAN_DIR} / entity={entity}）")
        return
    p = files[0]
    out.append(f"# 清單檔：{p.name}")
    wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
    ws = next((wb[s] for s in wb.sheetnames if s.lower().startswith("database")), wb[wb.sheetnames[0]])
    rows = []
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        rows.append(r)
        if i > maxrows:
            break
    hr, hdr = _find_header(rows, ["承批公司項目序號"])
    if hr is None:
        out.append("# ⚠ 清單揾唔到『承批公司項目序號』表頭"); return
    body = rows[hr + 1:]
    ncol = max((len(r) for r in rows if r), default=0)
    out.append(f"# 清單 sheet={ws.title} 表頭喺第{hr+1}行，{ncol} 欄，{len(body)} 行資料")
    for ci in range(ncol):
        st = _col_stats(body, ci, hdr[ci] if ci < len(hdr) else "", USED_QINGDAN)
        if st:
            out.append(" | ".join(["清單", ws.title] + st))


def audit_biao2(entity, maxrows, out):
    import biao2 as B2
    import inspect_biao2 as IB
    d = Path(BIAO2_DIR)
    if not d.exists():
        out.append(f"# 表2：揾唔到 {BIAO2_DIR}"); return
    files = [p for p in sorted(d.rglob("*.xls*"))
             if not p.name.startswith("~$") and B2._match_entity(p.name, entity.lower())
             and "提供附件" not in p.name]
    out.append(f"# 表2：{len(files)} 檔 match「{entity}」")
    for p in files:
        try:
            wb = IB.load_wb(p)
        except Exception as e:
            out.append(f"# ⚠ 開唔到 {p.name}: {e}"); continue
        for sn in wb.sheetnames:
            try:
                ws = wb[sn]
                rows = []
                for i, r in enumerate(ws.iter_rows(values_only=True)):
                    rows.append(r)
                    if i > maxrows:
                        break
                ncol = max((len(r) for r in rows if r), default=0)
                if ncol == 0:
                    continue
                # 表2 兩層表頭：detail 行（概念名）+ 上面 group 行
                hr, nmatch = _detail_header(rows)
                if hr is None:
                    hr, nmatch = 0, 0
                hdr = [_clean(v) for v in (rows[hr] or [])]
                grp = [_clean(v) for v in (rows[hr - 1] or [])] if hr > 0 else []
                body = rows[hr + 1:]
                out.append(f"# 表2 {p.name}｜{sn}：detail表頭第{hr+1}行（match {nmatch} 個概念）、"
                           f"group第{hr}行，{ncol} 欄，{len(body)} 行資料")
                for ci in range(ncol):
                    h = hdr[ci] if ci < len(hdr) else ""
                    g = grp[ci] if ci < len(grp) else ""
                    name = (f"{g}／{h}" if g and h and g != h else (h or g))
                    st = _col_stats(body, ci, name, [])      # 表2 全部當「未用」（而家係盲抓）
                    if st:
                        out.append(" | ".join(["表2", f"{p.name[:28]}｜{sn[:22]}"] + st))
            except Exception as e:
                out.append(f"# ⚠ {p.name}｜{sn}: {e}")


def audit_feed(entity, out):
    try:
        import pandas as pd
    except ImportError:
        out.append("# feed：冇 pandas"); return
    p = Path(FEED)
    if not p.exists():
        out.append(f"# feed：揾唔到 {FEED}"); return
    # ⚠ 唔可以 nrows 截：feed 幾十萬行、entity 唔一定排喺頭（之前出「0 行」就係咁）
    parts = []
    for ch in pd.read_csv(p, low_memory=False, chunksize=200000):
        parts.append(ch[ch["entity"].astype(str).str.lower() == entity.lower()]
                     if "entity" in ch.columns else ch)
    df = pd.concat(parts, ignore_index=True) if parts else pd.DataFrame()
    if df.empty:
        out.append(f"# feed {p.name}：0 行（{entity}）—— 檢查 entity 值同 feed 內容"); return
    out.append(f"# feed {p.name}：{len(df)} 行（{entity}），{len(df.columns)} 欄")
    for ci, c in enumerate(df.columns):
        s = df[c]
        nn = int(s.notna().sum())
        if nn == 0:
            continue
        samp = " ⁄ ".join(_clean(x)[:60] for x in s.dropna().unique()[:2])
        out.append(" | ".join(["feed", "-", str(ci), _clean(c)[:60], str(nn),
                               str(int(s.nunique(dropna=True))), _used(c, USED_FEED), samp]))


def find_all(entity, needles, out):
    """喺 清單 + 表2 嘅【全部 sheet 全部 cell】搵 keyword／數字（例：10年投資預算 1970000）。
    → 用嚟揾一啲唔喺標準欄嘅數（audit 只掃 Database / 標準表2 sheet）。"""
    import openpyxl
    import biao2 as B2
    import inspect_biao2 as IB
    pats = [str(n) for n in needles]

    def scan(label, wb):
        for sn in wb.sheetnames:
            try:
                for ri, row in enumerate(wb[sn].iter_rows(values_only=True), 1):
                    for ci, v in enumerate(row or []):
                        if v is None:
                            continue
                        t = _clean(v)
                        tn = t.replace(",", "")
                        if any(p in t or p in tn for p in pats):
                            ctx = " ⁄ ".join(_clean(x)[:40] for x in (row or [])[max(0, ci-2):ci+3] if x is not None)
                            out.append(f"HIT | {label} | {sn[:26]} | r{ri} c{ci} | {t[:70]} | …{ctx[:110]}")
            except Exception as e:
                out.append(f"# ⚠ {label}｜{sn}: {e}")
    d = Path(QINGDAN_DIR)
    for p in (sorted(d.rglob("*.xlsx")) if d.exists() else []):
        if entity.lower() in p.name.lower() and not p.name.startswith("~$"):
            scan(f"清單 {p.name[:24]}", openpyxl.load_workbook(p, data_only=True, read_only=True))
    d = Path(BIAO2_DIR)
    for p in (sorted(d.rglob("*.xls*")) if d.exists() else []):
        if p.name.startswith("~$") or not B2._match_entity(p.name, entity.lower()):
            continue
        try:
            scan(f"表2 {p.name[:24]}", IB.load_wb(p))
        except Exception as e:
            out.append(f"# ⚠ 開唔到 {p.name}: {e}")


def main():
    args = sys.argv[1:]
    entity, maxrows, batch = "mgm", 2500, 900
    for flag, cast in (("--entity", str), ("--rows", int), ("--batch", int)):
        if flag in args:
            i = args.index(flag); v = cast(args[i + 1]); del args[i:i + 2]
            entity = v if flag == "--entity" else entity
            maxrows = v if flag == "--rows" else maxrows
            batch = v if flag == "--batch" else batch
    entity = entity.lower()
    needles = []
    while "--find" in args:
        i = args.index("--find"); needles.append(args[i + 1]); del args[i:i + 2]
    if needles:
        out = [f"# ==== 全域搜尋｜entity={entity}｜keyword={needles} ===="]
        find_all(entity, needles, out)
        f = Path(f"{entity}_find.txt"); f.write_text("\n".join(out), encoding="utf-8")
        print(f"✓ {f.resolve()}（{len(out)-1} 個命中）")
        print("\n".join(out[:400]))
        return
    out = [f"# ==== report sources audit｜entity={entity} ====",
           "# 欄位（用 | 分隔，tab 一 paste 就會變空格）：來源 | 檔／sheet | 欄index | 欄名 | "
           "有值行數 | 相異值 | 用咗未 | 樣本",
           "# 『USED』＝已接入報告；『—』＝未用（可以攞嚟寫報告內容）"]
    audit_feed(entity, out)
    audit_qingdan(entity, maxrows, out)
    audit_biao2(entity, maxrows, out)

    f = Path(f"{entity}_sources_audit.txt")
    f.write_text("\n".join(out), encoding="utf-8")
    n_un = sum(1 for l in out if " | — | " in l)
    n_us = sum(1 for l in out if " | USED | " in l)
    print(f"✓ {f.resolve()}（{len(out)} 行；USED {n_us}、未用 {n_un}）")
    print(f"── 以下分批印出，每批 ≤{batch} 行，逐批 paste 返俾我 ──")
    for i in range(0, len(out), batch):
        print(f"\n===== BATCH {i//batch + 1}/{-(-len(out)//batch)} =====")
        print("\n".join(out[i:i + batch]))


if __name__ == "__main__":
    main()
