#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_narrative.py — 由投資項目清單逐項目抽「可抄」嘅 narrative 素材（報告 slide 28-40 主要發現 /
93-100 現場走訪概述 / 概述 project 描述）。全部係清單原文，機械抄，唔使 LLM。

清單 Database sheet 有年度 block（2023/2024/2025 欄重覆），按**表頭關鍵字**歸類收晒非空值：
  項目狀況｜計劃投資內容｜實際投資內容｜實施地點｜KPMG分析發現(投資偏離)｜管理層解釋(變更原因)｜調整事項備註｜期後調整內容
key = 承批公司項目序號（項目N）。

用法（Windows）：
    python scripts\\report\\build_narrative.py "data\\投資項目清單\\3. MGM.…投资项目清单.xlsx" --entity mgm
出 {entity}_narrative.xlsx（逐項目一行，每概念一欄，取最後一個非空＝最新年度）＋ console 預覽。
之後接 LLM（workbench）只需精煉/縮寫，唔使由零寫。
"""
import re
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("✗ pip install openpyxl"); sys.exit(1)

# 概念 → 表頭關鍵字（順序＝輸出欄序）
CONCEPTS = [
    ("項目狀況", ["項目狀況"]),
    ("實施地點", ["實際實施地點", "實際實施地點/空間"]),
    ("計劃投資內容", ["計劃投資內容"]),
    ("實際投資內容", ["實際投資內容"]),
    ("KPMG分析發現", ["分析發現", "投資偏離"]),
    ("管理層解釋", ["管理層解釋", "變更原因"]),
    ("期後調整內容", ["期後調整内容", "期後調整內容"]),
    ("調整事項備註", ["調整事項備註", "調整備註"]),
]


def _norm_code(v):
    s = re.sub(r"\s+", "", str(v if v is not None else ""))
    s = re.sub(r"^項目", "", s)
    m = re.match(r"^0*(\d+)$", s)
    return m.group(1) if m else s


OUT_COLS = ["項目序號", "項目名稱", "項目類型"] + [n for n, _ in CONCEPTS]


def load_narrative(path, log=lambda *a: None) -> dict:
    """讀清單 → {正規化項目編號: rec}，rec 有 項目序號/名稱/類型 + 每概念（取最新非空）。"""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = None
    for sn in wb.sheetnames:
        if sn.lower().startswith("database") or sn == "Database":
            ws = wb[sn]; break
    ws = ws or wb[wb.sheetnames[0]]
    rows = []
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        rows.append(r)
        if i > 2500:
            break
    hdr_r = code_c = name_c = type_c = None
    for ri in range(min(14, len(rows))):
        for ci, v in enumerate(rows[ri] or []):
            sv = "" if v is None else str(v)
            if "承批公司項目序號" in sv:
                hdr_r, code_c = ri, ci
            if name_c is None and sv.strip() == "項目名稱":
                name_c = ci
            if type_c is None and "項目類型" in sv:
                type_c = ci
        if hdr_r is not None:
            break
    if hdr_r is None:
        log("✗ 揾唔到『承批公司項目序號』表頭"); return {}
    hdr = [("" if v is None else str(v).replace("\n", "")) for v in rows[hdr_r]]
    concept_cols = {name: [ci for ci, h in enumerate(hdr) if any(k in h for k in keys)]
                    for name, keys in CONCEPTS}
    log(f"清單 header row {hdr_r+1}；概念欄命中：" +
        ", ".join(f"{n}:{len(concept_cols[n])}" for n, _ in CONCEPTS))
    out = {}
    for ri in range(hdr_r + 1, len(rows)):
        row = rows[ri]
        code = _norm_code(row[code_c]) if code_c < len(row) else ""
        if not code:
            continue
        rec = {"項目序號": f"項目{code}",
               "項目名稱": (row[name_c] if name_c is not None and name_c < len(row) else "") or "",
               "項目類型": (row[type_c] if type_c is not None and type_c < len(row) else "") or ""}
        for name, _ in CONCEPTS:
            vals = []
            for ci in concept_cols[name]:
                if ci < len(row) and row[ci] not in (None, "", 0, "0"):
                    sv = str(row[ci]).strip()
                    if sv and sv not in vals and sv.lower() not in ("n/a", "nan"):
                        vals.append(sv)
            rec[name] = vals[-1] if vals else ""
        out.setdefault(code, rec)
    return out


def main():
    args = sys.argv[1:]
    entity = None
    if "--entity" in args:
        i = args.index("--entity"); entity = args[i + 1].lower(); del args[i:i + 2]
    if not args:
        print("俾清單 xlsx 路徑（--entity mgm）"); return
    narr = load_narrative(args[0], log=lambda m: print(f"（{entity}）{m}"))
    out_cols = OUT_COLS
    out_rows = list(narr.values())

    out = Path(f"{entity or 'all'}_narrative.xlsx")
    wb2 = openpyxl.Workbook(); wsx = wb2.active; wsx.title = "narrative"
    wsx.append(out_cols)
    for rec in out_rows:
        wsx.append([rec.get(c, "") for c in out_cols])
    wb2.save(out)
    n_find = sum(1 for r in out_rows if r["KPMG分析發現"])
    n_adj = sum(1 for r in out_rows if r["調整事項備註"])
    print(f"\n✓ {out.resolve()}：{len(out_rows)} 個項目 × {len(out_cols)} 欄"
          f"（{n_find} 個有 KPMG分析發現、{n_adj} 個有調整備註）")
    # console 預覽：優先有『KPMG分析發現』嘅項目（＝主要發現 narrative 來源），跳過「如期開展」冇調整嘅
    finding_recs = [r for r in out_rows if r["KPMG分析發現"]]
    print(f"\n── 有『KPMG分析發現』嘅項目（主要發現/走訪概述來源）頭 8 個預覽 ──")
    for rec in finding_recs[:8]:
        print(f"\n[{rec['項目序號']}] {str(rec['項目名稱'])[:30]}（{rec['項目類型']}｜{rec['項目狀況']}）")
        if rec["KPMG分析發現"]:
            print(f"  KPMG分析發現: {rec['KPMG分析發現'][:150]}")
        if rec["管理層解釋"]:
            print(f"  管理層解釋: {rec['管理層解釋'][:150]}")


if __name__ == "__main__":
    main()
