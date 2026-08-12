#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_project_review_table.py — 由 prep_tableau feed 生成「單個項目審查結果匯總表」
（報告 slide 46-63 而家係 Tableau 截圖，native 砌返 + 對數）。版式對齊報告 IMG_0094：

  項目基本信息: 項目序號 | 項目名稱 | 計劃投資金額(a) | 報告投資金額(b) | 投資計劃完成率(c=b/a)
  潛在調整事項(7): ①一般支持性部門的人工成本 ②其他日常營運支出調整 ③超出可計入範圍的內部資源支出
                   ④酒店客房改造支出 ⑤不符合"吸引外國客源"定義的相關投資支出 ⑥未完全實現投資目的的投資支出
                   ⑦投資計劃獲批前發生且未被認可的投資支出 | 潛在調整合計(d)
  潛在調整後: 調整後投資金額(e=b+d) | 潛在調整後投資計劃完成率(f=e/a) | 設施建設/資本性支出 | 活動舉辦/營運性支出

分組: 博彩(按 vertical_label) → 博彩合計 → 非博彩(按 ng_label) → 每範疇小計。每年(25/24/23)一 sheet。
⚠ key = (ng_scope, 項目序號)：博彩/非博彩 共用同一「項目N」號係唔同項目（e.g. 項目19）。
⚠ MGM 跨年項目（舊號喺當年期後）暫未特別處理（user 2026-07-20 defer，到時提）。

用法（Windows，kpi-main 底下）：
    python scripts\\report\\build_project_review_table.py "tableau_combined_25.csv" --entity mgm ^
        --qingdan "data\\投資項目清單\\3. MGM.2025年度投資計劃執行情況審查專項工作.投资项目清单.xlsx"
"""
import re
import sys
from pathlib import Path

try:
    import pandas as pd
except ImportError:
    print("✗ 未裝 pandas → pip install pandas openpyxl"); sys.exit(1)

pd.set_option("display.max_columns", 40)
pd.set_option("display.width", 260)

# 調整一級 → canonical 報告 7 類（合併變體）
CANON = {
    "計入報告投資金額的高管及一般支持人員人工成本": "一般支持性部門的人工成本",
    "一般性支持部門的人工成本": "一般支持性部門的人工成本",
    "一般支持性部門的人工成本": "一般支持性部門的人工成本",
    "其他日常營運支出調整": "其他日常營運支出調整",
    "其他日常營運支出調整調整": "其他日常營運支出調整",
    "超過可計入範圍的内部資源支出": "超出可計入範圍的內部資源支出",
    "超出可計入範圍的內部資源支出": "超出可計入範圍的內部資源支出",
    "酒店客房改造支出": "酒店客房改造支出",
    "不符合“吸引外國客源”定義的相關投資支出": "不符合“吸引外國客源”定義的相關投資支出",
    "不符合吸引外國客源": "不符合“吸引外國客源”定義的相關投資支出",
    "未完全實現投資目的的投資支出": "未完全實現投資目的的投資支出",
    "未能實現投資目的的投資支出": "未完全實現投資目的的投資支出",
    "計劃獲批前的投資": "投資計劃獲批前發生且未被認可的投資支出",
    # 以下係跨年機制，報告 7 欄冇（user defer 跨年）→ 落 trailing other，之後 filter 當年就消失
    "將2024年的支出計入2025年報告投資金額": "〔跨年〕將往年支出計入本年報告投資金額",
    "2024年度計劃與2023年度計劃期後調整之間的報告投資金額跨期調整": "〔跨年〕年度計劃期後跨期調整",
}
ADJ7 = [
    "一般支持性部門的人工成本", "其他日常營運支出調整", "超出可計入範圍的內部資源支出",
    "酒店客房改造支出", "不符合“吸引外國客源”定義的相關投資支出",
    "未完全實現投資目的的投資支出", "投資計劃獲批前發生且未被認可的投資支出",
]
BASE_L = ["計劃投資金額", "報告投資金額", "投資計劃完成率"]
TAIL_L = ["潛在調整合計", "調整後投資金額", "潛在調整後投資計劃完成率", "設施建設/資本性支出", "活動舉辦/營運性支出"]

_PLAN_RE = {
    25: re.compile(r"2025.*預計投資金額.*合計|2025.*預計投資金額（萬"),
    24: re.compile(r"2024.*預計投資金額.*合計|2024.*預計投資金額（萬"),
    23: re.compile(r"2023.*預計投資金額"),
}


def _norm(c) -> str:
    s = re.sub(r"\s+", "", str(c if c is not None else ""))
    s = re.sub(r"^項目", "", s)
    m = re.match(r"^0*(\d+(?:\.\d+)?)$", s)
    return (m.group(1) if m else s).lower()


def _plan_year(yb) -> int:
    """year_bucket → 計劃年份（報告匯總表按計劃年分，唔係支出年）：
    25→25(2025計劃), 25_24SY→24(2024計劃嘅2025期後), 25_23SY→23, 24→24, 24_23SY→23, 23→23。"""
    s = str(yb).strip()
    m = re.search(r"_(\d+)SY", s)          # 有 _NNSY = 該計劃年嘅期後
    if m:
        return int(m.group(1))
    m2 = re.match(r"^(\d{2})", s)
    return int(m2.group(1)) if m2 else -1


def load_plan(path: Path, log=print) -> dict:
    """{報告年: {(is_gaming, 正規化項目編號): 計劃_萬}}（清單 database；博彩/非博彩共用項目N → key 帶 scope）。"""
    import openpyxl
    out = {25: {}, 24: {}, 23: {}}
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        log(f"  ⚠ 清單開唔到 {path}: {e}"); return out
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = []
        for i, r in enumerate(ws.iter_rows(values_only=True)):
            rows.append(r)
            if i > 2500:
                break
        hdr_r = code_c = type_c = None
        for ri in range(min(14, len(rows))):
            for ci, v in enumerate(rows[ri] or []):
                sv = "" if v is None else str(v)
                if "承批公司項目序號" in sv:
                    hdr_r, code_c = ri, ci
                if type_c is None and "項目類型" in sv:
                    type_c = ci
            if hdr_r is not None:
                break
        if hdr_r is None:
            continue
        hdr = [("" if v is None else str(v).replace("\n", "")) for v in rows[hdr_r]]
        plan_c = {}
        for yr, rgx in _PLAN_RE.items():
            for ci, h in enumerate(hdr):
                if rgx.search(h):
                    plan_c[yr] = ci; break
        if not plan_c:
            continue
        log("  清單 " + sn + ": 計劃欄 " + ", ".join(f"{yr}→{hdr[ci][:18]!r}" for yr, ci in plan_c.items())
            + (f"；項目類型欄 col{type_c}" if type_c is not None else "；⚠冇項目類型欄(scope 分唔到)"))
        for ri in range(hdr_r + 1, len(rows)):
            row = rows[ri]
            code = _norm(row[code_c]) if code_c < len(row) else ""
            if not code:
                continue
            gaming = False
            if type_c is not None and type_c < len(row) and row[type_c] is not None:
                gaming = str(row[type_c]).strip().startswith("博彩")
            for yr, ci in plan_c.items():
                if ci < len(row):
                    try:
                        out[yr].setdefault((gaming, code), float(row[ci]))
                    except (TypeError, ValueError):
                        pass
        break
    return out


def load_category(path: Path, log=lambda *a: None) -> dict:
    """{(is_gaming, 正規化項目編號): 項目性質(D)}（清單；用嚟將零投資項目計劃 attribute 返範疇）。
    含 feed 冇嘅零投資項目（清單有齊全部項目）。"""
    import openpyxl
    out = {}
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        log(f"  ⚠ 清單開唔到 {path}: {e}"); return out
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = []
        for i, r in enumerate(ws.iter_rows(values_only=True)):
            rows.append(r)
            if i > 2500:
                break
        hdr_r = code_c = type_c = sub_c = None
        for ri in range(min(14, len(rows))):
            for ci, v in enumerate(rows[ri] or []):
                sv = "" if v is None else str(v)
                if "承批公司項目序號" in sv:
                    hdr_r, code_c = ri, ci
                if type_c is None and "項目類型" in sv:
                    type_c = ci
                if sub_c is None and sv.strip() == "項目性質":
                    sub_c = ci
            if hdr_r is not None:
                break
        if hdr_r is None or sub_c is None:
            continue
        for ri in range(hdr_r + 1, len(rows)):
            row = rows[ri]
            code = _norm(row[code_c]) if code_c < len(row) else ""
            if not code:
                continue
            gaming = (type_c is not None and type_c < len(row) and row[type_c] is not None
                      and str(row[type_c]).strip().startswith("博彩"))
            sub = row[sub_c] if sub_c < len(row) else None
            if sub is not None and str(sub).strip():
                out.setdefault((gaming, code), str(sub).strip())
        log(f"  清單 項目性質(D)：{len(out)} 個項目")
        break
    return out


def _rate(rep, plan):
    try:
        return round(rep / plan, 4) if plan else None
    except Exception:
        return None


def build_year(df: pd.DataFrame, year: int, plan: dict | None = None):
    # 按計劃年份(plan year)分表：跨年期後(25_24SY 等)落返啱嘅計劃年，唔會爆 2025 表小計
    d = df[df["_plan_year"] == year].copy()
    # 只留乾淨「項目N」碼（丟 項目CAPEX-5 等 pseudo/分攤碼）
    d = d[d["dicj code"].astype(str).str.match(r"^項目\s*\d")]
    if d.empty:
        return None, []
    d["_adj"] = d["調整一級"].map(CANON).fillna(d["調整一級"])
    d["_sub"] = d.apply(lambda r: r["vertical_label"] if r["ng_scope"] == "gaming" else r["ng_label"], axis=1)
    key = ["ng_scope", "dicj code"]

    def _mode(s):
        m = s.dropna()
        return m.mode().iat[0] if len(m.mode()) else (m.iloc[0] if len(m) else "")
    base = d.groupby(key, dropna=False).agg(
        項目名稱=("project", "first"),
        _sub=("_sub", _mode),
        _ngcode=("ng_code", _mode),
        報告投資金額=("調整前_萬", "sum"),
        潛在調整合計=("調整_萬", "sum"),
        調整後投資金額=("調整後_萬", "sum"),
    )
    cap = d[d["final_capex_opex"] == "Capex"].groupby(key)["調整後_萬"].sum()
    ope = d[d["final_capex_opex"] == "Opex"].groupby(key)["調整後_萬"].sum()
    base["設施建設/資本性支出"] = cap.reindex(base.index).fillna(0)
    base["活動舉辦/營運性支出"] = ope.reindex(base.index).fillna(0)
    adj = d[d["調整一級"].notna()]
    pv = adj.pivot_table(index=key, columns="_adj", values="調整_萬", aggfunc="sum", fill_value=0)
    tab = base.join(pv).fillna(0).reset_index()
    tab = tab.rename(columns={"dicj code": "項目序號"})
    # 計劃 + 完成率（(scope,code) join；博彩/非博彩 collision → code-only fallback）
    if plan:
        codeonly = {}
        for (g, c), v in plan.items():
            codeonly.setdefault(c, v)
        tab["計劃投資金額"] = pd.to_numeric(tab.apply(
            lambda r: plan.get((r["ng_scope"] == "gaming", _norm(r["項目序號"])),
                               codeonly.get(_norm(r["項目序號"]))), axis=1), errors="coerce")
    else:
        tab["計劃投資金額"] = pd.NA
    tab["投資計劃完成率"] = tab.apply(lambda r: _rate(r["報告投資金額"], r["計劃投資金額"]), axis=1)
    tab["潛在調整後投資計劃完成率"] = tab.apply(lambda r: _rate(r["調整後投資金額"], r["計劃投資金額"]), axis=1)
    adj_cols = [c for c in ADJ7 if c in tab.columns]
    other = [c for c in pv.columns if c not in ADJ7]     # 跨年/未預期
    num = ["計劃投資金額", "報告投資金額"] + adj_cols + other + ["潛在調整合計",
          "調整後投資金額", "設施建設/資本性支出", "活動舉辦/營運性支出"]
    for c in num:
        tab[c] = pd.to_numeric(tab[c], errors="coerce").round(1)

    # 排序：博彩先，博彩內 vertical_label（娛樂場→設施設備），非博彩 ng_code(數字)；內部項目N
    def _pn(s):
        m = re.search(r"(\d+(?:\.\d+)?)", str(s)); return float(m.group(1)) if m else 9e9
    def _ngn(s):
        m = re.search(r"(\d+)", str(s)); return int(m.group(1)) if m else 99
    gorder = {"博彩娛樂場優化": 0, "博彩娛樂場場地的優化": 0, "博彩設施設備優化": 1, "博彩設施及設備的優化": 1}
    tab["_scope"] = (tab["ng_scope"] != "gaming").astype(int)          # 博彩=0 先
    tab["_go"] = tab["_sub"].map(lambda s: gorder.get(s, 5))
    tab["_ngn"] = tab["_ngcode"].map(_ngn)
    tab["_pn"] = tab["項目序號"].map(_pn)
    tab = tab.sort_values(["_scope", "_go", "_ngn", "_pn"]).reset_index(drop=True)

    # 砌 rows：section header + data + 範疇小計 + 類型合計
    ALL = ["項目序號", "項目名稱"] + BASE_L + adj_cols + other + TAIL_L
    numcols = ["計劃投資金額", "報告投資金額"] + adj_cols + other + ["潛在調整合計",
              "調整後投資金額", "設施建設/資本性支出", "活動舉辦/營運性支出"]
    rows = []

    def _agg_row(sub, label):
        r = {c: "" for c in ALL}
        r["項目序號"] = label
        for c in numcols:
            r[c] = round(pd.to_numeric(sub[c], errors="coerce").sum(), 1)
        r["投資計劃完成率"] = _rate(r["報告投資金額"], r["計劃投資金額"])
        r["潛在調整後投資計劃完成率"] = _rate(r["調整後投資金額"], r["計劃投資金額"])
        return r

    for scope in [s for s in ["gaming", "non_gaming"] if (tab["ng_scope"] == s).any()]:
        sc = tab[tab["ng_scope"] == scope]
        scope_name = "博彩項目" if scope == "gaming" else "非博彩項目"
        for sub_name, sub in sc.groupby(sc["_sub"], sort=False):
            rows.append({**{c: "" for c in ALL}, "項目序號": f"{scope_name}—{sub_name}"})  # section
            for _, row in sub.iterrows():
                rows.append({c: row.get(c, "") for c in ALL})
            rows.append(_agg_row(sub, f"{scope_name}—{sub_name} 小計"))
        rows.append(_agg_row(sc, f"{scope_name}合計"))
    rows.append(_agg_row(tab, "總計"))          # 全表總計（scan 每年表最後一行都有）

    out_df = pd.DataFrame(rows, columns=ALL)
    return out_df, other


def main():
    args = sys.argv[1:]
    entity = qingdan = None
    if "--entity" in args:
        i = args.index("--entity"); entity = args[i + 1].lower(); del args[i:i + 2]
    if "--qingdan" in args:
        i = args.index("--qingdan"); qingdan = args[i + 1]; del args[i:i + 2]
    if not args:
        print("俾 tableau feed csv 路徑（--qingdan <清單.xlsx> 加計劃金額+完成率）"); return
    df = pd.read_csv(Path(args[0]), low_memory=False)
    if entity and "entity" in df.columns:
        df = df[df["entity"].astype(str).str.lower() == entity]
    df["報告年"] = pd.to_numeric(df["報告年"], errors="coerce")
    df["_plan_year"] = df["year_bucket"].map(_plan_year)     # 計劃年份分表用

    plan = None
    if qingdan:
        print(f"── 讀清單計劃金額: {qingdan}")
        plan = load_plan(Path(qingdan))
        for yr in (25, 24, 23):
            print(f"    報告年{yr}: {len(plan.get(yr, {}))} 個項目有計劃金額")

    out = Path(f"{entity or 'all'}_項目審查匯總.xlsx")
    unmapped = set()
    with pd.ExcelWriter(out, engine="openpyxl") as xw:
        for yr in (25, 24, 23):
            tab, other = build_year(df, yr, plan.get(yr) if plan else None)
            if tab is None or tab.empty:
                continue
            unmapped |= set(other)
            tab.to_excel(xw, sheet_name=f"報告年{yr}", index=False)
            print(f"\n{'='*90}\n# 報告年 {yr}（{len(tab)} 行含 section/小計）")
            prev = tab.copy()
            prev["項目名稱"] = prev["項目名稱"].astype(str).str.slice(0, 12)
            print(prev.head(34).to_string(index=False))
            if len(tab) > 34:
                print(f"    …（另 {len(tab)-34} 行，見 xlsx）")
    if unmapped:
        print(f"\n⚠ 報告 7 欄以外嘅調整類型（多數係跨年，暫落 trailing）: {sorted(unmapped)}")
    print(f"\n✓ 寫入 {out.resolve()}")


if __name__ == "__main__":
    main()
