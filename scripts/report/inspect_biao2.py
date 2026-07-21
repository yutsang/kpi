#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
inspect_biao2.py — 拆解「表2」(表二審查底稿) 結構。表2 = per-project 審查工作底稿，2 行表頭
(group header + detail header)，有 finding/調整/跨司溝通嘅完整欄，係報告主要發現/概覽/單項審查嘅源。

用法（Windows）：
    python scripts\\report\\inspect_biao2.py "data\\表2"            # 成個資料夾
    python scripts\\report\\inspect_biao2.py "data\\表2\\MGM….xlsx"  # 單一檔
輸出 stdout（貼返嚟 / 大就掉 results\\）：逐檔逐 sheet →
  · dims / 資料行數
  · 每欄：Excel字母 | 組標題(r上) | 欄名(r下) | 第一個非空樣本值
  · key 欄(投資項目序號及名稱) + 金額/調整/發現欄位置
"""
import io
import re
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("✗ pip install openpyxl"); sys.exit(1)

PASSWORD = "dicj_kpmg"     # 表2 同報告一樣加密（msoffcrypto）


def load_wb(path, password=PASSWORD):
    """開 xlsx；『not a zip file』＝加密 → msoffcrypto 用密碼解。回 openpyxl workbook。"""
    try:
        return openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception:
        pass
    import msoffcrypto
    buf = io.BytesIO()
    with open(path, "rb") as f:
        off = msoffcrypto.OfficeFile(f)
        off.load_key(password=password)
        off.decrypt(buf)
    buf.seek(0)
    return openpyxl.load_workbook(buf, data_only=True, read_only=True)

KEY_HINT = ["投資項目序號", "項目序號及名稱", "項目序號"]
NARR_HINT = ["KPMG分析", "關注事項", "管理層解釋", "調整金額", "調整原因", "分析意見",
             "反饋意見", "調整後金額", "項目分類", "備註", "狀態"]


def _find_header(rows):
    """揾 detail header row（含投資項目序號）；上一行＝group header。回 (grp_r, det_r)。"""
    for ri in range(min(12, len(rows))):
        joined = "".join(str(v) for v in (rows[ri] or []) if v)
        if any(h in joined for h in KEY_HINT):
            return (ri - 1 if ri > 0 else ri), ri
    return None, None


def _ffill(seq):
    out, last = [], ""
    for v in seq:
        s = "" if v is None else str(v).replace("\n", "").strip()
        if s:
            last = s
        out.append(last)
    return out


def inspect_one(path: Path):
    print(f"\n{'#'*84}\n# {path.name}")
    try:
        wb = load_wb(path)
    except ImportError:
        print("  ✗ 加密檔要 msoffcrypto → pip install msoffcrypto-tool"); return
    except Exception as e:
        print(f"  ✗ 開唔到: {type(e).__name__}: {e}"); return
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = []
        for i, r in enumerate(ws.iter_rows(values_only=True)):
            rows.append(r)
            if i > 800:
                break
        grp_r, det_r = _find_header(rows)
        print(f"\n{'='*72}\n## sheet {sn!r}  約 {len(rows)} 行"
              + (f"  header：組 r{grp_r+1} / 欄 r{det_r+1}" if det_r is not None else "  ⚠揾唔到表頭"))
        if det_r is None:
            for ri in range(min(6, len(rows))):
                cells = [("" if v is None else str(v).replace("\n", " ")[:20]) for v in (rows[ri] or [])]
                print(f"  r{ri}:", " | ".join(c for c in cells if c)[:200])
            continue
        grp = _ffill(rows[grp_r]) if grp_r is not None and grp_r >= 0 else [""] * len(rows[det_r])
        det = [("" if v is None else str(v).replace("\n", "").strip()) for v in rows[det_r]]
        ncol = max(len(grp), len(det))
        ndata = sum(1 for ri in range(det_r + 1, len(rows))
                    if rows[ri] and any(v not in (None, "") for v in rows[ri]))
        print(f"  資料行約 {ndata}；共 {ncol} 欄：")
        for ci in range(ncol):
            g = grp[ci] if ci < len(grp) else ""
            dcol = det[ci] if ci < len(det) else ""
            sample = ""
            for ri in range(det_r + 1, min(det_r + 60, len(rows))):
                row = rows[ri]
                if row and ci < len(row) and row[ci] not in (None, ""):
                    sample = str(row[ci]).replace("\n", " ").strip()[:50]; break
            if not (g or dcol or sample):
                continue
            mark = ""
            if any(h in (dcol) for h in KEY_HINT):
                mark = "  ⟸ KEY(項目)"
            elif any(h in (g + dcol) for h in NARR_HINT):
                mark = "  ⟸ finding/調整"
            print(f"    {get_column_letter(ci+1):>3} | {g[:16]:<16} | {dcol[:24]:<24} | {sample}{mark}")


def main():
    args = sys.argv[1:]
    if not args:
        print('俾 data\\表2 資料夾或單一 xlsx'); return
    targets = []
    for a in args:
        p = Path(a)
        if p.is_dir():
            targets += sorted(x for x in p.rglob("*.xls*") if not x.name.startswith("~$"))
        elif p.exists():
            targets.append(p)
    if not targets:
        print("✗ 揾唔到 xlsx"); return
    for t in targets:
        inspect_one(t)


if __name__ == "__main__":
    main()
