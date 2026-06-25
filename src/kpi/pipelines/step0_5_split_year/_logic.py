"""Step 0.5: split JE rows into year buckets (25 / 25_24SY / 25_23SY).

Modes (configured per-entity in `year_split:` block of parameters.yml):

  1. direct_filter — Galaxy / Wynn / VML / Melco
     One column already tells you which year-bucket each row is. No row duplication.

  2. hybrid_direct_and_ratio — SJM
     A filter column tells you direct bucket for most rows. Rows with the special
     `ratio_bucket` value (e.g. '跨年項目') get split into 3 rows by ratio from a
     `ratio_source` (database tab). Rows with `proportional_distribute` values
     (e.g. NaN) get split using the OVERALL distribution of the entity.

  3. ratio_with_fallback — MGM
     Try matching every row's project_id (or project_name fuzzy) against Master tab
     ratios I:(J-S):K. Unmatched rows fall back to `report_year` column.

Output:
  data/<alias>/interim/<company>_split.parquet
      Same rows as raw.parquet, with these added columns (audit trail):
        report_period       — '25' / '25_24SY' / '25_23SY'
        split_ratio         — fraction of original amount kept by this output row (1.0 if no split)
        split_source        — provenance: 'direct_filter' / 'master_ratio' /
                              'name_fuzzy' / 'name_exact' / 'report_year_fallback' /
                              'edge_case_25' / 'proportional' / 'unconfig_default'
        original_amount     — amount BEFORE ratio multiplication (for back-check)
        _uid                — original _uid + suffix '__25' / '__24SY' / '__23SY' for ratio-split rows

  data/<alias>/interim/<company>_split_audit.xlsx
      Sheet 'summary'        — bucket totals per source
      Sheet 'project_calc'   — for ratio split entities: per-project ratio + amount calc
      Sheet 'edge_cases'     — projects with J<S or all-zero ratio (MGM)
      Sheet 'fuzzy_review'   — MGM only: parquet_name × candidate Master.B with manual_override_B

  data/<alias>/interim/<company>_fuzzy_match_review.xlsx  (MGM only — read on rerun)
      Persistent user-editable review file. `manual_override_B` overrides auto fuzzy.
"""
from __future__ import annotations

from pathlib import Path
import logging

import openpyxl
import pandas as pd

from kpi.lib.conf import load_config, path
from kpi.lib.io_setup import force_unbuffered_io

force_unbuffered_io()
log = logging.getLogger(__name__)


def _resolve_filter_col(df: pd.DataFrame, requested: str) -> str:
    """Resolve filter_col name against df columns — exact match preferred, else
    fall back to unique-substring match.

    Some 24 source files have year-marker columns with long Chinese descriptions
    embedding newlines and quotes (e.g. SJM 24's `項目所屬年份\n（①若為0...）`).
    Specifying the full literal in YAML is error-prone — allow a short unique
    substring to identify the column instead.

    Raises ValueError if no match or multiple matches.
    """
    if requested in df.columns:
        return requested
    candidates = [c for c in df.columns if requested in str(c)]
    if len(candidates) == 1:
        return candidates[0]
    if len(candidates) == 0:
        raise ValueError(
            f"filter_col {requested!r} not in raw.parquet (no exact or substring match). "
            f"Available cols (first 10): {list(df.columns[:10])}"
        )
    raise ValueError(
        f"filter_col {requested!r} matches {len(candidates)} cols ambiguously. "
        f"Be more specific. Candidates: {candidates[:3]}"
    )


# ─────────────────────────────────────────────────────────────────────────────
# Entry
# ─────────────────────────────────────────────────────────────────────────────
def main():
    cfg = load_config()
    company = cfg["company"]["code"]
    interim = path(cfg, "interim_dir")

    parquet_path = interim / f"{company}_raw.parquet"
    if not parquet_path.exists():
        raise FileNotFoundError(f"Run step0 first. Missing: {parquet_path}")

    # Idempotency check WITHOUT loading the (possibly huge) frame: read only the parquet schema.
    # Galaxy's 1.15M-row raw OOM'd here just reading to skip on a memory-tight box.
    try:
        import pyarrow.parquet as _pq
        _names = _pq.read_schema(parquet_path).names
    except Exception:
        _names = None
    if _names is not None and "report_period" in _names:
        print(f"  [skip] {parquet_path.name} already has 'report_period' column "
              f"(delete file to redo split)")
        return

    df = pd.read_parquet(parquet_path)
    n_in = len(df)

    # Idempotency (fallback, if schema read was unavailable): skip if already split.
    if "report_period" in df.columns:
        print(f"  [skip] {parquet_path.name} already has 'report_period' column "
              f"(delete file to redo split)")
        return

    print(f"step0.5: read {parquet_path.name} rows={n_in:,}")

    # year_split config lives at the entity level (master.companies.<entity_key>.year_split)
    master = cfg.get("_master") or {}
    entity_key = cfg.get("_entity") or company
    ent = (master.get("companies") or {}).get(entity_key, {})
    ys = ent.get("year_split") or {}
    if not ys:
        df_out = _no_config(df, cfg)
        df_out.to_parquet(parquet_path, index=False)
        print(f"  [no year_split config] all rows → '25'  rows={len(df_out):,}")
        _write_summary_audit(df_out, interim, company, {})
        return

    mode = ys.get("mode")
    # ── Per-year-mode dispatch ────────────────────────────────────────────────
    # If any per_year_rule has its own `mode` field, each year is dispatched
    # individually to its own handler. Used by SJM 24 (direct_filter) coexisting
    # with SJM 25 (hybrid_direct_and_ratio) in the same year_split block.
    per_year_rules = ys.get("per_year_rules") or []
    if per_year_rules and any(r.get("mode") for r in per_year_rules):
        df_out, audit = _do_per_year_mode(df, ys, cfg, interim, company)
    elif mode == "direct_filter":
        df_out, audit = _do_direct_filter(df, ys, cfg)
    elif mode == "hybrid_direct_and_ratio":
        df_out, audit = _do_hybrid(df, ys, cfg, interim, company)
    elif mode == "ratio_with_fallback":
        df_out, audit = _do_ratio_fallback(df, ys, cfg, interim, company)
    else:
        raise ValueError(f"Unknown year_split mode: {mode!r}")

    # ── DICJ bucket-scoped remap: golden 同一 project 喺唔同 bucket 用唔同 code ──
    # mgm 2025: 項目121/168/177… 喺 bucket 25 用 2025/Opex 碼 (202/244/230)，喺 25_24SY 用返 2024 碼。
    # report_period 喺 step0_5 split 完先有 → 喺呢度做 (step0 嘅 dicj_remap_by_year 太粗會跨 bucket 改錯)。
    # 只動 dicj_code，唔郁 Project_code → V map 安全。
    _dremap_b = ent.get("dicj_remap_by_bucket") or {}
    if _dremap_b and "dicj_code" in df_out.columns and "report_period" in df_out.columns:
        _rp = df_out["report_period"].astype(str).str.strip()
        _dc = df_out["dicj_code"].astype("string").fillna("")
        _chg = 0
        for _bk, _mp in _dremap_b.items():
            if not _mp:
                continue
            _bm = _rp.eq(str(_bk))
            for _src, _dst in _mp.items():
                _hit = _bm & _dc.eq(str(_src))
                _n = int(_hit.sum())
                if _n:
                    _dc = _dc.mask(_hit, str(_dst))
                    _chg += _n
        df_out["dicj_code"] = _dc
        _rules = sum(len(m or {}) for m in _dremap_b.values())
        print(f"  dicj_code ← remap_by_bucket ({_rules} 規則): {_chg:,} 行改寫 (對齊 golden per-bucket)")

    # OVERWRITE raw.parquet with split version. Downstream (step1/2/4) reads as-is.
    df_out.to_parquet(parquet_path, index=False)
    print(f"  rows in={n_in:,}  out={len(df_out):,}  → {parquet_path.name} (overwritten)")

    # Summary by report_period
    sumdf = df_out.groupby("report_period").agg(
        rows=("split_ratio", "size"),
        amount=("original_amount", lambda s: float(pd.to_numeric(s, errors="coerce").fillna(0).sum())),
    )
    print(f"\n  Bucket summary:\n{sumdf.to_string()}")

    _write_summary_audit(df_out, interim, company, audit)


# ─────────────────────────────────────────────────────────────────────────────
# Mode 0: no config (unconfigured entity)
# ─────────────────────────────────────────────────────────────────────────────
def _no_config(df: pd.DataFrame, cfg: dict) -> pd.DataFrame:
    df = df.copy()
    df["report_period"] = "25"
    df["split_ratio"] = 1.0
    df["split_source"] = "unconfig_default"
    amt = pd.to_numeric(df[cfg["columns"]["amount"]], errors="coerce")
    df["original_amount"] = amt
    return df


# ─────────────────────────────────────────────────────────────────────────────
# Mode 1: direct filter (Galaxy / Wynn / VML / Melco)
# ─────────────────────────────────────────────────────────────────────────────
def _do_direct_filter(df: pd.DataFrame, ys: dict, cfg: dict):
    # If per_year_rules present → multi-year mode (Wynn-style 6-bucket).
    if ys.get("per_year_rules"):
        return _do_direct_filter_multi_year(df, ys, cfg)

    fcol_req = ys["filter_col"]
    buckets = ys.get("buckets") or {}
    default = ys.get("default_bucket", "25")
    fcol = _resolve_filter_col(df, fcol_req)

    df = df.copy()
    amt = pd.to_numeric(df[cfg["columns"]["amount"]], errors="coerce")
    fval = df[fcol].astype(str).str.strip()

    df["report_period"] = default
    df["split_source"] = "direct_filter"
    for bk, expected in buckets.items():
        vals = expected if isinstance(expected, list) else [expected]
        mask = fval.isin([str(v).strip() for v in vals])
        df.loc[mask, "report_period"] = bk

    # Track which rows fell to default
    expected_all = set()
    for v in buckets.values():
        for vv in (v if isinstance(v, list) else [v]):
            expected_all.add(str(vv).strip())
    fallback_mask = ~fval.isin(expected_all)
    df.loc[fallback_mask, "split_source"] = "default_fallback"

    df["split_ratio"] = 1.0
    df["original_amount"] = amt

    audit = {
        "mode": "direct_filter",
        "unassigned_rows": int(fallback_mask.sum()),
        "unassigned_top_values": fval[fallback_mask].value_counts().head(10).to_dict(),
    }
    return df, audit


def _do_bucket_amount_cols(df: pd.DataFrame, ys: dict, cfg: dict):
    """Each row holds per-bucket amounts in SEPARATE columns; explode into one row per non-zero
    bucket and set the canonical amount = that bucket's column (+ optional adjust col).

    YAML (per-year rule):
        - report_year: 2024
          mode: bucket_amount_cols
          buckets:
            "24":      {amount: "2024年度金額", adjust: "2024年度調整"}
            "24_23SY": {amount: "2023年度金額", adjust: "2023年度調整"}
        - report_year: 2025                 # mgm 25 — already-final per-bucket cols, no adjust
          mode: bucket_amount_cols
          buckets:
            "25": {amount: "25_Amt"}
            "25_24SY": {amount: "24_Amt"}
            "25_23SY": {amount: "23_Amt"}
    A row contributing to 2 buckets is duplicated (one per bucket) — fine for a fact table; the
    H-join is on signature, not unique_id.
    """
    amt_col = cfg["columns"]["amount"]
    buckets = ys.get("buckets") or {}
    parts = []
    audit = {"mode": "bucket_amount_cols", "buckets": {}}
    for bk, spec in buckets.items():
        if not isinstance(spec, dict):
            raise ValueError(f"bucket_amount_cols['{bk}'] must be a mapping with 'amount' (got {spec!r})")
        a_name = spec.get("amount")
        adj_name = spec.get("adjust")
        if a_name not in df.columns:
            print(f"  [bucket_amount_cols] {bk}: amount col {a_name!r} not in raw — skipped", flush=True)
            continue
        a = pd.to_numeric(df[a_name], errors="coerce").fillna(0.0)
        _adj = (pd.to_numeric(df[adj_name], errors="coerce").fillna(0.0)
                if adj_name and adj_name in df.columns else pd.Series(0.0, index=df.index))
        a = a + _adj
        # 保留「post 金額=0 但有調整」嘅行（e.g. SJM 調整事項Y：25調整金額=−25跨年 全額撤回），
        # 等 調整前/調整數/調整後 三維度顯示得到（amount_mop=0 → 對 tie 零貢獻，prep_tableau
        # zero-drop 亦因 調整前_萬≠0 而保留）。原本 nz=a!=0 會將呢啲純調整行直接 drop。
        nz = (a != 0.0) | (_adj != 0.0)
        sub = df[nz].copy()
        sub[amt_col] = a[nz].values
        sub["report_period"] = bk
        sub["split_source"] = "bucket_amount_cols"
        sub["split_ratio"] = 1.0
        sub["original_amount"] = a[nz].values
        parts.append(sub)
        audit["buckets"][bk] = {"rows": int(nz.sum()), "sum": float(a[nz].sum()),
                                "amount_col": a_name, "adjust_col": adj_name}
        print(f"  [bucket_amount_cols] {bk}: {int(nz.sum()):,} rows  "
              f"Σ={a[nz].sum()/1e6:,.2f}M  ({a_name!r}{'+'+adj_name if adj_name else ''})", flush=True)
    out = pd.concat(parts, ignore_index=True) if parts else df.iloc[0:0].copy()
    return out, audit


def _do_direct_filter_multi_year(df: pd.DataFrame, ys: dict, cfg: dict):
    """Multi-year direct filter — per-(report_year, filter_col) rules → 6 buckets.

    YAML schema:
        year_split:
          mode: direct_filter
          default_bucket: "25"
          per_year_rules:
            - report_year: 2025
              filter_col: "是否2025年"
              buckets:
                "25":      "2025年度"
                "25_24SY": "2024年度"
                "25_23SY": "2023年度"
            - report_year: 2024
              filter_col: "是否2024年"
              buckets:
                "24":      "2024年度"
                "24_23SY": "2023年計劃2024年完成情況"
              default_bucket: "24"
            - report_year: 2023
              default_bucket: "23"   # no filter_col → all 2023 rows → "23"

    Requires raw.parquet to have a `report_year` column (populated by step0_convert
    yearly_sources or step0_prebuild prebuild_years).
    """
    if "report_year" not in df.columns:
        raise ValueError(
            "year_split.per_year_rules requires 'report_year' column in raw.parquet "
            "(via yearly_sources or prebuild_years)."
        )

    df = df.copy()
    amt = pd.to_numeric(df[cfg["columns"]["amount"]], errors="coerce")
    df["original_amount"] = amt
    df["split_ratio"] = 1.0
    df["split_source"] = "default_fallback"
    default = ys.get("default_bucket", "25")
    df["report_period"] = default

    ry = pd.to_numeric(df["report_year"], errors="coerce")

    rules = ys.get("per_year_rules") or []
    per_year_stats = {}
    for rule in rules:
        rule_year = int(rule["report_year"])
        rule_default = rule.get("default_bucket", default)
        ry_mask = ry == rule_year

        rule_fcol_req = rule.get("filter_col", "")
        rule_buckets = rule.get("buckets") or {}

        if not rule_fcol_req or not rule_buckets:
            # No filter for this year — all rows go to year-default
            df.loc[ry_mask, "report_period"] = rule_default
            df.loc[ry_mask, "split_source"] = "year_default"
            per_year_stats[rule_year] = {"rule_default": rule_default,
                                         "rows": int(ry_mask.sum())}
            continue

        try:
            rule_fcol = _resolve_filter_col(df, rule_fcol_req)
        except ValueError as e:
            print(f"  [warn] year {rule_year} filter_col unresolved: {e} — using year_default")
            df.loc[ry_mask, "report_period"] = rule_default
            df.loc[ry_mask, "split_source"] = "year_default"
            per_year_stats[rule_year] = {"rule_default": rule_default,
                                         "rows": int(ry_mask.sum()),
                                         "filter_col_error": str(e)}
            continue

        fval = df[rule_fcol].astype(str).str.strip()
        matched_in_year = pd.Series(False, index=df.index)
        bucket_hits = {}
        for bk, expected in rule_buckets.items():
            vals = expected if isinstance(expected, list) else [expected]
            bk_mask = ry_mask & fval.isin([str(v).strip() for v in vals])
            n = int(bk_mask.sum())
            df.loc[bk_mask, "report_period"] = bk
            df.loc[bk_mask, "split_source"] = "direct_filter"
            matched_in_year |= bk_mask
            bucket_hits[bk] = n

        # Year-specific unmatched → year-default
        unmatched_in_year = ry_mask & ~matched_in_year
        df.loc[unmatched_in_year, "report_period"] = rule_default
        df.loc[unmatched_in_year, "split_source"] = "year_default"

        per_year_stats[rule_year] = {
            "rule_default": rule_default,
            "rows": int(ry_mask.sum()),
            "bucket_hits": bucket_hits,
            "fallback_rows": int(unmatched_in_year.sum()),
        }

    audit = {"mode": "direct_filter_multi_year", "per_year_stats": per_year_stats}
    return df, audit


# ─────────────────────────────────────────────────────────────────────────────
# Mode 1b: Per-year mode dispatch — each year_rule chooses its own handler
# ─────────────────────────────────────────────────────────────────────────────
def _do_per_year_mode(df: pd.DataFrame, ys: dict, cfg: dict,
                       interim: Path, company: str):
    """Slice df by report_year and dispatch each slice to its rule.mode handler.

    Used when SJM (or any entity) needs mixed modes across years —
    e.g. 25 uses hybrid_direct_and_ratio (跨年項目 needs ratio split via database
    tab) while 24 uses direct_filter (24 file's `項目所屬年份` is already
    per-row classified, no ratio needed).

    YAML schema example (SJM):
        year_split:
          default_bucket: "25"
          per_year_rules:
            - report_year: 2025
              mode: hybrid_direct_and_ratio
              filter_col: "項目期間"
              buckets: {...}
              ratio_bucket: "跨年項目"
              proportional_distribute: ["nan", ""]
              ratio_source: {...}
            - report_year: 2024
              mode: direct_filter
              filter_col: "項目所屬年份"   # substring resolved by _resolve_filter_col
              buckets: {"24": [...], "24_23SY": [...]}
              default_bucket: "24"

    Each rule's config is lifted into a synthetic `ys` dict and passed to the
    matching handler. The `ratio_source` (if present) is per-rule, so different
    years can read from different workbook tabs.

    Requires raw.parquet to have `report_year` column (added by yearly_sources
    or prebuild_years upstream).
    """
    if "report_year" not in df.columns:
        raise ValueError(
            "per_year_mode requires 'report_year' column. Add yearly_sources or "
            "prebuild_years upstream to inject it."
        )

    rules = ys.get("per_year_rules") or []
    default_bucket = ys.get("default_bucket", "25")
    ry = pd.to_numeric(df["report_year"], errors="coerce")

    # Lifted from rule into synthetic ys for each per-year dispatch.
    LIFT_KEYS = ("filter_col", "buckets", "default_bucket",
                 "ratio_bucket", "proportional_distribute", "ratio_source",
                 "pid_je_col", "name_je_col", "fallback_col",
                 "fallback_mapping", "per_year_override")

    parts = []
    per_year_audits = {}
    handled_idx_sets = []

    for rule in rules:
        rule_year = int(rule["report_year"])
        rule_mode = rule.get("mode")
        if not rule_mode:
            # Skip mode-less rules here — they're handled by _do_direct_filter_multi_year
            # legacy path. (We only reach this fn when at least one rule has a mode.)
            continue

        ry_mask = ry == rule_year
        sub = df[ry_mask].copy()
        if len(sub) == 0:
            print(f"  [per_year_mode] year {rule_year} ({rule_mode}): 0 rows, skipped")
            continue

        synthetic_ys = {k: rule[k] for k in LIFT_KEYS if k in rule}
        synthetic_ys["mode"] = rule_mode
        synthetic_ys.setdefault("default_bucket", default_bucket)

        print(f"  [per_year_mode] year {rule_year} ({rule_mode}): "
              f"{len(sub):,} rows → handler")

        if rule_mode == "direct_filter":
            sub_out, sub_audit = _do_direct_filter(sub, synthetic_ys, cfg)
        elif rule_mode == "hybrid_direct_and_ratio":
            sub_out, sub_audit = _do_hybrid(sub, synthetic_ys, cfg, interim, company)
        elif rule_mode == "ratio_with_fallback":
            sub_out, sub_audit = _do_ratio_fallback(sub, synthetic_ys, cfg, interim, company)
        elif rule_mode == "bucket_amount_cols":
            sub_out, sub_audit = _do_bucket_amount_cols(sub, synthetic_ys, cfg)
        else:
            raise ValueError(f"Unknown per-year mode: {rule_mode!r} for year {rule_year}")

        parts.append(sub_out)
        per_year_audits[rule_year] = sub_audit
        handled_idx_sets.append(set(sub.index))

    # Catch-all for rows whose report_year doesn't match any rule with a mode
    all_handled = set().union(*handled_idx_sets) if handled_idx_sets else set()
    unhandled_mask = ~df.index.isin(all_handled)
    if unhandled_mask.any():
        unh = df[unhandled_mask].copy()
        amt = pd.to_numeric(unh[cfg["columns"]["amount"]], errors="coerce")
        unh["report_period"] = default_bucket
        unh["split_ratio"] = 1.0
        unh["split_source"] = "year_default_unmatched"
        unh["original_amount"] = amt
        parts.append(unh)
        print(f"  [per_year_mode] unhandled rows (no rule match): {len(unh):,} "
              f"→ default_bucket={default_bucket}")

    df_out = pd.concat(parts, ignore_index=True) if parts else df.iloc[0:0].copy()
    audit = {"mode": "per_year_mode", "per_year_audits": per_year_audits}
    return df_out, audit


# ─────────────────────────────────────────────────────────────────────────────
# Mode 2: SJM hybrid (direct + ratio + proportional)
# ─────────────────────────────────────────────────────────────────────────────
def _do_hybrid(df: pd.DataFrame, ys: dict, cfg: dict, interim: Path, company: str):
    fcol_req = ys["filter_col"]
    buckets = ys.get("buckets") or {}
    ratio_bucket_val = ys.get("ratio_bucket")
    proportional_vals = [str(v).strip() for v in (ys.get("proportional_distribute") or [])]

    fcol = _resolve_filter_col(df, fcol_req)

    amt_col = cfg["columns"]["amount"]
    df = df.copy()
    df["original_amount"] = pd.to_numeric(df[amt_col], errors="coerce").fillna(0)
    fval = df[fcol].astype(str).str.strip()

    # Replace literal 'nan' (from pandas str cast on NaN) so it matches proportional vals
    is_nan_actual = df[fcol].isna()
    fval_eff = fval.where(~is_nan_actual, "nan")

    # ── Direct bucket rows ────────────────────────────────────────────────
    direct_rows = []
    for bk, expected in buckets.items():
        vals = expected if isinstance(expected, list) else [expected]
        mask = fval_eff.isin([str(v).strip() for v in vals])
        sub = df[mask].copy()
        sub["report_period"] = bk
        sub["split_ratio"] = 1.0
        sub["split_source"] = "direct_filter"
        direct_rows.append(sub)
    direct_df = pd.concat(direct_rows, ignore_index=True) if direct_rows else df.iloc[0:0].copy()

    # ── Ratio split rows (跨年項目) ───────────────────────────────────────
    ratio_mask = fval_eff.eq(str(ratio_bucket_val).strip()) if ratio_bucket_val else pd.Series(False, index=df.index)
    ratio_in = df[ratio_mask].copy()
    print(f"  [hybrid] direct rows: {len(direct_df):,}, ratio rows: {len(ratio_in):,}, "
          f"proportional rows: {int(is_nan_actual.sum() | fval_eff.isin(proportional_vals).sum())}")

    ratio_out, ratio_audit = _apply_ratio_split_sjm(ratio_in, ys, cfg, interim, company)

    # ── Proportional distribution rows (nan / etc.) ──────────────────────
    prop_mask = fval_eff.isin(proportional_vals) | is_nan_actual
    # exclude rows already accounted for
    accounted_mask = ratio_mask | _direct_mask_union(fval_eff, buckets)
    prop_mask = prop_mask & ~accounted_mask
    prop_in = df[prop_mask].copy()

    # Compute global bucket ratio from direct + ratio split output
    combined_so_far = pd.concat([direct_df, ratio_out], ignore_index=True) if len(ratio_out) else direct_df
    if len(combined_so_far) > 0 and "report_period" in combined_so_far.columns:
        bk_totals = combined_so_far.groupby("report_period")["original_amount"].sum()
        total = float(pd.to_numeric(combined_so_far["original_amount"], errors="coerce").fillna(0).sum())
        global_ratio = {b: (float(v) / total if total > 0 else 0.0) for b, v in bk_totals.items()}
    else:
        global_ratio = {"25": 1.0}
    print(f"  [hybrid] global ratio for proportional: "
          f"{ {k: f'{v:.3f}' for k, v in global_ratio.items()} }")

    prop_out = []
    for _, r in prop_in.iterrows():
        for bk, ratio in global_ratio.items():
            if ratio <= 0:
                continue
            new = r.copy()
            new["report_period"] = bk
            new["split_ratio"] = float(ratio)
            new["split_source"] = "proportional"
            new[amt_col] = float(new["original_amount"]) * ratio
            if "_uid" in new and pd.notna(new["_uid"]):
                new["_uid"] = f"{new['_uid']}__{bk}"
            prop_out.append(new)
    prop_df = pd.DataFrame(prop_out) if prop_out else df.iloc[0:0].copy()

    # Catch-all: rows that don't match any of {direct, ratio, proportional}
    other_mask = ~(accounted_mask | prop_mask)
    other_in = df[other_mask].copy()
    if len(other_in):
        other_in["report_period"] = ys.get("default_bucket", "25")
        other_in["split_ratio"] = 1.0
        other_in["split_source"] = "default_fallback"

    parts = [d for d in [direct_df, ratio_out, prop_df, other_in] if len(d)]
    df_out = pd.concat(parts, ignore_index=True) if parts else df.iloc[0:0].copy()

    audit = {
        "mode": "hybrid_direct_and_ratio",
        "direct_rows": int(len(direct_df)),
        "ratio_rows_in": int(len(ratio_in)),
        "ratio_rows_out": int(len(ratio_out)),
        "proportional_rows_in": int(len(prop_in)),
        "proportional_rows_out": int(len(prop_df)),
        "other_rows": int(len(other_in)),
        "global_ratio_for_proportional": global_ratio,
        "ratio_audit": ratio_audit,
    }
    return df_out, audit


def _direct_mask_union(fval_eff: pd.Series, buckets: dict) -> pd.Series:
    vals_all = []
    for v in buckets.values():
        for vv in (v if isinstance(v, list) else [v]):
            vals_all.append(str(vv).strip())
    return fval_eff.isin(vals_all)


def _apply_ratio_split_sjm(df_in: pd.DataFrame, ys: dict, cfg: dict,
                             interim: Path, company: str):
    """Read SJM database tab, build {db_key: (r25, r24, r23)} dict, split rows by ratio."""
    if len(df_in) == 0:
        return df_in.copy(), {"projects": []}

    rs = ys.get("ratio_source") or {}
    # If `file` not specified, default to raw_filename (same workbook, different tab).
    rs_file = (path(cfg, "raw_xlsx").parent / rs["file"]) if rs.get("file") else path(cfg, "raw_xlsx")
    rs_sheet = rs["sheet"]
    header_row = int(rs.get("header_row", 1))
    je_key_col = rs.get("je_key_col")        # column name in raw.parquet
    db_key_col = rs.get("db_key_col")        # Excel column letter, e.g. 'E'
    ratio_cols = rs.get("ratio_cols") or {}  # {'25': 'EY', ...}

    if je_key_col not in df_in.columns:
        raise ValueError(f"ratio_source.je_key_col {je_key_col!r} not in parquet")

    if not rs_file.exists():
        raise FileNotFoundError(f"ratio_source file not found: {rs_file}")
    wb = openpyxl.load_workbook(rs_file, read_only=True, data_only=True)
    ws = wb[rs_sheet]

    db_lookup: dict[str, dict] = {}
    proj_calc_rows: list[dict] = []
    for ri in range(header_row + 1, ws.max_row + 1):
        key_v = ws[f"{db_key_col}{ri}"].value
        if key_v is None or str(key_v).strip() == "":
            continue
        key = str(key_v).strip()
        ratios = {}
        for bk, letter in ratio_cols.items():
            v = ws[f"{letter}{ri}"].value
            try:
                ratios[bk] = float(v) if v is not None else 0.0
            except (ValueError, TypeError):
                ratios[bk] = 0.0
        db_lookup[key] = ratios
        # Also pull project name (col F by SJM convention) for audit
        proj_name = ws[f"F{ri}"].value if ri <= ws.max_row else ""
        proj_calc_rows.append({
            "db_key": key,
            "project_name": str(proj_name).strip() if proj_name else "",
            **ratios,
        })

    amt_col = cfg["columns"]["amount"]
    df_in = df_in.copy()
    df_in["original_amount"] = pd.to_numeric(df_in[amt_col], errors="coerce").fillna(0)

    out_rows = []
    unmatched_keys: dict[str, int] = {}
    project_split_totals: dict[str, dict] = {}
    edge_count_allzero = 0

    je_keys = df_in[je_key_col].astype(str).str.strip()
    for idx, (key, orig_amt) in enumerate(zip(je_keys, df_in["original_amount"])):
        row = df_in.iloc[idx]
        if key not in db_lookup:
            unmatched_keys[key] = unmatched_keys.get(key, 0) + 1
            # Fallback: full amount to default
            new = row.copy()
            new["report_period"] = ys.get("default_bucket", "25")
            new["split_ratio"] = 1.0
            new["split_source"] = "ratio_unmatched_fallback"
            out_rows.append(new)
            continue
        ratios = db_lookup[key]
        total_r = sum(max(v, 0) for v in ratios.values())
        if total_r <= 0:
            edge_count_allzero += 1
            new = row.copy()
            new["report_period"] = ys.get("default_bucket", "25")
            new["split_ratio"] = 1.0
            new["split_source"] = "edge_case_25"
            out_rows.append(new)
            continue
        # Split into len(ratios) buckets
        proj_audit = project_split_totals.setdefault(
            key, {"db_key": key, "rows": 0, "orig_amount": 0.0,
                  "ratios": ratios, "splits": {b: 0.0 for b in ratios}}
        )
        proj_audit["rows"] += 1
        proj_audit["orig_amount"] += float(orig_amt)
        for bk, r in ratios.items():
            r_eff = max(r, 0)
            if r_eff == 0:
                continue
            frac = r_eff / total_r
            new = row.copy()
            new["report_period"] = bk
            new["split_ratio"] = float(frac)
            new["split_source"] = "ratio_split"
            new[amt_col] = float(orig_amt) * frac
            if "_uid" in new and pd.notna(new["_uid"]):
                new["_uid"] = f"{new['_uid']}__{bk}"
            out_rows.append(new)
            proj_audit["splits"][bk] += float(orig_amt) * frac

    df_out = pd.DataFrame(out_rows)
    audit = {
        "db_rows": len(db_lookup),
        "matched_rows": int(len(df_in) - sum(unmatched_keys.values())),
        "unmatched_rows": int(sum(unmatched_keys.values())),
        "unmatched_keys": unmatched_keys,
        "edge_case_allzero": edge_count_allzero,
        "project_split": list(project_split_totals.values()),
        "db_full_ratios": proj_calc_rows,
    }
    return df_out, audit


# ─────────────────────────────────────────────────────────────────────────────
# Mode 3: MGM ratio with fallback
# ─────────────────────────────────────────────────────────────────────────────
def _do_ratio_fallback(df: pd.DataFrame, ys: dict, cfg: dict,
                        interim: Path, company: str):
    rs = ys.get("ratio_source") or {}
    # If `file` not specified, default to raw_filename (same workbook, different tab).
    rs_file = (path(cfg, "raw_xlsx").parent / rs["file"]) if rs.get("file") else path(cfg, "raw_xlsx")
    rs_sheet = rs["sheet"]
    header_row = int(rs.get("header_row", 6))
    data_start_row = int(rs.get("data_start_row", 7))
    pid_letter = rs.get("pid_col", "B")
    name_letter = rs.get("name_col", "C")
    formula_cols = rs.get("ratio_formula") or {}

    fallback_col = ys.get("fallback_col", "report_year")
    fallback_mapping = ys.get("fallback_mapping") or {}
    fallback_mapping = {str(k): str(v) for k, v in fallback_mapping.items()}
    default_bucket = ys.get("default_bucket", "25")

    # NEW: per_year_override — for non-target reporting years, skip ratio split
    # entirely and route to a direct bucket (e.g. 2024 rows → "24", 2023 → "23"
    # because Master ratio is only for 2025 reporting).
    per_year_override = {str(k): str(v) for k, v in (ys.get("per_year_override") or {}).items()}

    pid_je_col = ys.get("pid_je_col", "project_id")
    name_je_col = ys.get("name_je_col", "project_name")
    amt_col = cfg["columns"]["amount"]

    if not rs_file.exists():
        raise FileNotFoundError(f"ratio_source file not found: {rs_file}")

    wb = openpyxl.load_workbook(rs_file, read_only=True, data_only=True)
    ws = wb[rs_sheet]

    # Load Master rows
    master_records = []
    for ri in range(data_start_row, ws.max_row + 1):
        pid_v = ws[f"{pid_letter}{ri}"].value
        name_v = ws[f"{name_letter}{ri}"].value
        if name_v is None or str(name_v).strip() == "":
            continue
        if pid_v is not None and str(pid_v).strip().lower() == "total":
            continue
        rec = {
            "B": str(pid_v).strip() if pid_v else "",
            "C": str(name_v).strip(),
        }
        # Evaluate ratio formulas. Support simple "X" or "max(X-Y,0)" forms.
        for bk, formula in formula_cols.items():
            rec[bk] = _eval_formula(formula, ws, ri)
        master_records.append(rec)
    print(f"  [mgm] Master loaded: {len(master_records)} projects from {rs_file.name}")

    # Build match indexes
    def _zfill_norm(v: str) -> str:
        s = str(v).strip()
        return s.zfill(3) if s.isdigit() else s

    by_pid_exact = {r["B"]: r for r in master_records if r["B"]}
    by_pid_zfill = {_zfill_norm(r["B"]): r for r in master_records if r["B"]}
    by_name_exact = {r["C"]: r for r in master_records}
    # For name substring: store as list to allow longest-match disambiguation
    master_by_name_substr = [(r["C"], r) for r in master_records]
    master_by_name_substr.sort(key=lambda kv: -len(kv[0]))  # longest first

    # Read manual override file (if exists from prior run)
    review_path = interim / f"{company}_fuzzy_match_review.xlsx"
    overrides: dict[str, str] = {}  # parquet_name → master.B
    if review_path.exists():
        try:
            rev = pd.read_excel(review_path)
            if "parquet_project_name" in rev.columns and "manual_override_B" in rev.columns:
                for _, r in rev.iterrows():
                    pn = str(r["parquet_project_name"]).strip()
                    mb = str(r.get("manual_override_B", "")).strip()
                    if pn and mb and mb.lower() not in ("nan", "skip", ""):
                        overrides[pn] = mb
            print(f"  [mgm] Loaded {len(overrides)} manual_override_B from existing review file")
        except Exception as e:
            print(f"  [mgm] Failed to read review file: {e}")

    # Match each row
    df = df.copy()
    df["original_amount"] = pd.to_numeric(df[amt_col], errors="coerce").fillna(0)

    pids = df[pid_je_col].astype(str).str.strip() if pid_je_col in df.columns else pd.Series([""] * len(df), index=df.index)
    pnames = df[name_je_col].astype(str).str.strip() if name_je_col in df.columns else pd.Series([""] * len(df), index=df.index)

    fuzzy_log: dict[str, dict] = {}  # pname → {method, master_B, ...}

    def _match_row(pid: str, pname: str):
        # 1. manual override
        if pname in overrides:
            mb = overrides[pname]
            r = by_pid_exact.get(mb) or by_pid_zfill.get(_zfill_norm(mb))
            if r:
                return r, "manual_override"
        # 2. pid exact
        if pid and pid != "None" and pid in by_pid_exact:
            return by_pid_exact[pid], "pid_exact"
        # 3. pid zfill
        if pid and pid != "None":
            r = by_pid_zfill.get(_zfill_norm(pid))
            if r:
                return r, "pid_zfill"
        # 4. name exact
        if pname and pname in by_name_exact:
            return by_name_exact[pname], "name_exact"
        # 5. name substring (Master.C is contained in parquet name)
        if pname:
            for mc, r in master_by_name_substr:
                if mc and mc in pname:
                    return r, "name_substring"
        return None, "unmatched"

    out_rows = []
    edge_negs = 0
    edge_zero = 0
    method_counts: dict[str, int] = {}
    for idx in range(len(df)):
        row = df.iloc[idx]
        pid = pids.iloc[idx]
        pname = pnames.iloc[idx]
        orig_amt = float(row["original_amount"])

        # NEW: per_year_override fires BEFORE ratio split logic.
        # E.g. for MGM, 2024/2023 rows skip Master ratio (which is 25-only) and
        # go direct to bucket "24" / "23".
        if per_year_override:
            yr_raw = row.get(fallback_col)
            yr_key = str(yr_raw).strip().split(".")[0] if yr_raw is not None else ""
            if yr_key in per_year_override:
                new = row.copy()
                new["report_period"] = per_year_override[yr_key]
                new["split_ratio"] = 1.0
                new["split_source"] = "per_year_override"
                out_rows.append(new)
                method_counts["per_year_override"] = method_counts.get("per_year_override", 0) + 1
                continue

        match, method = _match_row(pid, pname)
        method_counts[method] = method_counts.get(method, 0) + 1

        # Track fuzzy candidates for review file
        if pname and pname not in fuzzy_log:
            fuzzy_log[pname] = {"parquet_project_name": pname,
                                  "match_method": method,
                                  "matched_master_B": match["B"] if match else "",
                                  "matched_master_C": match["C"] if match else "",
                                  "rows": 0,
                                  "amount": 0.0}
        if pname:
            fuzzy_log[pname]["rows"] += 1
            fuzzy_log[pname]["amount"] += orig_amt

        if match is None:
            # Fallback to report_year
            new = row.copy()
            yr = row.get(fallback_col)
            bk = fallback_mapping.get(str(yr).strip().split(".")[0], default_bucket)
            new["report_period"] = bk
            new["split_ratio"] = 1.0
            new["split_source"] = "report_year_fallback"
            out_rows.append(new)
            continue

        # Apply ratio
        raw_ratios = {bk: match[bk] for bk in formula_cols}
        ratios = {bk: max(v, 0.0) for bk, v in raw_ratios.items()}
        denom = sum(ratios.values())
        if denom <= 0:
            # All ratio values are zero after clamping.  Two sub-cases:
            #   (a) all_negative: raw ratios were all < 0 (e.g. J < S for every period) →
            #       mark split_source so audit can distinguish from genuine zero-plan rows.
            #   (b) genuine zero: Master shows zero plan across all periods.
            # In both cases fall back to report_year bucket, but surface the sub-case.
            edge_zero += 1
            all_negative = all(v < 0 for v in raw_ratios.values())
            new = row.copy()
            yr = row.get(fallback_col)
            bk = fallback_mapping.get(str(yr).strip().split(".")[0], default_bucket)
            new["report_period"] = bk
            new["split_ratio"] = 1.0
            new["split_source"] = "edge_neg_fallback" if all_negative else "edge_zero_fallback"
            out_rows.append(new)
            continue

        # Negative J detection: track for audit (J already capped by formula)
        # Split
        for bk, r_v in ratios.items():
            if r_v <= 0:
                continue
            frac = r_v / denom
            new = row.copy()
            new["report_period"] = bk
            new["split_ratio"] = float(frac)
            new["split_source"] = f"ratio_{method}"
            new[amt_col] = orig_amt * frac
            if "_uid" in new and pd.notna(new["_uid"]):
                new["_uid"] = f"{new['_uid']}__{bk}"
            out_rows.append(new)

    df_out = pd.DataFrame(out_rows)

    # Write fuzzy review xlsx (preserve manual_override_B if exists)
    fuzzy_df = pd.DataFrame(list(fuzzy_log.values()))
    if len(fuzzy_df):
        fuzzy_df = fuzzy_df.sort_values(["match_method", "amount"], ascending=[True, False])
        fuzzy_df["manual_override_B"] = ""
        # Apply existing overrides back
        for pn, mb in overrides.items():
            fuzzy_df.loc[fuzzy_df["parquet_project_name"] == pn, "manual_override_B"] = mb
        fuzzy_df["manual_notes"] = ""
        fuzzy_df.to_excel(review_path, index=False)
        print(f"  [mgm] Fuzzy review written: {review_path.name}  ({len(fuzzy_df)} unique names)")

    audit = {
        "mode": "ratio_with_fallback",
        "match_method_counts": method_counts,
        "edge_zero": edge_zero,
        "overrides_applied": len(overrides),
        "fuzzy_review_path": str(review_path),
    }
    return df_out, audit


def _eval_formula(formula: str, ws, ri: int) -> float:
    """Evaluate simple Excel-cell formula strings.

    Supported forms:
      "I"               → ws.I{ri}.value
      "max(J-S, 0)"     → max(ws.J{ri} - ws.S{ri}, 0)
    """
    f = formula.strip()
    # Case 1: single letter
    if f.isalpha() and len(f) <= 2:
        v = ws[f"{f}{ri}"].value
        try:
            return float(v) if v is not None else 0.0
        except (ValueError, TypeError):
            return 0.0
    # Case 2: max(X-Y, 0) pattern
    if f.startswith("max(") and f.endswith(")"):
        inner = f[4:-1]
        # Split last comma to handle the "0" bound
        parts = inner.rsplit(",", 1)
        if len(parts) == 2:
            expr, bound = parts[0].strip(), parts[1].strip()
            try:
                bound_v = float(bound)
            except ValueError:
                bound_v = 0.0
            # Support "X-Y"
            if "-" in expr:
                a, b = expr.split("-", 1)
                a_v = ws[f"{a.strip()}{ri}"].value or 0
                b_v = ws[f"{b.strip()}{ri}"].value or 0
                try:
                    return max(float(a_v) - float(b_v), bound_v)
                except (ValueError, TypeError):
                    return bound_v
    raise ValueError(f"Unsupported ratio formula: {formula!r}")


# ─────────────────────────────────────────────────────────────────────────────
# Audit xlsx writer
# ─────────────────────────────────────────────────────────────────────────────
def _write_summary_audit(df_out: pd.DataFrame, interim: Path, company: str, audit: dict):
    out_path = interim / f"{company}_split_audit.xlsx"
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        # Sheet 1: summary
        if "report_period" in df_out.columns:
            summary = df_out.groupby(["report_period", "split_source"]).agg(
                rows=("split_ratio", "size"),
                amount=("original_amount", lambda s: float(pd.to_numeric(s, errors="coerce").fillna(0).sum())),
            ).reset_index()
            summary.to_excel(writer, sheet_name="summary", index=False)
        else:
            pd.DataFrame([{"note": "no report_period column"}]).to_excel(writer, sheet_name="summary", index=False)

        # Sheet 2: project_calc (only for ratio split)
        ratio_audit = audit.get("ratio_audit") if audit else None
        proj_splits = ratio_audit.get("project_split") if ratio_audit else audit.get("project_split")
        if proj_splits:
            rows = []
            for p in proj_splits:
                row = {
                    "db_key": p["db_key"],
                    "rows_je": p["rows"],
                    "orig_amount": p["orig_amount"],
                }
                for bk, r in p["ratios"].items():
                    row[f"ratio_{bk}"] = r
                for bk, a in p["splits"].items():
                    row[f"split_amount_{bk}"] = a
                rows.append(row)
            pd.DataFrame(rows).to_excel(writer, sheet_name="project_calc", index=False)

        # Sheet 3: edge cases (raw audit dump)
        edge_rows = [{"key": k, "value": str(v)[:200]} for k, v in (audit or {}).items()]
        pd.DataFrame(edge_rows).to_excel(writer, sheet_name="audit_meta", index=False)

    print(f"  Audit xlsx → {out_path.name}")


if __name__ == "__main__":
    main()
